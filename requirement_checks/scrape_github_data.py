import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils.github_code_documentation_quality.environment_instructions_existance_check import check_setup_with_nlp
from utils.common.find_readme import get_readme_info
from transformers import pipeline
import time
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")

def log(msg):
    """Print timestamped log message"""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def parse_github_url(url):
    parts = url.rstrip("/").split("/")
    return parts[-2], parts[-1]

def get_repo_info(owner, repo, headers=None):
    log(f"📦 Fetching repo info: {owner}/{repo}")
    url = f"https://api.github.com/repos/{owner}/{repo}"
    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        try:
            error_msg = response.json().get("message", "Unknown error")
        except:
            error_msg = response.text[:100]
        log(f"❌ Failed [{response.status_code}]: {error_msg}")
        return None

    # Check for README
    log(f"   📄 Checking for README...")
    has_readme = get_readme_info(owner, repo, headers)
    log(f"   ✓ README check done: {has_readme}")

    data = response.json()
    
    log(f"   🔍 Running NLP setup detection (this may take time)...")
    start_nlp = time.time()
    setup_result = check_setup_with_nlp(owner, repo, headers)
    nlp_time = time.time() - start_nlp
    log(f"   ✓ NLP detection done ({nlp_time:.1f}s): {setup_result}")
    
    return {
        "name": data["full_name"],
        "stars": data["stargazers_count"],
        "forks": data["forks_count"],
        "license": data["license"]["name"] if data["license"] else "N/A",
        "open_issues": data["open_issues_count"],
        "language": data["language"] or "N/A",
        "pushed_at": data["pushed_at"][:10] if data["pushed_at"] else "N/A",
        "updated_at": data["updated_at"][:10] if data["updated_at"] else "N/A",
        "html_url": data["html_url"]or "N/A",
        "has_readme": has_readme,
        "has_setup": setup_result,
    }

def save_to_excel(results, filename="github_stats.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GitHub Stats"

    # Styles
    LIGHT_BLUE_FILL = PatternFill("solid", fgColor="BDD7EE")
    ALT_ROW_FILL    = PatternFill("solid", fgColor="EBF3FB")
    WHITE_FILL      = PatternFill("solid", fgColor="FFFFFF")
    HEADER_FONT     = Font(name="Arial", bold=True, size=11, color="1F4E79")
    DATA_FONT       = Font(name="Arial", size=10)
    THIN_BORDER     = Border(
        left=Side(style="thin", color="B8CCE4"),
        right=Side(style="thin", color="B8CCE4"),
        top=Side(style="thin", color="B8CCE4"),
        bottom=Side(style="thin", color="B8CCE4"),
    )

    headers    = ["Repository",  "Stars", "Forks", "License", "Open Issues", "Language", "Last Push", "Last Update", "URL", "Has README", "Setup Instructions"]
    col_widths = [32,             12,      12,      42,        14,            14,         20,          20,            60,    10,           60]

    # Header row
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = LIGHT_BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, repo in enumerate(results, start=2):
        row_fill = ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL
        values = [
            repo["name"],
            repo["stars"],
            repo["forks"],
            repo["license"],
            repo["open_issues"],
            repo["language"],
            repo["pushed_at"],
            repo["updated_at"],
            repo["html_url"],
            repo["has_readme"],
            repo["has_setup"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center",
                vertical="center"
            )
            if col_idx in (2, 3, 5):  # Stars, Forks, Open Issues
                cell.number_format = "#,##0"
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"Saved to {filename}")

log("\n" + "="*80)
log("🚀 Starting GitHub Data Scraper")
log("="*80)

log("⏳ Loading NLP model (first time only, may take 1-2 minutes)...")
start_model = time.time()
classifier = pipeline("zero-shot-classification",
                      model="roberta-large-mnli")
model_time = time.time() - start_model
log(f"✓ Model loaded in {model_time:.1f}s\n")

# --- Your list of GitHub URLs ---
repo_urls = [
    "https://github.com/Interactions-HSG/GEAR",
    "https://github.com/pallets/flask",
    "https://github.com/tensorflow/tensorflow",
    "https://github.com/minimaxir/big-list-of-naughty-strings",
    "https://github.com/nvbn/thefuck",
    "https://github.com/kennethreitz/records",
    "https://github.com/jakubroztocil/httpie",
    "https://github.com/nickcoutsos/keyswitch-layout-selector",
    "https://github.com/apache/airflow",
    "https://github.com/Textualize/rich"
]

# 🔐 Optional: add GitHub token (recommended to avoid rate limits)
headers = {"Authorization": f"Bearer {GITHUB_TOKEN}"} if GITHUB_TOKEN else None

log(f"Processing {len(repo_urls)} repositories...\n")
results = []
start_total = time.time()

for idx, url in enumerate(repo_urls, 1):
    log(f"[{idx}/{len(repo_urls)}] {url}")
    owner, repo = parse_github_url(url)
    info = get_repo_info(owner, repo, headers)
    if info:
        results.append(info)
    log("")

total_time = time.time() - start_total
log("="*80)
log(f"✓ All repositories processed in {total_time:.1f}s")
log("="*80)

if results:
    log(f"\n💾 Saving {len(results)} repositories to Excel...")
    save_to_excel(results, "github_stats.xlsx")
    log("✓ Done!")