"""
Question 5.1.3 — Pre-processing & Pipeline Code classifier (repo-only)

For each paper in data/papers_from_database.py, inspect the linked GitHub
repository and decide whether it contains real preprocessing / pipeline
source code in ANY modality (text, image, audio, video, multimodal):
tokenization / chunking / patching, filtering / cleaning / dedup, or
prompt / input construction. Paper text is not consulted — only the
repo content.

Output: results/preprocessing_code_results.xlsx — one Results, Detail,
and Summary sheet per LLM model in AZURE_OPENAI_DEPLOYMENT, plus a Model
Comparison sheet when more than one model is configured.

Pipeline:
  1. Parse the GitHub URL from each paper (github.com or github.io).
     Non-GitHub repos → NOT_APPLICABLE.                                 → check_paper()
  2. List every file in the repo via the GitHub API.                    → collect_repo_content()
  3. Bucket source files by priority:
       - NAMED    files matching preprocessing patterns (preprocess.py, …)
       - FOLDER   files under preprocessing folders (data/, pipeline/, …)
       - GENERIC  other source files (fallback context)
  4. Fetch each bucket under its OWN character budget so NAMED files
     (highest signal) cannot be crowded out by 70 unrelated scripts.    → collect_repo_content()
  5. Send the combined text to the LLM and parse the verdict — including
     per-file (category, short evidence) pairs.                         → make_check_paper_fn()
  6. Cheap remediation (only when triggered): if the verdict is
     NOT_APPLICABLE but the NAMED bucket had files, re-run the LLM on
     NAMED-only content. If that flips to EXISTS, prefer it. Same token
     budget as the first pass.                                          → make_check_paper_fn()
  7. Repeat across every configured model; write the Excel output.      → main(), save_results()

Labels:
  EXISTS         — repo contains real preprocessing / pipeline code
  NOT_APPLICABLE — no code found, non-GitHub URL, empty repo, or LLM error
"""

import os
import sys
from functools import lru_cache
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()

_this = Path(__file__).resolve()
# requirement_checks/ is three levels up — needed for common/, data/, openai_client
sys.path.insert(0, str(_this.parents[2]))

from common.github_helpers import (
    parse_github_repo, list_all_repo_files, fetch_file_content,
    fetch_paths_with_char_budget,
)
from common.llm_helpers import llm_call_parse_retry
from common.checker_pipeline import run_pipeline
from common.excel_output import (
    write_header_row, write_results_data_rows, write_summary_sheet,
    write_comparison_sheet, safe_sheet_name,
    thin_border, auto_row_height, alignment_wrap_left,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from openai_client import client, AZURE_OPENAI_DEPLOYMENTS
from data.papers_from_database import PAPERS
from prompts import SYSTEM_PROMPT, USER_PROMPT
from config import (
    PREPROC_NAME_RE, PREPROC_FOLDER_PREFIXES, SOURCE_EXTS,
    SKIP_FOLDER_PREFIXES,
    BUCKET_BUDGET_NAMED, BUCKET_BUDGET_FOLDER, BUCKET_BUDGET_GENERIC,
    PER_FILE_CHAR_CAP_NAMED, PER_FILE_CHAR_CAP,
    MAX_NAMED_FILES, MAX_FOLDER_FILES, MAX_GENERIC_FILES,
)

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")

# Display labels for the three categories — keys must match the lowercase
# values the LLM emits in its preprocessing_files[*].category field.
CATEGORY_DISPLAY = {
    "tokenization": "Tokenization",
    "filtering":    "Filtering",
    "prompting":    "Prompting wrappers",
}

# Per-bucket fetch parameters: (budget_chars, per_file_cap, header_label).
# The budget values are STARTING ceilings — collect_repo_content() pools
# the total across buckets in priority order (NAMED -> FOLDER -> GENERIC),
# so unused budget from earlier buckets cascades to later ones.
BUCKET_PARAMS = {
    "named":   (BUCKET_BUDGET_NAMED,   PER_FILE_CHAR_CAP_NAMED, "PREPROCESS-NAMED"),
    "folder":  (BUCKET_BUDGET_FOLDER,  PER_FILE_CHAR_CAP,       "PREPROCESS-FOLDER"),
    "generic": (BUCKET_BUDGET_GENERIC, PER_FILE_CHAR_CAP,       "GENERIC SOURCE"),
}


# ---------------------------------------------------------------------------
# Status -> classification label
# ---------------------------------------------------------------------------

def _status_to_label(status):
    """Map the per-paper status ('yes' vs anything else) to a human label."""
    return "EXISTS" if status == "yes" else "NOT_APPLICABLE"


# ---------------------------------------------------------------------------
# Repo content collection
# ---------------------------------------------------------------------------

def _is_skippable(path_lower):
    return any(path_lower.startswith(p) for p in SKIP_FOLDER_PREFIXES)


def _is_source(path_lower):
    return os.path.splitext(path_lower)[1] in SOURCE_EXTS


def _bucket_paths(all_files):
    """Sort repo paths into (named, folder, generic) lists, dropping skip/non-source."""
    named, folder, generic = [], [], []
    for f in all_files:
        fpath     = f["path"] if isinstance(f, dict) else f
        fpath_low = fpath.lower()
        basename  = fpath_low.split("/")[-1]

        if _is_skippable(fpath_low) or not _is_source(fpath_low):
            continue

        if PREPROC_NAME_RE.search(basename):
            named.append(fpath)
        elif any(fpath_low.startswith(p) for p in PREPROC_FOLDER_PREFIXES):
            folder.append(fpath)
        else:
            generic.append(fpath)
    return named, folder, generic


def _fetch_bucket(owner, repo, paths, budget, per_file_cap, header_label):
    """Fetch one bucket's files under a char budget. Returns (text, fetched_paths, chars_used)."""
    if not paths or budget <= 0:
        return "", [], 0
    blocks, got, chars_used = fetch_paths_with_char_budget(
        owner, repo, paths, budget,
        fetch_content=fetch_file_content,
        start_total_chars=0,
        pause_seconds=0.15,
        header_label=header_label,
        per_file_char_cap=per_file_cap,
    )
    return "\n\n".join(blocks), got, chars_used


def collect_repo_content(owner, repo):
    """Fetch preprocessing-relevant files from a GitHub repo.

    Files are sorted into NAMED / FOLDER / GENERIC priority buckets, then
    fetched against a SHARED char budget that cascades in priority order:
    whatever NAMED does not consume flows to FOLDER, whatever FOLDER does
    not consume flows to GENERIC. This stops the GENERIC bucket from
    starving when real preprocessing code lives under a "neutrally-named"
    file like `msformer/utils.py` (the MacFrag fragmentation algorithm).

    Returns a dict:
      {
        "combined_text":      str,    # all buckets joined
        "fetched_paths":      list,   # paths actually retrieved
        "total_source_files": int,    # source files BEFORE bucket truncation
        "buckets": {
          "named":   {"text": str, "files": [...], "candidates": int},
          "folder":  {"text": str, "files": [...], "candidates": int},
          "generic": {"text": str, "files": [...], "candidates": int},
        },
      }
    """
    print(f"  [GitHub] Listing {owner}/{repo}")
    try:
        all_files = list_all_repo_files(owner, repo)
    except Exception as e:
        print(f"  [GitHub] ERROR during listing: {type(e).__name__}: {e}")
        return _empty_repo_content()

    if not all_files:
        print(f"  [GitHub] Repo empty or unreachable")
        return _empty_repo_content()

    print(f"  [GitHub] {len(all_files)} files in repo")

    named, folder, generic = _bucket_paths(all_files)
    total_source_files = len(named) + len(folder) + len(generic)

    # Sort by depth then name. Shallower paths first — top-level scripts
    # are usually the canonical entry points.
    for lst in (named, folder, generic):
        lst.sort(key=lambda p: (p.count("/"), p.lower()))

    # Cap each bucket's candidate count BEFORE fetching, then let the
    # shared char budget further trim if needed.
    candidates = {
        "named":   named[:MAX_NAMED_FILES],
        "folder":  folder[:MAX_FOLDER_FILES],
        "generic": generic[:MAX_GENERIC_FILES],
    }
    print(f"  [GitHub] Selected: {len(candidates['named'])} named, "
          f"{len(candidates['folder'])} in folders, {len(candidates['generic'])} generic "
          f"(of {total_source_files} source files)")

    # Pool the char budget across buckets in priority order.
    total_budget = sum(b[0] for b in BUCKET_PARAMS.values())
    buckets = {}
    remaining = total_budget
    for name in ("named", "folder", "generic"):
        paths = candidates[name]
        _, per_file_cap, header_label = BUCKET_PARAMS[name]
        text, files, used = _fetch_bucket(
            owner, repo, paths, remaining, per_file_cap, header_label,
        )
        remaining -= used
        buckets[name] = {
            "text":       text,
            "files":      files,
            "candidates": len(paths),
        }

    combined_text = "\n\n".join(b["text"] for b in buckets.values() if b["text"])
    fetched_paths = [p for b in buckets.values() for p in b["files"]]

    return {
        "combined_text":      combined_text,
        "fetched_paths":      fetched_paths,
        "total_source_files": total_source_files,
        "buckets":            buckets,
    }


def _empty_repo_content():
    return {
        "combined_text":      "",
        "fetched_paths":      [],
        "total_source_files": 0,
        "buckets": {
            "named":   {"text": "", "files": [], "candidates": 0},
            "folder":  {"text": "", "files": [], "candidates": 0},
            "generic": {"text": "", "files": [], "candidates": 0},
        },
    }


@lru_cache(maxsize=None)
def _cached_collect_repo_content(owner, repo):
    """Cached collect_repo_content — avoids re-fetching when running multiple models."""
    return collect_repo_content(owner, repo)


# ---------------------------------------------------------------------------
# LLM result handling
# ---------------------------------------------------------------------------

EMPTY_PREPROCESSING_PAYLOAD = {
    "classification":      None,
    "preprocessing_files": [],
    "diagnostic_notes":    "Empty LLM response after retry — content may exceed context window",
    "final_reason":        "",
}


def _normalise_preprocessing_files(raw):
    """Coerce the LLM's preprocessing_files into a list of clean dicts.

    Tolerates legacy string entries (path only) and trims/lowercases the
    fields we display. Drops entries without a path.
    """
    normalised = []
    for entry in raw or []:
        if isinstance(entry, dict):
            path     = (entry.get("path")     or "").strip()
            category = (entry.get("category") or "").strip().lower()
            evidence = (entry.get("evidence") or "").strip()
        elif isinstance(entry, str):
            path, category, evidence = entry.strip(), "", ""
        else:
            continue
        if path:
            normalised.append({"path": path, "category": category, "evidence": evidence})
    return normalised


def _derive_categories(pf_list):
    """Distinct sorted lowercase categories from per-file entries."""
    cats = {pf.get("category", "") for pf in pf_list}
    cats.discard("")
    return sorted(cats)


def _derive_top_evidence(pf_list, n=3):
    """Concatenate up to n evidence snippets from per-file entries."""
    snippets = [pf["evidence"] for pf in pf_list if pf.get("evidence")]
    return "; ".join(snippets[:n])


def _coverage_pct(files_checked, total_source_files):
    """Return a coverage string like '23%', or '' when total is 0/unknown."""
    if not total_source_files:
        return ""
    return f"{100.0 * len(files_checked) / total_source_files:.0f}%"


# ---------------------------------------------------------------------------
# Per-paper orchestration
# ---------------------------------------------------------------------------

def make_check_paper_fn(deployment, token_usage):
    """Factory: return a check_paper function bound to a specific LLM deployment.

    The factory shares one repo-content cache across all model instances for
    the same repo, so multi-model runs do not re-fetch GitHub content.
    """

    def _llm_check(repo_content, paper_title, repo_url):
        """One LLM call. Returns the parsed JSON payload (dict).

        Output-token budget scales with input size (2500 for small inputs,
        4000 for >50k chars of input).
        """
        def build_msg(content):
            return USER_PROMPT.format(
                title=paper_title,
                repo_url=repo_url,
                repo_content=content if content else "[No source code files found in the repository]",
            )

        max_tokens = 4000 if len(repo_content) > 50_000 else 2500

        return llm_call_parse_retry(
            client=client,
            deployment=deployment,
            system_prompt=SYSTEM_PROMPT,
            build_user_message=build_msg,
            content=repo_content,
            token_usage=token_usage,
            empty_payload=EMPTY_PREPROCESSING_PAYLOAD,
            required_list_fields=("preprocessing_files",),
            max_completion_tokens=max_tokens,
            retry_truncate_chars=40_000,
            retry_max_tokens=max(max_tokens, 4000),
            preview_chars=500,
        )

    def _classify_with_remediation(rc, paper_title, repo_url):
        """Two-pass classification.

        Layer 1 — first pass: normal LLM call on combined-bucket content.

        Layer 2 — NAMED-only retry (cheap, no extra tokens):
          If the verdict is NOT_APPLICABLE but NAMED files exist, the
          high-signal files may have been diluted by FOLDER/GENERIC noise.
          Re-prompt the LLM with NAMED-bucket text only. If that flips to
          EXISTS, prefer it.
        """
        # ─── Layer 1 ────────────────────────────────────────────────────────
        current = _llm_check(rc["combined_text"], paper_title, repo_url)
        clf     = current.get("classification")

        # ─── Layer 2 ────────────────────────────────────────────────────────
        named = rc["buckets"]["named"]
        if clf == "NOT_APPLICABLE" and named["candidates"] > 0 and named["text"]:
            print(f"  [Retry] First pass = NOT_APPLICABLE but {named['candidates']} "
                  f"NAMED files present. Re-checking with NAMED-only content...")
            retry = _llm_check(named["text"], paper_title, repo_url)
            if retry.get("classification") == "EXISTS":
                print(f"  [Retry] Second pass = EXISTS — using corrected verdict.")
                return retry
            print(f"  [Retry] Second pass = {retry.get('classification')!r} — keeping first verdict.")

        return current

    def check_paper(paper):
        title = paper.get("title", "")
        url   = paper.get("repo", "") or ""

        result = {
            "title":               title,
            "semanticscholarid":   paper.get("semanticscholarid", ""),
            "repo":                url,
            "status":              None,
            "classification":      "NOT_APPLICABLE",
            "evidence":            "",
            "matched_categories":  [],
            "preprocessing_files": [],   # list of {path, category, evidence} dicts
            "files_checked":       [],
            "total_source_files":  0,
            "coverage":            "",
            "note":                "",
        }

        owner, repo_name = parse_github_repo(url)
        if not owner:
            result["status"] = "skipped"
            result["note"] = (
                f"Not a GitHub repo — manual review needed. URL: {url}"
                if url else "No repo URL provided"
            )
            return result

        try:
            rc = _cached_collect_repo_content(owner, repo_name)
            result["files_checked"]      = rc["fetched_paths"]
            result["total_source_files"] = rc["total_source_files"]
            result["coverage"]           = _coverage_pct(
                rc["fetched_paths"], rc["total_source_files"]
            )

            if not rc["fetched_paths"]:
                result["status"] = "no"
                result["note"] = "No source files retrievable from the repo"
                return result

            llm_result = _classify_with_remediation(rc, title, url)
            clf = llm_result.get("classification")
            pf  = _normalise_preprocessing_files(llm_result.get("preprocessing_files"))

            result["classification"]      = clf or "NOT_APPLICABLE"
            result["preprocessing_files"] = pf
            result["matched_categories"]  = _derive_categories(pf)
            result["evidence"]            = (
                _derive_top_evidence(pf)
                or llm_result.get("diagnostic_notes", "")
            )
            result["note"]                = llm_result.get("final_reason", "")

            if clf is None:
                result["status"] = "error"
                result["note"]   = "Empty LLM response — content may be too large for this model"
            elif clf == "EXISTS":
                result["status"] = "yes"
            else:
                result["status"] = "no"

        except RuntimeError as e:
            result["status"] = "error"
            result["note"]   = str(e)
        except Exception as e:
            result["status"] = "error"
            result["note"]   = f"Unexpected error: {e}"

        return result

    return check_paper


# ---------------------------------------------------------------------------
# Console reporting
# ---------------------------------------------------------------------------

def print_results(results):
    """Print a compact console table of per-paper results."""
    W = 110
    print("\n" + "=" * W)
    print(f"{'#':<4} {'STATUS':<10} {'CATEGORIES':<36} TITLE")
    print("=" * W)

    for i, r in enumerate(results, 1):
        # NOT_APPL (truncated) keeps the status column narrow on the console.
        icon = "EXISTS" if r["status"] == "yes" else "NOT_APPL"

        cats_str = ", ".join(r.get("matched_categories", []))[:34]
        if not cats_str and r.get("note"):
            cats_str = r["note"][:34]

        title_short = r["title"][:46] + ("..." if len(r["title"]) > 46 else "")
        print(f"{i:<4} {icon:<10} {cats_str:<36} {title_short}")

        if r.get("evidence"):
            print(f"       evidence: {r['evidence'][:120]}")
        if r.get("files_checked"):
            extras = len(r["files_checked"]) - 5
            print(f"       files: {', '.join(r['files_checked'][:5])}"
                  + (f" (+{extras} more)" if extras > 0 else ""))

    print("=" * W)
    exists_count = sum(1 for r in results if r["status"] == "yes")
    na_count     = len(results) - exists_count
    print(
        f"\nSUMMARY: {exists_count} EXISTS | "
        f"{na_count} NOT_APPLICABLE | {len(results)} total\n"
    )


# ---------------------------------------------------------------------------
# Excel: Detail sheet builder (extracted)
# ---------------------------------------------------------------------------

def _write_detail_sheet(ws, results, border, wrap):
    """Write the per-file detail sheet for a single model's results.

    Columns: Paper #, Title, Repo, Preprocessing File, Category, Evidence,
    GitHub Link. Paper#/Title/Repo cells are vertically merged when a paper
    has multiple preprocessing files.
    """
    arial10   = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")

    hdrs   = ["Paper #", "Paper Title", "Repo",
              "Preprocessing File", "Category", "Evidence", "GitHub Link"]
    widths = [8, 45, 35, 45, 18, 60, 60]
    write_header_row(ws, hdrs, widths, fill_hex="2F4F6F", border=border)

    merge_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    merge_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    row = 2
    for paper_num, r in enumerate(results, 1):
        owner, repo_name = parse_github_repo(r.get("repo", ""))
        files = [
            pf for pf in r.get("preprocessing_files", [])
            if isinstance(pf, dict) and pf.get("path")
        ]
        if not files:
            continue

        start_row = row
        for pf in files:
            path     = pf["path"]
            category = CATEGORY_DISPLAY.get(
                pf.get("category", ""),
                pf.get("category", "").capitalize(),
            )
            evidence = pf.get("evidence", "")
            link     = (
                f"https://github.com/{owner}/{repo_name}/blob/HEAD/{path}"
                if owner else ""
            )

            for col_idx, value in enumerate([path, category, evidence, link], 4):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font      = arial10
                cell.border    = border
                cell.alignment = wrap
                if col_idx == 7 and value:
                    cell.hyperlink = value
                    cell.font = link_font
            ws.row_dimensions[row].height = auto_row_height(
                [path, category, evidence, link]
            )
            row += 1

        end_row = row - 1
        ws.cell(row=start_row, column=1, value=paper_num).font            = arial10
        ws.cell(row=start_row, column=2, value=r.get("title") or "").font = arial10
        ws.cell(row=start_row, column=3, value=r.get("repo") or "").font  = arial10
        for col_idx in (1, 2, 3):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.border    = border
            cell.alignment = merge_center if col_idx == 1 else merge_left
        if end_row > start_row:
            for col_idx in (1, 2, 3):
                ws.merge_cells(start_row=start_row, start_column=col_idx,
                               end_row=end_row,   end_column=col_idx)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_results(all_model_results, path=None):
    """Write per-model Results + Detail + Summary sheets, and a Model Comparison sheet."""
    if path is None:
        path = Path(__file__).resolve().parent / "results" / "preprocessing_code_results.xlsx"

    wb     = Workbook()
    border = thin_border()
    wrap   = alignment_wrap_left()

    deployments = list(all_model_results.keys())
    first_sheet = True

    for deployment in deployments:
        model_data  = all_model_results[deployment]
        results     = model_data["results"]
        token_usage = model_data["token_usage"]

        # ── Results sheet (per model) ────────────────────────────────────
        if first_sheet:
            ws1 = wb.active
            first_sheet = False
        else:
            ws1 = wb.create_sheet()
        ws1.title = safe_sheet_name("Results", deployment)

        hdrs1   = ["#", "Classification", "Title", "Repo",
                   "Categories", "Preprocessing Files", "Evidence",
                   "Files Checked", "Note", "Coverage %", "Tokens Used"]
        widths1 = [5, 14, 45, 40, 30, 40, 55, 50, 35, 12, 12]
        write_header_row(ws1, hdrs1, widths1, fill_hex="2F4F6F", border=border)

        def results_row_data(r, num):
            preproc_paths = ", ".join(
                pf.get("path", "")
                for pf in (r.get("preprocessing_files") or [])
                if isinstance(pf, dict) and pf.get("path")
            )
            return [
                num,
                _status_to_label(r.get("status")),
                r.get("title") or "",
                r.get("repo") or "",
                ", ".join(r.get("matched_categories") or []),
                preproc_paths,
                r.get("evidence") or "",
                ", ".join(r.get("files_checked") or []),
                r.get("note") or "",
                r.get("coverage") or "",
                r.get("tokens_used") or "",
            ]

        write_results_data_rows(
            ws1, results, results_row_data,
            border=border,
            row_height_fn=lambda vals: auto_row_height(vals),
        )

        # ── Per-File Detail sheet (per model) ────────────────────────────
        ws2 = wb.create_sheet(safe_sheet_name("Detail", deployment))
        _write_detail_sheet(ws2, results, border, wrap)

        # ── Summary sheet (per model) ────────────────────────────────────
        ws3 = wb.create_sheet(safe_sheet_name("Summary", deployment))
        write_summary_sheet(
            ws3, results,
            positive_label="EXISTS",
            negative_label="NOT_APPLICABLE",
            token_usage=token_usage,
            deployment=deployment,
            fill_hex="2F4F6F",
            border=border,
        )

    # ── Model Comparison sheet ───────────────────────────────────────────
    if len(deployments) > 1:
        ws_cmp = wb.create_sheet("Model Comparison")
        write_comparison_sheet(
            ws_cmp, all_model_results, fill_hex="2F4F6F", border=border,
        )

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Full results saved to {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    """Entry point: run the preprocessing-code analysis pipeline."""
    _cached_collect_repo_content.cache_clear()
    run_pipeline(
        papers=PAPERS,
        deployments=AZURE_OPENAI_DEPLOYMENTS,
        make_check_paper_fn=make_check_paper_fn,
        print_results_fn=print_results,
        save_results_fn=save_results,
        description="preprocessing / pipeline code",
        github_token=GITHUB_TOKEN,
    )


if __name__ == "__main__":
    main()
