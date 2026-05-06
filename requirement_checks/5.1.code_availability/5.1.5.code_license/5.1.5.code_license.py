"""
Question 5.1.5 — Code License (automatic extraction).

Question: "Is the code released under an explicit license?"

Expected values: free text (the license name) or "MISSING".
Examples: "MIT License", "Apache License 2.0", "GNU General Public License v3.0",
"Custom (license file present)", "MISSING".

How it works:
  1. Primary path — call GitHub's licensee endpoint:
        GET /repos/{owner}/{repo}/license
     GitHub runs the licensee tool on the repo and returns a structured
     answer with SPDX id, human name, and the file path.  This is the
     most reliable signal when it works.
  2. Fallback — when GitHub's licensee can't classify the license (or
     returns 404), scan the repo file listing for a root-level file
     whose name looks like a license file (LICENSE, COPYING, …).  If
     one exists, report "Custom (license file present)" so the reviewer
     knows there's *something* there even if the type is unknown.
  3. If neither signal triggers, report "MISSING".

The output mirrors the 5.2.3 / 5.2.2 pattern: one Excel row per paper
with the license name, SPDX id, the path to the license file, and a
short "source" tag (github_api / file_scan / missing) so reviewers can
see how each verdict was derived.

Usage:
  python 5.1.5.code_license.py
"""

import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import load_dotenv

load_dotenv()

_this = Path(__file__).resolve()
# Folder layout for path resolution:
#   _this.parent             = 5.1.5.code_license/
#   _this.parent.parent      = 5.1.code_availability/
#   _this.parent.parent.parent = requirement_checks/   ← needed for common/, data/
sys.path.insert(0, str(_this.parent.parent.parent))

from common.github_helpers import (
    github_get, parse_github_repo, list_all_repo_files,
)
from common.excel_output import thin_border, write_header_row
from data.papers_from_database import PAPERS

MISSING = "MISSING"

# Filenames at the repo root that we treat as "license file present" when
# GitHub's licensee tool didn't return a verdict.  Lowercased for matching.
LICENSE_FILE_BASENAMES = {
    "license", "license.md", "license.txt", "license.rst",
    "licence", "licence.md", "licence.txt",
    "copying", "copying.txt", "copying.lesser", "copying.md",
    "mit-license", "mit-license.txt",
    "gpl", "gpl.txt",
    "unlicense", "unlicense.txt",
}


# =============================================================================
# LOGGING
# =============================================================================

def log(msg):
    """Print a timestamped log message."""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


# =============================================================================
# LICENSE DETECTION
# =============================================================================

def _license_via_github_api(owner, repo):
    """Ask GitHub's licensee endpoint for the repo's license.

    Returns a dict ``{name, spdx_id, path}`` when GitHub recognised a
    license, or None when the endpoint returned 404 / null fields.

    GitHub returns a top-level ``license`` object that is sometimes a
    "no-license" placeholder with key=="other" / spdx_id=="NOASSERTION".
    We treat those as "no verdict" so the file-scan fallback can run.
    """
    data = github_get(f"repos/{owner}/{repo}/license")
    if not data:
        return None

    license_obj = data.get("license") or {}
    spdx_id = license_obj.get("spdx_id") or ""
    name    = license_obj.get("name") or ""

    # Reject GitHub's "we found a file but couldn't classify it" sentinel.
    if not spdx_id or spdx_id == "NOASSERTION":
        return None

    return {
        "name":    name or spdx_id,
        "spdx_id": spdx_id,
        "path":    data.get("path") or "",
    }


def _license_file_via_scan(owner, repo):
    """Fallback: look for a root-level license file in the repo tree.

    Returns the file path string when a license-like basename is found at
    the repo root, otherwise None.  Only the root level is checked because
    license files at greater depth (e.g. ``vendor/foo/LICENSE``) usually
    belong to bundled dependencies, not the project itself.
    """
    files = list_all_repo_files(owner, repo)
    if not files:
        return None

    for entry in files:
        path = entry.get("path", "")
        if "/" in path:
            continue                  # not at repo root
        if path.lower() in LICENSE_FILE_BASENAMES:
            return path
    return None


def detect_license(owner, repo):
    """Detect the license for one repo via GitHub API → file scan → MISSING.

    Returns a dict with keys ``name``, ``spdx_id``, ``path``, ``source``:
      * ``source == "github_api"``  → GitHub's licensee tool classified it.
      * ``source == "file_scan"``    → A license file exists but couldn't be
        classified (custom or unrecognised license text).
      * ``source == "missing"``       → No license signal anywhere.
    """
    via_api = _license_via_github_api(owner, repo)
    if via_api is not None:
        return {**via_api, "source": "github_api"}

    via_scan = _license_file_via_scan(owner, repo)
    if via_scan is not None:
        return {
            "name":    "Custom (license file present)",
            "spdx_id": "",
            "path":    via_scan,
            "source":  "file_scan",
        }

    return {
        "name":    MISSING,
        "spdx_id": "",
        "path":    "",
        "source":  "missing",
    }


# =============================================================================
# PER-PAPER ASSESSMENT
# =============================================================================

def assess_paper(paper):
    """Run license detection on one paper entry.

    Result dict: ``{title, repo, license, spdx_id, path, source, note}``.
    Non-GitHub URLs and unreachable repos are reported as ``MISSING`` with
    a note explaining the reason.
    """
    title = paper.get("title", "")
    url   = paper.get("repo", "") or ""

    base = {
        "title":   title,
        "repo":    url,
        "license": MISSING,
        "spdx_id": "",
        "path":    "",
        "source":  "missing",
        "note":    "",
    }

    if not url:
        base["note"] = "No repo URL provided"
        return base

    owner, repo = parse_github_repo(url)
    if not owner:
        base["note"] = "Not a GitHub repo URL"
        return base

    log(f"  Detecting license: {owner}/{repo}")
    try:
        result = detect_license(owner, repo)
    except RuntimeError as e:
        # github_helpers raises this on rate-limit (HTTP 403)
        base["note"] = str(e)
        return base
    except Exception as e:
        base["note"] = f"Unexpected error: {e}"
        return base

    base["license"] = result["name"]
    base["spdx_id"] = result["spdx_id"]
    base["path"]    = result["path"]
    base["source"]  = result["source"]
    return base


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
DATA_FONT    = Font(name="Arial", size=10)

EXCEL_HEADERS = [
    "#", "Paper Title", "Repo URL",
    "License", "SPDX ID", "License File", "Source", "Notes",
]
EXCEL_COL_WIDTHS = [5, 50, 55, 36, 14, 22, 14, 50]
CENTER_COLS      = {1, 5, 7}   # #, SPDX ID, Source


def save_to_excel(results, filename):
    """Write license-detection results to a one-sheet Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Code License"
    border = thin_border("B8CCE4")

    # Header row + column widths via the shared helper.
    write_header_row(
        ws, EXCEL_HEADERS, EXCEL_COL_WIDTHS,
        fill_hex="BDD7EE", font_color="1F4E79", height=28, border=border,
    )

    # ── Data rows ───────────────────────────────────────────────────────
    for row_idx, r in enumerate(results, start=2):
        row_fill = ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL
        values = [
            row_idx - 1,
            r.get("title", ""),
            r.get("repo", ""),
            r.get("license", MISSING),
            r.get("spdx_id", ""),
            r.get("path", ""),
            r.get("source", ""),
            r.get("note", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in CENTER_COLS else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "A2"

    Path(filename).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filename)
    log(f"Saved to {filename}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Run license detection over every paper and save to Excel."""
    log("\n" + "=" * 80)
    log("Question 5.1.5 — Code License (automatic extraction)")
    log("=" * 80)
    log(f"Processing {len(PAPERS)} papers...\n")

    results = []
    started = time.time()

    for idx, paper in enumerate(PAPERS, 1):
        title = paper.get("title", "")
        log(f"[{idx}/{len(PAPERS)}] {title[:60]}")
        result = assess_paper(paper)
        results.append(result)
        log(f"  => {result['license']} (source={result['source']})")
        log("")
        time.sleep(0.3)

    elapsed = time.time() - started
    log("=" * 80)
    log(f"Done in {elapsed:.1f}s")
    log("=" * 80)

    output_path = _this.parent / "code_license.xlsx"
    save_to_excel(results, str(output_path))


if __name__ == "__main__":
    main()
