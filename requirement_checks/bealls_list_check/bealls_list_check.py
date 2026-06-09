"""Beall's List predatory-venue check — main entry point.

Loads every Semantic Scholar dump in the corpus folder, matches each paper's
publication venue against the vendored Beall's List snapshot, and writes one
combined Excel workbook with three sheets:

  * Results      — one row per paper, colored by status, with full evidence
                   (what matched, on which signal, at what confidence) so any
                   flag can be audited by hand.
  * Summary      — counts per status + per Beall list, and the run metadata
                   (snapshot date, corpus files) needed to interpret them.
  * Flagged only — just the on_list + review rows, so you are not scrolling
                   tens of thousands of clean/preprint rows to find the hits.

Run the scraper first (scrape_bealls_list.py) to produce the snapshot, then:
  python bealls_list_check.py
"""

import json
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import (
    CORPUS_DIR, CORPUS_GLOB, SNAPSHOT_PATH, RESULTS_DIR, RESULTS_FILENAME,
    STATUS_FILL_COLORS,
)
from match import BeallIndex, match_paper


# =============================================================================
# LOADING
# =============================================================================

def load_snapshot():
    """Load the vendored Beall's List snapshot, or exit with guidance."""
    if not SNAPSHOT_PATH.exists():
        sys.exit(
            f"Snapshot not found: {SNAPSHOT_PATH}\n"
            "Run scrape_bealls_list.py first to vendor the Beall's List."
        )
    with open(SNAPSHOT_PATH, encoding="utf-8") as f:
        return json.load(f)


def load_corpus():
    """Yield (source_file, paper) for every record across all corpus files.

    Deduplicates by ``paperId`` across files (the corpus files are disjoint in
    practice, but this guards against accidental overlap).
    """
    files = sorted(CORPUS_DIR.glob(CORPUS_GLOB))
    if not files:
        sys.exit(f"No corpus files matching {CORPUS_GLOB} in {CORPUS_DIR}")
    seen = set()
    for fp in files:
        with open(fp, encoding="utf-8") as f:
            records = json.load(f)
        kept = 0
        for paper in records:
            pid = paper.get("paperId") or paper.get("title")
            if pid in seen:
                continue
            seen.add(pid)
            kept += 1
            yield fp.name, paper
        print(f"  loaded {kept:>6} papers from {fp.name}")


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

# Per-paper layout: each paper occupies a BLOCK of ROWS_PER_PAPER rows.  The #
# and Status cells are merged down the block; the Paper / Venue / Match columns
# each show one labelled field per row (built by the _*_subrows helpers below),
# so every field is its own selectable row instead of text crammed into one
# cell.  Only the DOI row (-> article) and the Matched row (-> Beall entry) are
# hyperlinks.
COLUMNS = [
    ("#",      6),
    ("Status", 12),
    ("Paper",  60),
    ("Venue",  44),
    ("Match",  60),
]

# 1-based column indices, named so the writer reads clearly.
_NUM_COL, _STATUS_COL, _PAPER_COL, _VENUE_COL, _MATCH_COL = 1, 2, 3, 4, 5

# Each paper block is this many rows tall (one labelled field per row).
ROWS_PER_PAPER = 4

_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
_DATA_FONT = Font(name="Arial", size=9)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
# Subtle shade on alternating paper blocks so adjacent papers are easy to tell apart.
_BLOCK_SHADE = "F2F4F8"


def _border():
    edge = Side(style="thin", color="D0D0D0")
    return Border(left=edge, right=edge, top=edge, bottom=edge)


def _paper_subrows(r):
    """Four labelled rows for the Paper column: Title / Year / Source / DOI.

    Returns a list of ``(text, link_url_or_None)``; the DOI row links to the
    article (DOI landing page or open-access PDF). 'Source' is the corpus file.
    """
    return [
        (f"Title: {r.get('title') or '(untitled)'}", None),
        (f"Year: {r.get('year') or '—'}", None),
        (f"Source: {r.get('source_file') or '—'}", None),
        (f"DOI: {r.get('doi') or '—'}", r.get("paper_url") or None),
    ]


def _venue_subrows(r):
    """Four labelled rows for the Venue column: Venue / Type / Domain / ISSN."""
    return [
        (f"Venue: {r.get('venue') or '—'}", None),
        (f"Type: {r.get('venue_type') or '—'}", None),
        (f"Domain: {r.get('venue_domain') or '—'}", None),
        (f"ISSN: {r.get('issn') or '—'}", None),
    ]


# Plain-language description of each match signal, used both in the Match column
# and to build the Legend so the two never drift.  Tuple = (short label shown in
# the cell, which verdict it leads to, thorough explanation with a FIXED real
# example).  The examples are constant real Beall's List cases — they do NOT
# depend on the corpus, so every generated workbook shows the same illustrations.
SIGNAL_INFO = {
    "domain": (
        "the journal's own website is on the list",
        "on_list",
        "The journal's website address is itself on Beall's List. Example: a "
        "paper in a journal hosted at www.scirp.org is matched because "
        "scirp.org (Scientific Research Publishing) is a publisher on the list "
        "— the journal's own home is a listed site. This is the most reliable "
        "signal.",
    ),
    "domain_subdomain": (
        "the journal's website is part of a listed site",
        "on_list",
        "The journal's website is a section of a bigger website that is on the "
        "list. Example: a journal at ojs.bilpublishing.com is matched because "
        "it sits under bilpublishing.com (Bilingual Publishing Co.), a listed "
        "publisher.",
    ),
    "issn": (
        "the journal's ISSN is on the list",
        "on_list",
        "An ISSN is the unique ID number every journal has. Here the journal's "
        "ISSN equals the ISSN of a journal on the list. (Rarely used, because "
        "Beall's List mostly records names and websites, not ISSNs.)",
    ),
    "name_exact": (
        "the journal/publisher name exactly matches the list",
        "on_list",
        "The journal or publisher name is exactly the same as a name on the "
        "list (capitalisation, accents and punctuation are ignored). Example: a "
        "venue named 'OMICS Publishing Group' matches the listed 'OMICS "
        "Publishing Group'.",
    ),
    "name_fuzzy": (
        "the journal name is very similar to a listed name",
        "review",
        "The name is almost — but not exactly — the same as a listed name (at "
        "least 93% similar): a spelling variant or typo, e.g. 'Internatonal "
        "Journal of Science' vs the listed 'International Journal of Science'. "
        "Because it is not exact it is only ever flagged for review, never as a "
        "definite hit.",
    ),
    "domain_alternate_url": (
        "an alternate web link points to a listed site",
        "review",
        "The venue's main website looks fine, but one of the OTHER web links "
        "Semantic Scholar lists for it points to a listed site (e.g. a "
        "scirp.org link). Semantic Scholar sometimes merges two journals that "
        "share a name, so that link may belong to a different journal — a human "
        "should check.",
    ),
    "domain_open_access_pdf": (
        "the PDF is hosted on a listed site",
        "review",
        "The journal itself could not be identified, but the free PDF of the "
        "paper is hosted on a listed site (e.g. scirp.org). Suggestive but not "
        "conclusive, so it is flagged for review.",
    ),
}

# Plain-language name + weight for each Beall list.
LIST_INFO = {
    "publishers":         ("predatory publisher", "core list"),
    "standalone_journal": ("standalone predatory journal", "core list"),
    "hijacked":           ("hijacked / cloned journal", "core list"),
    "vanity_press":       ("vanity press (book publisher)", "weak — review only"),
    "misleading_metric":  ("fake-metrics company", "weak — review only"),
}


def _signal_plain(matched_on):
    """Short plain-language phrase for a match signal (falls back to the code)."""
    info = SIGNAL_INFO.get(matched_on)
    return info[0] if info else (matched_on or "—")


def _list_plain(list_source):
    """Plain-language name(s) for a (possibly '; '-joined) list_source."""
    parts = [p.strip() for p in (list_source or "").split(";") if p.strip()]
    if not parts:
        return "—"
    return "; ".join(LIST_INFO.get(p, (p,))[0] for p in parts)


# Short note shown on the first Match row for statuses that matched nothing.
_NONMATCH_NOTES = {
    "clean":    "Venue identified — not on the list",
    "no_venue": "No venue to classify (preprint / no info)",
}


def _match_subrows(r):
    """Four labelled rows for the Match column.

    For a flagged paper:
      Listed as   — the Beall's List entry that matched (links to it)
      Beall list  — which Beall list it came from, in plain language
      Matched by  — how it matched, in plain language (see the Legend)
      Why flagged — a plain reason (a high-confidence note for on_list, or the
                    verify-by-hand reason for review)
    For clean / no_venue / error rows the first row carries a short note and the
    rest are blank.
    """
    status = r.get("status")
    if status in _NONMATCH_NOTES:
        note = _NONMATCH_NOTES[status]
    elif status == "error":
        note = f"Error: {r.get('matched_on') or 'could not process record'}"
    else:
        reason = (r.get("review_reason")
                  or "Appears on Beall's List — a high-confidence match.")
        return [
            (f"Listed as: {r.get('matched_name') or '(matched)'}", r.get("matched_url") or None),
            (f"Beall list: {_list_plain(r.get('list_source'))}", None),
            (f"Matched by: {_signal_plain(r.get('matched_on'))}", None),
            (f"Why flagged: {reason}", None),
        ]
    return [(note, None), ("", None), ("", None), ("", None)]


def _estimate_row_height(text, width):
    """Row height (px) tall enough for one cell's text to wrap without clipping.

    openpyxl does not auto-fit rows, so a long Title would clip at the default
    height.  Approximate wrapping by dividing the text length by the column's
    character width.
    """
    lines = max(1, -(-len(str(text)) // max(1, width)))      # ceil(len / width)
    return min(140, max(16, lines * 15))


def _write_results_sheet(ws, results):
    """Write the Results sheet: one multi-row block per paper.

    Each paper spans ROWS_PER_PAPER rows.  The # and Status cells are merged
    down the block; the Paper / Venue / Match columns show one labelled field
    per row, so each field is its own selectable row.  Cells are fully styled
    *before* the merge so borders render across the merged ranges.
    """
    border = _border()
    for col, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font, cell.fill, cell.alignment, cell.border = (
            _HEADER_FONT, _HEADER_FILL, _CENTER, border
        )
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    link_font = Font(name="Arial", size=9, color="0563C1", underline="single")
    status_font = Font(name="Arial", size=9, bold=True)
    # Heavier top edge on each paper's first row draws a clear divider between
    # one paper's block and the next, so blocks read as distinct units.
    _edge = Side(style="thin", color="D0D0D0")
    block_top_border = Border(left=_edge, right=_edge, bottom=_edge,
                              top=Side(style="medium", color="7F9BC4"))

    row = 2
    for i, r in enumerate(results):
        top = row
        bottom = top + ROWS_PER_PAPER - 1
        shade_fill = PatternFill("solid", fgColor=_BLOCK_SHADE) if i % 2 else None
        status = r.get("status", "")
        status_fill = PatternFill(
            "solid", fgColor=STATUS_FILL_COLORS.get(status, "FFFFFF")
        )
        col_subrows = (
            (_PAPER_COL, _paper_subrows(r)),
            (_VENUE_COL, _venue_subrows(r)),
            (_MATCH_COL, _match_subrows(r)),
        )

        for sub in range(ROWS_PER_PAPER):
            rr = top + sub
            cell_border = block_top_border if sub == 0 else border   # divider on row 1
            # # and Status: value only on the block's first row; merged later.
            num = ws.cell(row=rr, column=_NUM_COL, value=(i + 1) if sub == 0 else None)
            num.font, num.border, num.alignment = _DATA_FONT, cell_border, _CENTER
            if shade_fill:
                num.fill = shade_fill
            st = ws.cell(row=rr, column=_STATUS_COL, value=status if sub == 0 else None)
            st.font, st.border, st.alignment, st.fill = (
                status_font, cell_border, _CENTER, status_fill
            )

            height = 16
            for col, subrows in col_subrows:
                text, url = subrows[sub]
                cell = ws.cell(row=rr, column=col, value=text)
                cell.font, cell.border, cell.alignment = _DATA_FONT, cell_border, _LEFT
                if shade_fill:
                    cell.fill = shade_fill
                if url:
                    cell.hyperlink = url
                    cell.font = link_font
                height = max(height, _estimate_row_height(text, COLUMNS[col - 1][1]))
            ws.row_dimensions[rr].height = height

        # Merge # and Status down the block (after styling, so borders draw).
        if ROWS_PER_PAPER > 1:
            ws.merge_cells(start_row=top, start_column=_NUM_COL,
                           end_row=bottom, end_column=_NUM_COL)
            ws.merge_cells(start_row=top, start_column=_STATUS_COL,
                           end_row=bottom, end_column=_STATUS_COL)

        row = bottom + 1


def _write_summary_sheet(ws, results, snapshot_meta, corpus_files):
    """Write counts + run metadata so results are interpretable."""
    border = _border()
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60

    def header(row, a, b):
        for col, val in ((1, a), (2, b)):
            c = ws.cell(row=row, column=col, value=val)
            c.font, c.fill, c.alignment, c.border = (
                _HEADER_FONT, _HEADER_FILL, _CENTER, border
            )

    def kv(row, label, value):
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        c = ws.cell(row=row, column=2, value=value)
        c.font = Font(name="Arial", size=10)
        c.alignment = _LEFT

    total = len(results)
    status_counts = {s: 0 for s in STATUS_FILL_COLORS}
    list_counts = {}
    signal_counts = {}        # matched_on -> count, over flagged rows only
    for r in results:
        status_counts[r["status"]] = status_counts.get(r["status"], 0) + 1
        if r["status"] in ("on_list", "review"):
            if r["list_source"]:
                list_counts[r["list_source"]] = list_counts.get(r["list_source"], 0) + 1
            mo = r.get("matched_on") or "—"
            signal_counts[mo] = signal_counts.get(mo, 0) + 1

    row = 1
    header(row, "Metric", "Value"); row += 1
    kv(row, "Total papers checked", total); row += 1
    kv(row, "On Beall's List (high confidence)", status_counts.get("on_list", 0)); row += 1
    kv(row, "Needs review (soft match)", status_counts.get("review", 0)); row += 1
    kv(row, "Clean (venue identified, not listed)", status_counts.get("clean", 0)); row += 1
    kv(row, "No venue (preprint / no info)", status_counts.get("no_venue", 0)); row += 1
    kv(row, "Errors", status_counts.get("error", 0)); row += 2

    # Breakdown by HOW each flagged paper matched (the 'Matched by' signal),
    # e.g. how many were 'an alternate web link points to a listed site'.
    header(row, "Flagged papers by 'Matched by' signal", "Count"); row += 1
    for mo in sorted(signal_counts, key=lambda k: (-signal_counts[k], k)):
        kv(row, f"  {_signal_plain(mo)}  [{mo}]", signal_counts[mo]); row += 1
    if not signal_counts:
        kv(row, "  (none)", 0); row += 1
    row += 1

    # Breakdown by which Beall list each flagged paper matched.
    header(row, "Flagged papers by Beall list", "Count"); row += 1
    for src in sorted(list_counts, key=lambda k: (-list_counts[k], k)):
        kv(row, f"  {_list_plain(src)}", list_counts[src]); row += 1
    if not list_counts:
        kv(row, "  (none)", 0); row += 1
    row += 1

    header(row, "Run metadata", ""); row += 1
    kv(row, "Snapshot scraped (UTC)", snapshot_meta.get("scraped_utc", "")); row += 1
    kv(row, "Snapshot entries", snapshot_meta.get("total_entries", "")); row += 1
    kv(row, "Snapshot source", snapshot_meta.get("source", "")); row += 1
    kv(row, "Corpus files", ", ".join(corpus_files)); row += 1
    kv(row, "Check run (local)", datetime.now().strftime("%Y-%m-%d %H:%M:%S")); row += 2

    header(row, "Caveat", ""); row += 1
    kv(row, "Interpretation", snapshot_meta.get("note", "")); row += 1
    ws.cell(row=row, column=1,
            value="'on_list' = exact domain/ISSN/name match. 'review' = fuzzy or "
                  "soft match — verify by hand before trusting.").font = Font(
        name="Arial", size=9, italic=True)


def _write_flagged_sheet(ws, results):
    """Write only the on_list + review rows (the actionable subset)."""
    flagged = [r for r in results if r["status"] in ("on_list", "review")]
    # Sort: on_list first, then by list source, then title.
    order = {"on_list": 0, "review": 1}
    flagged.sort(key=lambda r: (order.get(r["status"], 9), r["list_source"], r["title"]))
    _write_results_sheet(ws, flagged)
    return len(flagged)


# Legend-sheet content as one declarative structure.  Each section is a titled
# mini-table of (col-A, col-B, col-C) rows.  headers=None renders the rows as
# full-width bullet lines (the caveats); swatch=True tints column A by status
# to match the Results sheet colors.  Add/edit documentation here only — the
# renderer below is generic.
_LEGEND_TITLE = "Beall's List Check — Legend / Data Dictionary"
_LEGEND_SECTIONS = [
    {
        "title": "STATUS values (and the row color used in Results)",
        "headers": ("Status", "Confidence", "Meaning"),
        "swatch": True,
        "rows": [
            ("on_list", "high", "Venue matched a CORE Beall list (predatory "
             "publishers, standalone journals, or hijacked journals) by a "
             "high-confidence signal: exact domain, subdomain of a listed "
             "domain, exact ISSN, or exact name. Read as \"appears on Beall's "
             "List.\""),
            ("review", "verify", "A softer or contested match: fuzzy name, "
             "open-access PDF host, an alternate venue URL, or a WEAK list "
             "(vanity-press / misleading-metrics). Check by hand."),
            ("clean", "—", "The venue was identified and did NOT match any "
             "Beall entry."),
            ("no_venue", "—", "Preprint server (e.g. arXiv) or no venue "
             "metadata — there is no journal/publisher to classify."),
            ("error", "—", "The record could not be processed; the Python "
             "error is shown in the 'Matched On' column."),
        ],
    },
    {
        "title": "'Matched by' — how the venue was matched to Beall's List",
        "headers": ("How it matched (shown in the 'Match' column)",
                    "Result", "What it means — with a fixed real example"),
        # Rows are generated from SIGNAL_INFO so this table always matches the
        # phrases shown in the Match column.  Examples are constant (corpus-
        # independent) so every workbook documents the same illustrations.
        "rows": [
            (SIGNAL_INFO[k][0], SIGNAL_INFO[k][1], SIGNAL_INFO[k][2])
            for k in ("domain", "domain_subdomain", "issn", "name_exact",
                      "name_fuzzy", "domain_alternate_url", "domain_open_access_pdf")
        ],
    },
    {
        "title": "'Beall list' — which Beall list the entry came from",
        "headers": ("Beall list (shown in the 'Match' column)", "Weight", "Meaning"),
        "rows": [
            ("predatory publisher", "core list",
             "A predatory publisher — Beall's main list."),
            ("standalone predatory journal", "core list",
             "A single predatory journal that is not part of a larger publisher."),
            ("hijacked / cloned journal", "core list",
             "A fake journal impersonating a legitimate one by reusing its name "
             "and look (only the fake side is listed). A name-only match is "
             "downgraded to review, since the clone reuses the real journal's "
             "name; only a website match is asserted as on_list."),
            ("vanity press (book publisher)", "weak — review only",
             "A book/monograph publisher (e.g. IGI Global). Not a journal-quality "
             "signal, so a match here is downgraded to review."),
            ("fake-metrics company", "weak — review only",
             "A company selling fake impact-factor metrics — not a venue at all, "
             "so a match is downgraded to review."),
        ],
    },
    {
        "title": "Layout of the Results / Flagged-only sheets",
        "headers": ("Column", "Rows (top -> bottom)", "Meaning"),
        "rows": [
            ("(blocks)", "4 rows per paper", "Each paper occupies a 4-row "
             "block; alternating blocks are lightly shaded so they are easy to "
             "tell apart."),
            ("#", "(merged)", "Sequential paper number; the cell spans the "
             "paper's whole block."),
            ("Status", "(merged)", "Classification of the paper (see the Status "
             "section); spans the block and is colored by status."),
            ("Paper", "Title / Year / Source / DOI",
             "Paper identity, one field per row. 'Source' is which corpus JSON "
             "the paper came from. The DOI row is a link to the article at its "
             "publisher."),
            ("Venue", "Venue / Type / Domain / ISSN",
             "The publication venue from Semantic Scholar, one field per row. "
             "'Type' is journal/conference (conferences are out of scope); "
             "'Domain' is the venue website host — the main match key."),
            ("Match", "Listed as / Beall list / Matched by / Why flagged",
             "Why the paper was flagged, one field per row. 'Listed as' is the "
             "Beall's List entry that matched (a link to it); 'Beall list' is "
             "which list it came from (see the 'Beall list' section); 'Matched "
             "by' is how it matched (see the 'Matched by' section); 'Why "
             "flagged' is a plain-English reason. For clean / no_venue / error "
             "rows the first row shows a short note and the rest are blank."),
        ],
    },
    {
        "title": "Caveats — read before interpreting results",
        "headers": None,
        "rows": [
            ("Beall's List is an unofficial, archived, frozen (~2021), CONTESTED "
             "list. A match is an allegation as of the snapshot date, not proof "
             "a venue is predatory.",),
            ("Being on the list is contested. Frontiers, for example, is on the "
             "list but widely considered legitimate, so 'on_list' = \"appears "
             "on the list,\" not \"is predatory.\" (MDPI is in the site's "
             "'Excluded — decide after reading' section, so it is NOT flagged.)",),
            ("This check classifies the VENUE, not the paper's content/quality.",),
            ("It uses NO LLM (0 tokens). Every result is deterministic and "
             "auditable from the Matched On / Matched Entry columns.",),
            ("no_venue is mostly arXiv/preprints, which have no journal "
             "publisher and so cannot be on Beall's List.",),
        ],
    },
]


def _write_legend_sheet(ws):
    """Render the Legend tab from the declarative ``_LEGEND_SECTIONS``."""
    border = _border()
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 95
    bold = Font(name="Arial", size=10, bold=True)
    data = Font(name="Arial", size=10)
    section_font = Font(name="Arial", size=12, bold=True, color="2F5496")

    ws.cell(row=1, column=1, value=_LEGEND_TITLE).font = Font(
        name="Arial", size=14, bold=True)
    row = 3

    for sec in _LEGEND_SECTIONS:
        ws.cell(row=row, column=1, value=sec["title"]).font = section_font
        row += 1

        if sec["headers"] is None:                      # caveats: bullet lines
            for (text,) in sec["rows"]:
                cell = ws.cell(row=row, column=1, value="• " + text)
                cell.font, cell.alignment = data, _LEFT
                ws.merge_cells(start_row=row, start_column=1,
                               end_row=row, end_column=3)
                row += 1
            row += 1
            continue

        for col, header in enumerate(sec["headers"], 1):   # header row
            cell = ws.cell(row=row, column=col, value=header)
            cell.font, cell.fill, cell.alignment, cell.border = (
                _HEADER_FONT, _HEADER_FILL, _CENTER, border)
        row += 1

        for cells in sec["rows"]:                          # data rows
            for col, val in enumerate(cells, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = bold if col == 1 else data
                cell.border, cell.alignment = border, _LEFT
                if col == 1 and sec.get("swatch"):
                    cell.fill = PatternFill(
                        "solid", fgColor=STATUS_FILL_COLORS.get(val, "FFFFFF"))
            row += 1
        row += 1


def save_workbook(results, snapshot_meta, corpus_files):
    """Build and save the combined workbook; return its path."""
    wb = openpyxl.Workbook()
    ws_results = wb.active
    ws_results.title = "Results"
    _write_results_sheet(ws_results, results)

    ws_flagged = wb.create_sheet("Flagged only")
    n_flagged = _write_flagged_sheet(ws_flagged, results)

    ws_summary = wb.create_sheet("Summary")
    _write_summary_sheet(ws_summary, results, snapshot_meta, corpus_files)

    ws_legend = wb.create_sheet("Legend")
    _write_legend_sheet(ws_legend)

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = RESULTS_DIR / RESULTS_FILENAME
    try:
        wb.save(out_path)
    except PermissionError:
        # The target is almost certainly open in Excel (Windows locks it).
        # Fall back to a timestamped name so a run is never lost.
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = RESULTS_DIR / f"{Path(RESULTS_FILENAME).stem}_{stamp}.xlsx"
        wb.save(out_path)
        print(f"  (default file was locked — saved to {out_path.name} instead)")
    return out_path, n_flagged


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 70)
    print("Beall's List predatory-venue check")
    print("=" * 70)

    snapshot = load_snapshot()
    index = BeallIndex(snapshot["entries"])
    print(f"Loaded snapshot: {len(snapshot['entries'])} entries "
          f"(scraped {snapshot['meta'].get('scraped_utc', '?')})")

    print(f"Loading corpus from {CORPUS_DIR} ...")
    started = time.time()
    results = []
    corpus_files = set()
    for source_file, paper in load_corpus():
        corpus_files.add(source_file)
        # match_paper never raises — a bad record returns an 'error' result.
        results.append(match_paper(index, paper, source_file))

    elapsed = time.time() - started
    out_path, n_flagged = save_workbook(
        results, snapshot["meta"], sorted(corpus_files)
    )

    # Console summary.
    counts = {s: 0 for s in STATUS_FILL_COLORS}
    for r in results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    print("-" * 70)
    print(f"Checked {len(results)} papers in {elapsed:.1f}s")
    print(f"  on_list : {counts.get('on_list', 0)}")
    print(f"  review  : {counts.get('review', 0)}")
    print(f"  clean   : {counts.get('clean', 0)}")
    print(f"  no_venue: {counts.get('no_venue', 0)}")
    print(f"  error   : {counts.get('error', 0)}")
    print(f"Flagged (on_list + review): {n_flagged}")
    print(f"Saved -> {out_path}")


if __name__ == "__main__":
    main()
