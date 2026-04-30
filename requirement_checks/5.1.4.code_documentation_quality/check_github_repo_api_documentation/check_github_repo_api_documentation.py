"""
Checks whether a GitHub repository linked to an academic paper provides API
documentation — either generated documentation (Sphinx, Doxygen, MkDocs, etc.)
or docstrings written directly in the source code.

Logic:
  1. List every file in the repo via the GitHub API.                → collect_repo_content()
  2. Sort files into four buckets:
       doc-tool config files (mkdocs.yml, Doxyfile, conf.py, etc.)
       API spec files (openapi.yaml, swagger.json, etc.)
       documentation pages (docs/api/, docs/reference/, etc.)
       source code files (scanned for docstrings).                  → collect_repo_content()
  3. Fetch all buckets within a shared character budget, prioritising
     doc-tool configs first, then API specs, then doc pages,
     then source code.                                               → collect_repo_content()
  4. Pass the combined text to the LLM, asking it to identify every
     form of API documentation present.                             → make_check_paper_fn()
  5. Enrich the LLM output with direct GitHub hyperlinks.           → build_doc_file_entries()
  6. Run across all configured models and export to Excel.          → main(), save_results()

Results export to Excel.
"""

import os
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Add parent directories to Python path for local imports
_this = Path(__file__).resolve()
sys.path.insert(0, str(_this.parent.parent))          # 5.1.4.code_documentation_quality/
sys.path.insert(0, str(_this.parent.parent.parent))   # requirement_checks/

from common.fetch_and_parse_github_repo import (
    load_dotenv, parse_github_repo, list_all_repo_files, fetch_file_content,
)
from common.repo_content_helpers import fetch_paths_with_char_budget
from common.llm_helpers import llm_call_parse_retry
from common.checker_pipeline import run_pipeline
from common.excel_output import (
    STATUS_FILL_COLORS, count_statuses,
    write_header_row, write_results_data_rows, write_summary_sheet,
    write_comparison_sheet, safe_sheet_name,
    thin_border, auto_row_height, alignment_center, alignment_wrap_left,
)
from shared import check_paper_generic

load_dotenv()

from papers_from_database import PAPERS
from openai_client import client, AZURE_OPENAI_DEPLOYMENTS as DEPLOYMENTS
from prompts import SYSTEM_PROMPT
from config import (
    SOURCE_CODE_EXTENSIONS, SKIP_FOLDER_PREFIXES, SKIP_BASENAMES,
    TARGET_DOC_FILENAMES, DOC_FOLDER_PREFIXES, DOC_PAGE_EXTENSIONS,
    MAX_SOURCE_FILES, MAX_DOC_FILES,
    MAX_CONTENT_CHARS, SOURCE_FILE_CHAR_CAP, DOC_FILE_CHAR_CAP,
)

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")


# =============================================================================
# GITHUB API HELPERS
# =============================================================================


def _is_source_file(path_lower, basename):
    """Return True if the file is a source code file worth scanning for docstrings.

    Excludes vendored dependencies, auto-generated files, and test fixtures so
    that only code authored by the paper's developers is evaluated.
    """
    ext = "." + basename.rsplit(".", 1)[-1] if "." in basename else ""
    if ext not in SOURCE_CODE_EXTENSIONS:
        return False
    if any(path_lower.startswith(prefix) for prefix in SKIP_FOLDER_PREFIXES):
        return False
    if basename in SKIP_BASENAMES:
        return False
    return True


def _diversified_source_sample(paths, max_files):
    """Pick up to ``max_files`` source paths, spread across top-level directories.

    Files are bucketed by their first path component (root-level files form
    their own bucket). Within each bucket, files are ordered alphabetically.
    The function then walks the buckets round-robin — one file from each on
    each pass — so that no single directory can monopolise the budget.

    Root-level files are visited first on each round because they tend to be
    entry points or core modules; the remaining buckets are visited in
    alphabetical order for determinism.
    """
    if not paths or max_files <= 0:
        return []

    buckets = {}
    for p in paths:
        top = "" if "/" not in p else p.split("/", 1)[0]
        buckets.setdefault(top, []).append(p)

    for key in buckets:
        buckets[key].sort(key=str.lower)

    # Root-level bucket ("") first, then remaining directories alphabetically.
    ordered_keys = [""] if "" in buckets else []
    ordered_keys += sorted((k for k in buckets if k != ""), key=str.lower)

    selected = []
    round_idx = 0
    while len(selected) < max_files:
        added_this_round = False
        for key in ordered_keys:
            if round_idx < len(buckets[key]):
                selected.append(buckets[key][round_idx])
                added_this_round = True
                if len(selected) >= max_files:
                    break
        if not added_this_round:
            break
        round_idx += 1

    return selected


def collect_repo_content(owner, repo):
    """Fetch documentation artefacts and source code; return combined text and metadata.

    Files are collected in four prioritised buckets:
      1. Documentation tool config files (mkdocs.yml, conf.py, Doxyfile, etc.)
         and API spec files — strongest evidence of generated documentation.
      2. Documentation page files from docs/api/, docs/reference/, docs/, etc.
      3. Source code files — inspected for docstrings.

    Returns a tuple of (combined_text, fetched_paths).
    """
    all_files = list_all_repo_files(owner, repo)
    if not all_files:
        return "", []

    doc_config_paths = []   # tool config + api spec files
    doc_page_paths   = []   # documentation pages from doc folders
    source_paths     = []   # source code files for docstring scanning

    for f in all_files:
        fpath     = f["path"]
        fpath_low = fpath.lower()
        basename  = fpath_low.split("/")[-1]
        ext       = "." + basename.rsplit(".", 1)[-1] if "." in basename else ""

        # Named documentation config / spec files (exact match or nested)
        if fpath_low in TARGET_DOC_FILENAMES or basename in TARGET_DOC_FILENAMES:
            doc_config_paths.append(fpath)

        # Documentation pages inside designated doc folders
        elif any(fpath_low.startswith(p) for p in DOC_FOLDER_PREFIXES):
            if ext in DOC_PAGE_EXTENSIONS:
                doc_page_paths.append(fpath)

        # Source code files for docstring inspection
        elif _is_source_file(fpath_low, basename):
            source_paths.append(fpath)

    # Sort config/spec files: root-level first, then nested
    doc_config_paths.sort(key=lambda p: (p.count("/"), p.lower()))

    # Sort doc pages: api/reference subfolders first, then by depth
    def _doc_page_priority(p):
        p_low = p.lower()
        if any(p_low.startswith(x) for x in ("docs/api/", "docs/reference/",
                                               "apidoc/", "api_docs/", "api-docs/")):
            return (0, p_low)
        return (1, p_low)

    doc_page_paths.sort(key=_doc_page_priority)
    doc_page_paths = doc_page_paths[:MAX_DOC_FILES]

    # Sample source files with directory diversification.
    #
    # A plain `sort + [:MAX_SOURCE_FILES]` would fill the budget with whichever
    # top-level folder sorts first alphabetically (e.g. `a_test_online_model/`,
    # `fintune/`) and silently drop every file in late-alphabet folders
    # (`utils/`, `src/`, `tools/`). Docstrings commonly live in utility/helper
    # modules, so that systematically hides positive cases.
    #
    # Instead: bucket files by top-level directory, then round-robin across
    # buckets so every directory contributes before any single one dominates.
    source_paths = _diversified_source_sample(source_paths, MAX_SOURCE_FILES)

    fetched     = []
    combined    = []
    total_chars = 0

    # ── 1. Fetch documentation tool config / API spec files ──────────────────
    config_blocks, fetched_configs, total_chars = fetch_paths_with_char_budget(
        owner, repo, doc_config_paths, MAX_CONTENT_CHARS,
        fetch_content=fetch_file_content,
        start_total_chars=total_chars,
        pause_seconds=0.15,
        header_label="DOC CONFIG",
        per_file_char_cap=DOC_FILE_CHAR_CAP,
    )
    combined.extend(config_blocks)
    fetched.extend(fetched_configs)

    # ── 2. Fetch documentation page files ────────────────────────────────────
    page_blocks, fetched_pages, total_chars = fetch_paths_with_char_budget(
        owner, repo, doc_page_paths, MAX_CONTENT_CHARS,
        fetch_content=fetch_file_content,
        start_total_chars=total_chars,
        pause_seconds=0.15,
        header_label="DOC PAGE",
        per_file_char_cap=DOC_FILE_CHAR_CAP,
    )
    combined.extend(page_blocks)
    fetched.extend(fetched_pages)

    # ── 3. Fetch source code files for docstring inspection ──────────────────
    source_blocks, fetched_sources, total_chars = fetch_paths_with_char_budget(
        owner, repo, source_paths, MAX_CONTENT_CHARS,
        fetch_content=fetch_file_content,
        start_total_chars=total_chars,
        pause_seconds=0.15,
        header_label="SOURCE FILE",
        per_file_char_cap=SOURCE_FILE_CHAR_CAP,
    )
    combined.extend(source_blocks)
    fetched.extend(fetched_sources)

    return "\n\n".join(combined), fetched


# =============================================================================
# LLM CALL WRAPPER
# =============================================================================

EMPTY_API_DOC_PAYLOAD = {
    "has_api_documentation": None,
    "evidence": "Empty LLM response after retry — content may exceed context window",
    "documentation_types": [],
    "doc_tool": "none",
    "doc_files": [],
}

# Content cache — avoids re-fetching GitHub repos when running multiple models.
_content_cache = {}


def _cached_collect_repo_content(owner, repo):
    """Return cached collect_repo_content result, fetching only on first call."""
    key = f"{owner}/{repo}"
    if key not in _content_cache:
        _content_cache[key] = collect_repo_content(owner, repo)
    return _content_cache[key]


# =============================================================================
# PER-PAPER ORCHESTRATION
# =============================================================================

def build_doc_file_entries(doc_files_raw, owner, repo):
    """Enrich raw LLM doc_files list with GitHub hyperlinks and normalised fields."""
    entries = []
    for item in (doc_files_raw or []):
        path = (item.get("path") or "").strip()
        if not path:
            continue
        entries.append({
            "path":        path,
            "type":        item.get("type", "other"),
            "description": item.get("description", ""),
            "link":        f"https://github.com/{owner}/{repo}/blob/HEAD/{path}",
        })
    return entries


def make_check_paper_fn(deployment, token_usage):
    """Factory: return a check_paper function bound to a specific LLM deployment.

    Using a factory instead of a plain function lets run_pipeline swap in different
    models without re-fetching GitHub content — the content cache is shared across
    all model instances for the same repo.
    """

    def _llm_check(repo_content, paper_title):
        """Call the LLM and parse the API-documentation verdict for one paper."""
        def build_msg(content):
            return (
                f"Paper: {paper_title}\n\n"
                "Below are the contents of documentation configuration files, "
                "documentation pages, and source code files fetched from its GitHub repository.\n"
                "Determine whether the repository contains API documentation "
                "(generated docs or docstrings).\n\n"
                + (content if content else "[No relevant files found in the repository]")
            )

        token_budget = 6000 if len(repo_content) > 50_000 else 3000

        return llm_call_parse_retry(
            client=client,
            deployment=deployment,
            system_prompt=SYSTEM_PROMPT,
            build_user_message=build_msg,
            content=repo_content,
            token_usage=token_usage,
            empty_payload=EMPTY_API_DOC_PAYLOAD,
            required_list_fields=("documentation_types", "doc_files"),
            max_completion_tokens=token_budget,
            retry_truncate_chars=30_000,
            retry_max_tokens=12000,
            preview_chars=1000,
        )

    def _map_api_doc(result, llm_result, owner, repo):
        result["documentation_types"] = llm_result.get("documentation_types", [])
        result["doc_tool"]            = llm_result.get("doc_tool", "none") or "none"
        result["doc_files"]           = build_doc_file_entries(
            llm_result.get("doc_files", []), owner, repo
        )

    def check_paper(paper):
        return check_paper_generic(
            paper,
            extra_defaults={
                "documentation_types": [],
                "doc_tool":            "none",
                "doc_files":           [],
            },
            collect_content_fn=_cached_collect_repo_content,
            llm_check_fn=_llm_check,
            map_llm_result_fn=_map_api_doc,
            boolean_key="has_api_documentation",
        )

    return check_paper


# =============================================================================
# CONSOLE REPORTING
# =============================================================================

def print_results(results):
    """Print a compact console table of per-paper results with documentation types."""
    W = 110
    print("\n" + "=" * W)
    print(f"{'#':<4} {'STATUS':<10} {'DOC TOOL':<14} {'DOC TYPES':<30} TITLE")
    print("=" * W)

    for i, r in enumerate(results, 1):
        icon = {
            "yes": "YES", "no": "NO", "skipped": "SKIP", "error": "ERR",
        }.get(r["status"], "?")

        types_str = ", ".join(r.get("documentation_types", []))[:28]
        if not types_str and r.get("note"):
            types_str = r["note"][:28]

        doc_tool    = (r.get("doc_tool") or "none")[:12]
        title_short = r["title"][:46] + ("..." if len(r["title"]) > 46 else "")
        print(f"{i:<4} {icon:<10} {doc_tool:<14} {types_str:<30} {title_short}")

        if r.get("evidence"):
            print(f"       evidence: {r['evidence']}")
        for df in r.get("doc_files", [])[:3]:
            print(f"       [{df['type']}] {df['path']}  →  {df['description']}")

    print("=" * W)
    counts  = count_statuses(results)
    yes     = counts.get("yes", 0)
    no      = counts.get("no", 0)
    skipped = counts.get("skipped", 0)
    errors  = counts.get("error", 0)
    total_files = sum(len(r.get("doc_files", [])) for r in results)
    print(
        f"\nSUMMARY: {yes} repos have API docs | {no} missing | "
        f"{skipped} skipped | {errors} errors | {total_files} total doc files found\n"
    )


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def save_results(all_model_results, path=None):
    """Write per-model Excel sheets + Model Comparison sheet.

    Per model: Results, Doc Files Detail, Skipped & Errors, Summary.
    """
    if path is None:
        path = Path(__file__).resolve().parent / "results/api_documentation_results.xlsx"

    wb     = Workbook()
    border = thin_border()
    center = alignment_center()
    wrap   = alignment_wrap_left()

    deployments = list(all_model_results.keys())
    first_sheet = True

    for deployment in deployments:
        model_data  = all_model_results[deployment]
        results     = model_data["results"]
        token_usage = model_data["token_usage"]

        # =====================================================================
        # Results sheet (one row per paper)
        # =====================================================================
        if first_sheet:
            ws1 = wb.active
            first_sheet = False
        else:
            ws1 = wb.create_sheet()
        ws1.title = safe_sheet_name("Results", deployment)

        hdrs1   = ["#", "Status", "Title", "Repo",
                   "# Doc Files", "Doc Tool", "Documentation Types", "Evidence",
                   "All Doc Links", "Files Checked", "Note", "Tokens Used"]
        widths1 = [5, 10, 45, 40, 10, 18, 40, 55, 60, 50, 35, 12]
        write_header_row(ws1, hdrs1, widths1, fill_hex="2F5496", border=border)

        def results_row_data(r, num):
            all_links_text = "\n".join(
                f"{df['path']}  ({df['type']})"
                for df in r.get("doc_files", [])
            )
            return [
                num,
                (r.get("status") or "").upper(),
                r.get("title") or "",
                r.get("repo") or "",
                len(r.get("doc_files", [])),
                r.get("doc_tool") or "none",
                ", ".join(r.get("documentation_types") or []),
                r.get("evidence") or "",
                all_links_text,
                ", ".join(r.get("files_checked") or []),
                r.get("note") or "",
                r.get("tokens_used") or "",
            ]

        write_results_data_rows(
            ws1, results, results_row_data,
            border=border,
            row_height_fn=lambda vals: auto_row_height(vals, line_height=15 * 1.3),
        )

        # =====================================================================
        # Doc Files Detail sheet (one row per individual documented file)
        # =====================================================================
        ws2     = wb.create_sheet(safe_sheet_name("Doc Files", deployment))
        hdrs2   = ["Paper #", "Paper Title", "Repo", "File Path",
                   "Type", "Description", "GitHub Link"]
        widths2 = [8, 50, 45, 55, 30, 65, 70]
        write_header_row(ws2, hdrs2, widths2, fill_hex="2F5496", border=border)

        detail_row   = 2
        merge_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merge_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        for paper_num, r in enumerate(results, 1):
            file_rows = [
                df for df in r.get("doc_files", [])
                if (df.get("path") or "").strip()
            ]
            if not file_rows:
                continue

            start_row = detail_row
            for df in file_rows:
                fpath = df["path"].strip()
                link  = df.get("link") or ""
                for col_idx, value in enumerate([fpath, df.get("type") or "", df.get("description") or "", link], 4):
                    cell = ws2.cell(row=detail_row, column=col_idx, value=value)
                    cell.font      = Font(name="Arial", size=10)
                    cell.border    = border
                    cell.alignment = wrap
                    if col_idx == 7 and value:
                        cell.hyperlink = value
                        cell.font = Font(name="Arial", size=10,
                                         color="0563C1", underline="single")
                ws2.row_dimensions[detail_row].height = auto_row_height(
                    [fpath, df.get("description") or ""], line_height=15 * 1.3,
                )
                detail_row += 1

            end_row = detail_row - 1
            ws2.cell(row=start_row, column=1, value=paper_num).font = Font(name="Arial", size=10)
            ws2.cell(row=start_row, column=2, value=r.get("title") or "").font = Font(name="Arial", size=10)
            ws2.cell(row=start_row, column=3, value=r.get("repo") or "").font = Font(name="Arial", size=10)
            for col_idx in (1, 2, 3):
                cell           = ws2.cell(row=start_row, column=col_idx)
                cell.border    = border
                cell.alignment = merge_center if col_idx == 1 else merge_left
            if end_row > start_row:
                ws2.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws2.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                ws2.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

        # =====================================================================
        # Skipped / Errors sheet
        # =====================================================================
        ws3     = wb.create_sheet(safe_sheet_name("Skipped", deployment))
        hdrs3   = ["#", "Status", "Title", "Repo / URL", "Note"]
        widths3 = [5, 10, 55, 55, 60]
        write_header_row(ws3, hdrs3, widths3, fill_hex="7B3F00", border=border)

        se_row = 2
        for paper_num, r in enumerate(results, 1):
            if r.get("status") not in ("skipped", "error"):
                continue
            fill_color = STATUS_FILL_COLORS.get(r.get("status"), "FFFFFF")
            row_fill   = PatternFill("solid", start_color=fill_color)
            row_vals   = [
                paper_num,
                (r.get("status") or "").upper(),
                r.get("title") or "",
                r.get("repo") or "",
                r.get("note") or "",
            ]
            for col_idx, value in enumerate(row_vals, 1):
                cell           = ws3.cell(row=se_row, column=col_idx, value=value)
                cell.font      = Font(name="Arial", size=10)
                cell.border    = border
                cell.alignment = center if col_idx <= 2 else wrap
                if col_idx == 2:
                    cell.fill = row_fill
            se_row += 1

        # =====================================================================
        # Summary sheet
        # =====================================================================
        counts      = count_statuses(results)
        yes         = counts.get("yes", 0)
        total_files = sum(len(r.get("doc_files", [])) for r in results)

        tool_counter = Counter()
        type_counter = Counter()
        for r in results:
            tool = r.get("doc_tool") or "none"
            if tool and tool != "none":
                tool_counter[tool] += 1
            for t in (r.get("documentation_types") or []):
                type_counter[t] += 1

        extra_summary = [
            ("Total Doc Files Found", total_files),
            ("Avg Doc Files per Repo (w/ any)",
             f"={total_files}/{yes if yes else 1}"),
            ("", ""),
            ("Doc Tool Breakdown", ""),
        ]
        for tool, cnt in tool_counter.most_common():
            extra_summary.append((f"  • {tool}", cnt))
        extra_summary.append(("", ""))
        extra_summary.append(("Documentation Type Breakdown", ""))
        for t, cnt in type_counter.most_common():
            extra_summary.append((f"  • {t}", cnt))

        ws4 = wb.create_sheet(safe_sheet_name("Summary", deployment))
        write_summary_sheet(
            ws4, results,
            positive_label="Repos with API Documentation",
            negative_label="Repos Missing API Documentation",
            extra_rows=extra_summary,
            token_usage=token_usage,
            deployment=deployment,
            fill_hex="2F5496",
            border=border,
        )

    # =========================================================================
    # Model Comparison sheet (only when multiple models)
    # =========================================================================
    if len(deployments) > 1:
        ws_cmp = wb.create_sheet("Model Comparison")
        write_comparison_sheet(ws_cmp, all_model_results,
                               fill_hex="2F5496", border=border)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    # ── Console summary ───────────────────────────────────────────────────────
    for deployment in deployments:
        results     = all_model_results[deployment]["results"]
        counts      = count_statuses(results)
        total       = len(results)
        skipped     = counts.get("skipped", 0)
        errors      = counts.get("error", 0)
        total_files = sum(len(r.get("doc_files", [])) for r in results)
        print(f"\n[{deployment}]")
        print(f"  → Results:    {total} papers")
        print(f"  → Doc Files:  {total_files} individual documentation files")
        print(f"  → Skipped/Err: {skipped + errors} entries")
    print(f"\nFull results saved to: {path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Run the API documentation analysis pipeline over all configured papers."""
    _content_cache.clear()

    def _format_check_extra(result):
        return f"{len(result.get('doc_files', []))} doc files"

    run_pipeline(
        papers=PAPERS,
        deployments=DEPLOYMENTS,
        make_check_paper_fn=make_check_paper_fn,
        print_results_fn=print_results,
        save_results_fn=save_results,
        description="API documentation",
        github_token=GITHUB_TOKEN,
        format_check_extra=_format_check_extra,
    )


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    main()