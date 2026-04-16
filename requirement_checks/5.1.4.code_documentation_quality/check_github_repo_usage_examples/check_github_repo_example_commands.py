"""Usage-examples checker for academic paper GitHub repositories.

Fetches README, docs, notebooks, and example scripts via the GitHub API,
then asks the LLM to list every usage example. Results export to Excel.
"""

import os
import sys
import time
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Add parent directories to Python path for local imports
_this = Path(__file__).resolve()
sys.path.insert(0, str(_this.parent.parent))          # 5.1.4.code_documentation_quality/
sys.path.insert(0, str(_this.parent.parent.parent))   # requirement_checks/

from common.fetch_and_parse_github_repo import (
    load_dotenv, parse_github_repo, list_all_repo_files, fetch_file_content,
)
from common.result_status import STATUS_FILL_COLORS, count_statuses
from common.repo_content_helpers import (
    fetch_paths_with_char_budget,
    path_priority_with_readme_first,
)
from common.llm_helpers import llm_call_parse_retry
from common.checker_pipeline import run_multi_model_pipeline
from common.excel_output import (
    write_header_row, write_results_data_rows, write_summary_sheet,
    write_comparison_sheet, safe_sheet_name,
    thin_border, auto_row_height, alignment_center, alignment_wrap_left,
)
from shared import check_paper_generic

load_dotenv()

from papers_from_database import PAPERS
from openai_client import client, AZURE_OPENAI_DEPLOYMENTS as DEPLOYMENTS
from prompts import SYSTEM_PROMPT
from config import (
    TARGET_FILENAMES, USAGE_EXTENSIONS, SCRIPT_EXTENSIONS,
    EXAMPLE_FOLDER_PREFIXES, MAX_CONTENT_CHARS, MAX_NOTEBOOKS, MAX_SCRIPTS,
)

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")


# =============================================================================
# GITHUB API HELPERS
# =============================================================================


def summarise_notebook(raw_json_text, max_chars=8_000):
    """Distil a Jupyter notebook JSON into compact plain text (code + markdown only)."""
    try:
        nb = json.loads(raw_json_text)
    except Exception:
        return raw_json_text[:max_chars]

    lines = []
    for cell in nb.get("cells", []):
        ct  = cell.get("cell_type", "")
        src = "".join(cell.get("source", []))
        if not src.strip():
            continue
        if ct == "code":
            lines.append(f"[CODE]\n{src}")
            for out in cell.get("outputs", []):
                text = "".join(out.get("text", []))
                if text and len(text) < 500:
                    lines.append(f"[OUTPUT]\n{text}")
        elif ct == "markdown":
            lines.append(f"[MARKDOWN]\n{src}")

    return "\n\n".join(lines)[:max_chars]


def collect_repo_content(owner, repo):
    """Fetch docs, notebooks, and example scripts; return combined text and metadata."""
    all_files = list_all_repo_files(owner, repo)
    if not all_files:
        return "", [], []

    docs_paths     = []   # named documentation files
    notebook_paths = []   # .ipynb / .Rmd / .qmd files
    script_paths   = []   # example/demo scripts

    for f in all_files:
        fpath     = f["path"]
        fpath_low = fpath.lower()
        basename  = fpath_low.split("/")[-1]
        ext       = "." + basename.rsplit(".", 1)[-1] if "." in basename else ""

        # Notebooks anywhere in the tree
        if ext in USAGE_EXTENSIONS:
            notebook_paths.append(fpath)

        # Named doc files
        elif fpath_low in TARGET_FILENAMES or basename in TARGET_FILENAMES:
            docs_paths.append(fpath)

        # Markdown inside special folders
        elif any(fpath_low.startswith(p) for p in EXAMPLE_FOLDER_PREFIXES):
            if ext in (".md", ".rst", ".txt"):
                docs_paths.append(fpath)
            elif ext in SCRIPT_EXTENSIONS:
                script_paths.append(fpath)

        # Script files whose name suggests they are examples
        elif ext in SCRIPT_EXTENSIONS and any(
            kw in basename for kw in
            ["example", "demo", "tutorial", "quickstart", "sample", "run_", "test_usage"]
        ):
            script_paths.append(fpath)

    # Sort docs: root README first, then other root-level files, then nested
    docs_paths.sort(key=path_priority_with_readme_first)

    # Sort notebooks and scripts: shallowest first (root-level ones are most
    # likely to be the "main" usage example)
    notebook_paths.sort(key=lambda p: (p.count("/"), p.lower()))
    script_paths.sort(key=lambda p: (p.count("/"), p.lower()))

    # Build the metadata list for ALL example-like files (even unfetched ones),
    # so we can produce GitHub links for every one of them later.
    all_example_meta = (
        [{"path": p, "kind": "notebook"} for p in notebook_paths] +
        [{"path": p, "kind": "script"}   for p in script_paths]
    )

    # Cap how many we actually fetch to stay within the character budget
    notebooks_to_fetch = notebook_paths[:MAX_NOTEBOOKS]
    scripts_to_fetch   = script_paths[:MAX_SCRIPTS]

    fetched     = []
    combined    = []
    total_chars = 0

    # ── 1. Fetch documentation files ─────────────────────────────────────────
    doc_blocks, fetched_docs, total_chars = fetch_paths_with_char_budget(
        owner,
        repo,
        docs_paths,
        MAX_CONTENT_CHARS,
        fetch_content=fetch_file_content,
        start_total_chars=total_chars,
        pause_seconds=0.15,
        header_label="FILE",
    )
    combined.extend(doc_blocks)
    fetched.extend(fetched_docs)

    # ── 2. Fetch and summarise notebooks ─────────────────────────────────────
    remaining     = MAX_CONTENT_CHARS - total_chars
    budget_per_nb = min(8_000, remaining // max(len(notebooks_to_fetch), 1))

    for nb_path in notebooks_to_fetch:
        if total_chars >= MAX_CONTENT_CHARS:
            break
        raw = fetch_file_content(owner, repo, nb_path)
        if raw:
            summary = summarise_notebook(raw, max_chars=budget_per_nb)
            combined.append(f"### NOTEBOOK: {nb_path}\n{summary}")
            total_chars += len(summary)
            fetched.append(nb_path)
        time.sleep(0.15)

    # ── 3. Fetch example scripts ──────────────────────────────────────────────
    script_blocks, fetched_scripts, total_chars = fetch_paths_with_char_budget(
        owner,
        repo,
        scripts_to_fetch,
        MAX_CONTENT_CHARS,
        fetch_content=fetch_file_content,
        start_total_chars=total_chars,
        pause_seconds=0.15,
        header_label="SCRIPT",
        per_file_char_cap=4_000,
    )
    combined.extend(script_blocks)
    fetched.extend(fetched_scripts)

    return "\n\n".join(combined), fetched, all_example_meta


# =============================================================================
# LLM CALL WRAPPER
# =============================================================================

EMPTY_USAGE_PAYLOAD = {
    "has_usage_examples": None,
    "evidence": "Empty LLM response after retry",
    "example_types": [],
    "example_files": [],
}

# Content cache — avoids re-fetching GitHub repos when running multiple models.
_content_cache = {}


def _cached_collect_repo_content(owner, repo):
    """Return cached collect_repo_content result, fetching only on first call."""
    key = f"{owner}/{repo}"
    if key not in _content_cache:
        _content_cache[key] = collect_repo_content(owner, repo)
    return _content_cache[key]


# =============================================================================
# PER-PAPER ORCHESTRATION
# =============================================================================

def build_example_entries(example_files_raw, owner, repo):
    """Enrich raw LLM example_files with GitHub hyperlinks."""
    entries = []
    for item in (example_files_raw or []):
        path = (item.get("path") or "").strip()
        if not path:
            continue
        cmds = item.get("commands", [])
        if isinstance(cmds, str):   # graceful fallback if LLM returns a string
            cmds = [cmds] if cmds else []
        entry = {
            "path":        path,
            "type":        item.get("type", "other"),
            "description": item.get("description", ""),
            "commands":    cmds,
            "link":        f"https://github.com/{owner}/{repo}/blob/HEAD/{path}",
        }
        entries.append(entry)
    return entries


def make_check_paper_fn(deployment, token_usage):
    """Factory: create a check_paper function bound to a specific model."""

    def _llm_check(repo_content, paper_title):
        def build_msg(content):
            return (
                f"Paper: {paper_title}\n\n"
                "Below are the contents of key files from its GitHub repository.\n"
                "Find ALL usage examples present.\n\n"
                + (content if content else "[No relevant files found in the repository]")
            )

        token_budget = 8000 if len(repo_content) > 50_000 else 4000

        result = llm_call_parse_retry(
            client=client,
            deployment=deployment,
            system_prompt=SYSTEM_PROMPT,
            build_user_message=build_msg,
            content=repo_content,
            token_usage=token_usage,
            empty_payload=EMPTY_USAGE_PAYLOAD,
            required_list_fields=("example_types", "example_files"),
            max_completion_tokens=token_budget,
            retry_truncate_chars=30_000,
            retry_max_tokens=16000,
            preview_chars=1500,
        )

        if "example_file" in result and not result.get("example_files"):
            ef = result.pop("example_file")
            result["example_files"] = (
                [{"path": ef, "type": "other", "description": ""}] if ef else []
            )
        else:
            result.pop("example_file", None)

        return result

    def _collect_wrapper(owner, repo, result):
        repo_content, files_checked, all_example_meta = _cached_collect_repo_content(owner, repo)
        result["all_example_meta"] = all_example_meta
        return repo_content, files_checked

    def _map_usage(result, llm_result, owner, repo):
        result["example_types"]   = llm_result.get("example_types", [])
        result["example_entries"] = build_example_entries(
            llm_result.get("example_files", []), owner, repo
        )

    def check_paper(paper):
        return check_paper_generic(
            paper,
            extra_defaults={
                "example_types":    [],
                "example_entries":  [],
                "all_example_meta": [],
            },
            collect_content_fn=_collect_wrapper,
            llm_check_fn=_llm_check,
            map_llm_result_fn=_map_usage,
            boolean_key="has_usage_examples",
        )

    return check_paper


# =============================================================================
# GROUND-TRUTH EVALUATION
# =============================================================================

def _normalise_ground_truth_label(value):
    """Map common ground-truth label formats to a boolean, or None if unknown."""
    if isinstance(value, bool):
        return value

    if value is None:
        return None

    text = str(value).strip().lower()
    positive = {
        "yes", "true", "1", "has_usage_examples", "has examples",
        "present", "exists", "with_examples", "with examples",
    }
    negative = {
        "no", "false", "0", "missing", "none", "absent",
        "without_examples", "without examples", "no examples",
    }

    if text in positive:
        return True
    if text in negative:
        return False
    return None


def evaluate_against_ground_truth(results):
    """Compute binary metrics when PAPERS includes a usable ground_truth field."""
    gt_by_title = {}
    for paper in PAPERS:
        gt = _normalise_ground_truth_label(paper.get("ground_truth"))
        if gt is not None:
            gt_by_title[paper.get("title", "")] = gt

    per_paper = []
    true_pos = false_pos = true_neg = false_neg = 0
    evaluated = 0

    for row in results:
        title = row.get("title", "")
        if title not in gt_by_title:
            continue

        gt_bool = gt_by_title[title]
        status = (row.get("status") or "").lower()
        if status == "yes":
            pred_bool = True
        elif status == "no":
            pred_bool = False
        else:
            pred_bool = None

        if pred_bool is None:
            match = "unknown"
            note = f"prediction unavailable (status={status or 'n/a'})"
            if row.get("note"):
                note += f"; {row['note']}"
        else:
            evaluated += 1
            if pred_bool and gt_bool:
                true_pos += 1
            elif pred_bool and not gt_bool:
                false_pos += 1
            elif not pred_bool and not gt_bool:
                true_neg += 1
            else:
                false_neg += 1

            match = "correct" if pred_bool == gt_bool else "wrong"
            note = row.get("note") or ""

        per_paper.append({
            "title": title,
            "prediction": "yes" if pred_bool is True else "no" if pred_bool is False else "unknown",
            "ground_truth": "yes" if gt_bool else "no",
            "match": match,
            "note": note,
        })

    labelled = len(per_paper)
    accuracy = ((true_pos + true_neg) / evaluated) if evaluated else 0.0
    precision = (true_pos / (true_pos + false_pos)) if (true_pos + false_pos) else 0.0
    recall = (true_pos / (true_pos + false_neg)) if (true_pos + false_neg) else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0

    return {
        "labelled": labelled,
        "evaluated": evaluated,
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1": f1,
        "true_pos": true_pos,
        "false_pos": false_pos,
        "true_neg": true_neg,
        "false_neg": false_neg,
        "per_paper": per_paper,
    }


def print_ground_truth_report(metrics):
    """Print a compact summary of ground-truth comparison metrics."""
    if metrics.get("labelled", 0) == 0:
        print("\nGround-truth evaluation skipped: no usable ground_truth labels in PAPERS.")
        return

    print("\n" + "=" * 78)
    print("GROUND-TRUTH REPORT")
    print("=" * 78)
    print(f"Labelled papers : {metrics['labelled']}")
    print(f"Evaluated       : {metrics['evaluated']}")
    print(f"Accuracy        : {metrics['accuracy']:.1%}")
    print(f"Precision       : {metrics['precision']:.1%}")
    print(f"Recall          : {metrics['recall']:.1%}")
    print(f"F1 score        : {metrics['f1']:.3f}")
    print(
        "Confusion matrix: "
        f"TP={metrics['true_pos']}, FP={metrics['false_pos']}, "
        f"TN={metrics['true_neg']}, FN={metrics['false_neg']}"
    )

# =============================================================================
# CONSOLE REPORTING
# =============================================================================

def print_results(results):
    """Print a compact console table of per-paper results with example counts."""
    W = 110
    print("\n" + "=" * W)
    print(f"{'#':<4} {'STATUS':<10} {'CONF':<8} {'# EX':<6} {'EXAMPLE TYPES':<30} TITLE")
    print("=" * W)

    for i, r in enumerate(results, 1):
        icon = {
            "yes": "YES", "no": "NO", "skipped": "SKIP", "error": "ERR",
        }.get(r["status"], "?")

        types_str = ", ".join(r.get("example_types", []))[:28]
        if not types_str and r.get("note"):
            types_str = r["note"][:28]

        n_ex        = len(r.get("example_entries", []))
        title_short = r["title"][:48] + ("..." if len(r["title"]) > 48 else "")
        conf        = r.get("confidence") or "-"
        print(f"{i:<4} {icon:<10} {conf:<8} {n_ex:<6} {types_str:<30} {title_short}")

        if r.get("evidence"):
            print(f"       evidence: {r['evidence']}")
        for ex in r.get("example_entries", []):
            print(f"       [{ex['type']}] {ex['path']}  →  {ex['description']}")

    print("=" * W)
    counts  = count_statuses(results)
    yes     = counts.get("yes", 0)
    no      = counts.get("no", 0)
    skipped = counts.get("skipped", 0)
    errors  = counts.get("error", 0)
    total_ex = sum(len(r.get("example_entries", [])) for r in results)
    print(
        f"\nSUMMARY: {yes} repos have examples | {no} missing | "
        f"{skipped} skipped | {errors} errors | {total_ex} total example files found\n"
    )


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def save_results(all_model_results, path=None):
    """Write per-model Excel sheets + Model Comparison sheet.

    *all_model_results* is ``{deployment: {"results": [...], "token_usage": tracker}}``.
    Per model: Results, All Examples, Skipped & Errors, Summary, Ground Truth.
    """
    if path is None:
        path = Path(__file__).resolve().parent / "results/usage_examples_results.xlsx"

    wb = Workbook()
    border = thin_border()
    center = alignment_center()
    wrap   = alignment_wrap_left()

    deployments = list(all_model_results.keys())
    first_sheet = True

    for deployment in deployments:
        model_data   = all_model_results[deployment]
        results      = model_data["results"]
        token_usage  = model_data["token_usage"]

        # =====================================================================
        # Results (one row per paper)
        # =====================================================================
        if first_sheet:
            ws1 = wb.active
            first_sheet = False
        else:
            ws1 = wb.create_sheet()
        ws1.title = safe_sheet_name("Results", deployment)

        hdrs1   = ["#", "Status", "Title", "Repo",
                   "# Examples", "Example Types", "Evidence",
                   "All Example Links", "Files Checked", "Note"]
        widths1 = [5, 10, 45, 40, 10, 35, 55, 60, 50, 35]
        write_header_row(ws1, hdrs1, widths1, fill_hex="2F5496", border=border)

        def results_row_data(r, num):
            all_links_text = "\n".join(
                f"{ex['path']}  ({ex['type']})"
                for ex in r.get("example_entries", [])
            )
            return [
                num,
                (r.get("status") or "").upper(),
                r.get("title") or "",
                r.get("repo") or "",
                len(r.get("example_entries", [])),
                ", ".join(r.get("example_types") or []),
                r.get("evidence") or "",
                all_links_text,
                ", ".join(r.get("files_checked") or []),
                r.get("note") or "",
            ]

        write_results_data_rows(
            ws1, results, results_row_data,
            border=border,
            row_height_fn=lambda vals: auto_row_height(vals, line_height=15 * 1.3),
        )

        # =====================================================================
        # All Examples (one row per individual example file)
        # =====================================================================
        ws2    = wb.create_sheet(safe_sheet_name("All Examples", deployment))
        hdrs2  = ["Paper #", "Paper Title", "Repo", "File Path",
                  "Type", "Description", "Example Command", "GitHub Link"]
        widths2 = [8, 50, 45, 55, 22, 60, 70, 70]
        write_header_row(ws2, hdrs2, widths2, fill_hex="2F5496", border=border)

        ex_row = 2
        for paper_num, r in enumerate(results, 1):
            for ex in r.get("example_entries", []):
                row_vals = [
                    paper_num,
                    r.get("title") or "",
                    r.get("repo") or "",
                    ex.get("path") or "",
                    ex.get("type") or "",
                    ex.get("description") or "",
                    "\n---\n".join(ex.get("commands") or []),
                    ex.get("link") or "",
                ]
                for col_idx, value in enumerate(row_vals, 1):
                    cell = ws2.cell(row=ex_row, column=col_idx, value=value)
                    cell.font      = Font(name="Arial", size=10)
                    cell.border    = border
                    cell.alignment = center if col_idx == 1 else wrap
                    if col_idx == 8 and value:
                        cell.hyperlink = value
                        cell.font = Font(name="Arial", size=10,
                                         color="0563C1", underline="single")
                ws2.row_dimensions[ex_row].height = auto_row_height(
                    row_vals, line_height=15 * 1.3,
                )
                ex_row += 1

        # =====================================================================
        # Skipped / Errors
        # =====================================================================
        ws3    = wb.create_sheet(safe_sheet_name("Skipped", deployment))
        hdrs3  = ["#", "Status", "Title", "Repo / URL", "Note"]
        widths3 = [5, 10, 55, 55, 60]
        write_header_row(ws3, hdrs3, widths3, fill_hex="7B3F00", border=border)

        se_row = 2
        for paper_num, r in enumerate(results, 1):
            if r.get("status") not in ("skipped", "error"):
                continue
            fill_color = STATUS_FILL_COLORS.get(r.get("status"), "FFFFFF")
            row_fill   = PatternFill("solid", start_color=fill_color)
            row_vals   = [
                paper_num,
                (r.get("status") or "").upper(),
                r.get("title") or "",
                r.get("repo") or "",
                r.get("note") or "",
            ]
            for col_idx, value in enumerate(row_vals, 1):
                cell           = ws3.cell(row=se_row, column=col_idx, value=value)
                cell.font      = Font(name="Arial", size=10)
                cell.border    = border
                cell.alignment = center if col_idx <= 2 else wrap
                if col_idx == 2:
                    cell.fill = row_fill
            se_row += 1

        # =====================================================================
        # Summary (uses shared helper)
        # =====================================================================
        counts   = count_statuses(results)
        yes      = counts.get("yes", 0)
        total_ex = sum(len(r.get("example_entries", [])) for r in results)

        type_counter = Counter()
        for r in results:
            for t in (r.get("example_types") or []):
                type_counter[t] += 1

        extra_summary = [
            ("Total Example Files Found", total_ex),
            ("Avg Examples per Repo (w/ any)",
             f"={total_ex}/{yes if yes else 1}"),
            ("", ""),
            ("Example Type Breakdown", ""),
        ]
        for t, cnt in type_counter.most_common():
            extra_summary.append((f"  • {t}", cnt))

        ws4 = wb.create_sheet(safe_sheet_name("Summary", deployment))
        write_summary_sheet(
            ws4, results,
            positive_label="Repos with Usage Examples",
            negative_label="Repos Missing Usage Examples",
            extra_rows=extra_summary,
            token_usage=token_usage,
            deployment=deployment,
            fill_hex="2F5496",
            border=border,
        )

        # =====================================================================
        # Ground Truth Comparison (only when labels exist)
        # =====================================================================
        metrics = evaluate_against_ground_truth(results)
        if metrics["labelled"] > 0:
            ws5    = wb.create_sheet(safe_sheet_name("Ground Truth", deployment))
            hdrs5  = ["#", "Paper Title", "Prediction", "Ground Truth", "Match", "Note"]
            widths5 = [5, 60, 14, 14, 10, 55]
            write_header_row(ws5, hdrs5, widths5, fill_hex="2E7D32", border=border)

            MATCH_COLORS = {"correct": "C6EFCE", "wrong": "FFC7CE"}
            for r_idx, p in enumerate(metrics["per_paper"], 2):
                row_vals = [
                    r_idx - 1, p["title"], p["prediction"],
                    p["ground_truth"], p["match"], p["note"],
                ]
                gt_fill_hex = MATCH_COLORS.get(p["match"], "FFFFFF")
                row_fill = PatternFill("solid", start_color=gt_fill_hex)
                for col_idx, value in enumerate(row_vals, 1):
                    cell           = ws5.cell(row=r_idx, column=col_idx, value=value)
                    cell.font      = Font(name="Arial", size=10)
                    cell.border    = border
                    cell.alignment = center if col_idx in (1, 3, 4, 5) else wrap
                    if col_idx == 5:
                        cell.fill = row_fill

            summary_start = len(metrics["per_paper"]) + 3
            summary_pairs = [
                ("Labelled papers",  metrics["labelled"]),
                ("Accuracy",         f"{metrics['accuracy']:.1%}"),
                ("Precision",        f"{metrics['precision']:.1%}"),
                ("Recall",           f"{metrics['recall']:.1%}"),
                ("F1 score",         f"{metrics['f1']:.3f}"),
                ("True positives",   metrics["true_pos"]),
                ("False positives",  metrics["false_pos"]),
                ("True negatives",   metrics["true_neg"]),
                ("False negatives",  metrics["false_neg"]),
            ]
            for i, (label, value) in enumerate(summary_pairs, summary_start):
                ws5.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
                ws5.cell(row=i, column=2, value=value).font = Font(name="Arial", size=10)
                ws5.cell(row=i, column=2).alignment = center

    # =========================================================================
    # Model Comparison sheet (only when multiple models)
    # =========================================================================
    if len(deployments) > 1:
        ws_cmp = wb.create_sheet("Model Comparison")
        write_comparison_sheet(ws_cmp, all_model_results,
                               fill_hex="2F5496", border=border)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    # ── Console summary ──────────────────────────────────────────────────
    for deployment in deployments:
        results  = all_model_results[deployment]["results"]
        counts   = count_statuses(results)
        total    = len(results)
        skipped  = counts.get("skipped", 0)
        errors   = counts.get("error", 0)
        total_ex = sum(len(r.get("example_entries", [])) for r in results)
        metrics  = evaluate_against_ground_truth(results)
        print(f"\n[{deployment}]")
        print(f"  → Results:       {total} papers")
        print(f"  → All Examples:  {total_ex} individual example files")
        print(f"  → Skipped/Err:   {skipped + errors} entries")
        if metrics["labelled"] > 0:
            print(f"  → Ground Truth:  {metrics['labelled']} labelled, accuracy {metrics['accuracy']:.1%}")
    print(f"\nFull results saved to: {path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Run the analysis pipeline over all configured papers."""
    _content_cache.clear()

    def _format_check_extra(result):
        return f"{len(result.get('example_entries', []))} examples"

    def _finalize(results):
        gt_metrics = evaluate_against_ground_truth(results)
        print_ground_truth_report(gt_metrics)

    run_multi_model_pipeline(
        papers=PAPERS,
        deployments=DEPLOYMENTS,
        make_check_paper_fn=make_check_paper_fn,
        print_results_fn=print_results,
        save_results_fn=save_results,
        description="usage examples",
        github_token=GITHUB_TOKEN,
        format_check_extra=_format_check_extra,
        finalize_fn=_finalize,
    )


if __name__ == "__main__":
    main()