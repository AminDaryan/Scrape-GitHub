"""
Inline-comments checker for academic paper GitHub repositories.

Logic:
  1. List every file in the repo via the GitHub API.                  → collect_repo_content()
  2. Keep only files with a recognized source extension (.py, .java,
     .cpp, .js, .r, .cu, etc.) and skip folders that contain
     third-party or generated code (node_modules/, vendor/,
     build/, .github/, tests/, __pycache__/, etc.).                  → _is_source_file()
  3. Sort files by path depth (fewest slashes first) so files closest
     to the repo root — more likely to be the main authored code —
     are fetched first; hard-cap at 25 files (MAX_SOURCE_FILES) so
     large repos don't trigger hundreds of individual GitHub API calls.
                                                                      → collect_repo_content()
  4. Fetch the selected files, stopping once the total character
     budget is reached; a per-file cap prevents one large file from
     crowding out the rest of the LLM context.                        → collect_repo_content()
  5. Pass the concatenated file contents to the LLM, asking whether
     meaningful inline comments are present, which types they are,
     and which files contain them.                                     → make_check_paper_fn()
  6. Run across all configured models and export to Excel.            → main(), save_results()

Results export to Excel.
"""

import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Add parent directories to Python path for local imports
_this = Path(__file__).resolve()
sys.path.insert(0, str(_this.parent.parent))          # 5.1.4.code_documentation_quality/
sys.path.insert(0, str(_this.parent.parent.parent))   # requirement_checks/

from common.fetch_and_parse_github_repo import (
    load_dotenv, parse_github_repo, list_all_repo_files, fetch_file_content,
)
from common.repo_content_helpers import fetch_paths_with_char_budget
from common.llm_helpers import llm_call_parse_retry
from common.checker_pipeline import run_pipeline
from common.excel_output import (
    count_statuses,
    write_header_row, write_results_data_rows, write_summary_sheet,
    write_comparison_sheet, safe_sheet_name,
    thin_border, auto_row_height, alignment_center, alignment_wrap_left,
)
from shared import check_paper_generic

load_dotenv()

from papers_from_database import PAPERS
from openai_client import client, AZURE_OPENAI_DEPLOYMENTS
from prompts import SYSTEM_PROMPT
from config import (
    SOURCE_CODE_EXTENSIONS, SKIP_FOLDER_PREFIXES, SKIP_BASENAMES,
    MAX_SOURCE_FILES, MAX_CONTENT_CHARS, PER_FILE_CHAR_CAP,
)

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")


# =============================================================================
# GITHUB API HELPERS
# =============================================================================


def _is_source_file(path_lower, basename):
    """Return True if the file is a source code file worth evaluating for comments.

    Excludes vendored dependencies, auto-generated files, and test fixtures
    (configured in SKIP_FOLDER_PREFIXES / SKIP_BASENAMES) because comments in
    those are not authored by the paper's developers and would skew the result.
    """
    ext = "." + basename.rsplit(".", 1)[-1] if "." in basename else ""
    if ext not in SOURCE_CODE_EXTENSIONS:
        return False
    if any(path_lower.startswith(prefix) for prefix in SKIP_FOLDER_PREFIXES):
        return False
    if basename in SKIP_BASENAMES:
        return False
    return True


def collect_repo_content(owner, repo):
    """Fetch source code files from a GitHub repo for inline-comment analysis."""
    all_files = list_all_repo_files(owner, repo)
    if not all_files:
        return "", []

    source_paths = []
    for f in all_files:
        fpath = f["path"]
        fpath_lower = fpath.lower()
        basename = fpath_lower.split("/")[-1]
        if _is_source_file(fpath_lower, basename):
            source_paths.append(fpath)

    if not source_paths:
        return "", []

    # Sort by path depth (fewest "/" first), then alphabetically within each depth level.
    source_paths.sort(key=lambda p: (p.count("/"), p.lower()))

    # Cap to avoid fetching too many files.
    source_paths = source_paths[:MAX_SOURCE_FILES]

    blocks, fetched, _ = fetch_paths_with_char_budget(
        owner,
        repo,
        source_paths,
        MAX_CONTENT_CHARS,
        fetch_content=fetch_file_content,
        pause_seconds=0.15,
        header_label="SOURCE FILE",
        per_file_char_cap=PER_FILE_CHAR_CAP,
    )

    return "\n\n".join(blocks), fetched


# =============================================================================
# LLM CALL WRAPPER
# =============================================================================

EMPTY_INLINE_COMMENTS_PAYLOAD = {
    "has_inline_comments": None,
    "evidence": "Empty LLM response after retry — content may exceed context window",
    "comment_types": [],
    "files_with_comments": [],
}

# Content cache — avoids re-fetching GitHub repos when running multiple models.
_content_cache = {}


def _cached_collect_repo_content(owner, repo):
    """Return cached collect_repo_content result, fetching only on first call."""
    key = f"{owner}/{repo}"
    if key not in _content_cache:
        _content_cache[key] = collect_repo_content(owner, repo)
    return _content_cache[key]


# =============================================================================
# PER-PAPER ORCHESTRATION
# =============================================================================

def _map_inline_comments(result, llm_result, owner, repo):
    """Copy inline-comment-specific fields from LLM result into the paper result."""
    result["comment_types"]       = llm_result.get("comment_types", [])
    result["files_with_comments"] = llm_result.get("files_with_comments", [])


def make_check_paper_fn(deployment, token_usage):
    """Factory: return a check_paper function bound to a specific LLM deployment.

    Using a factory instead of a plain function lets run_pipeline swap in different
    models without re-fetching GitHub content — the content cache is shared across
    all model instances for the same repo.
    """

    def _llm_check(repo_content, paper_title):
        """Call the LLM and parse the inline-comments verdict for one paper."""
        def build_msg(content):
            return (
                f"Paper: {paper_title}\n\n"
                "Below are the contents of source code files fetched from its GitHub repository.\n"
                "Evaluate whether the code contains meaningful inline comments.\n\n"
                + (content if content else "[No source code files found in the repository]")
            )

        # Large repos need more output tokens so the LLM can list all commented files.
        token_budget = 4000 if len(repo_content) > 50_000 else 2000

        return llm_call_parse_retry(
            client=client,
            deployment=deployment,
            system_prompt=SYSTEM_PROMPT,
            build_user_message=build_msg,
            content=repo_content,
            token_usage=token_usage,
            empty_payload=EMPTY_INLINE_COMMENTS_PAYLOAD,
            required_list_fields=("comment_types", "files_with_comments"),
            max_completion_tokens=token_budget,
            retry_truncate_chars=30_000,
            retry_max_tokens=8000,
            preview_chars=1000,
        )

    def check_paper(paper):
        return check_paper_generic(
            paper,
            extra_defaults={
                "comment_types":       [],
                "files_with_comments": [],
            },
            collect_content_fn=_cached_collect_repo_content,
            llm_check_fn=_llm_check,
            map_llm_result_fn=_map_inline_comments,
            boolean_key="has_inline_comments",
            require_files=True,
            no_files_message="No source code files found in the repository",
        )

    return check_paper


# =============================================================================
# CONSOLE REPORTING
# =============================================================================

def print_results(results):
    """Print a compact console table of per-paper results."""
    W = 110
    print("\n" + "=" * W)
    print(f"{'#':<4} {'STATUS':<10} {'COMMENT TYPES':<36} TITLE")
    print("=" * W)

    for i, r in enumerate(results, 1):
        icon = {
            "yes": "YES", "no": "NO", "skipped": "SKIP", "error": "ERR",
        }.get(r["status"], "?")

        types_str = ", ".join(r.get("comment_types", []))[:34]
        if not types_str and r.get("note"):
            types_str = r["note"][:34]

        title_short = r["title"][:46] + ("..." if len(r["title"]) > 46 else "")
        print(f"{i:<4} {icon:<10} {types_str:<36} {title_short}")

        if r.get("evidence"):
            print(f"       evidence: {r['evidence']}")
        if r.get("files_checked"):
            print(f"       files: {', '.join(r['files_checked'][:5])}"
                  + (f" (+{len(r['files_checked'])-5} more)" if len(r["files_checked"]) > 5 else ""))

    print("=" * W)
    counts  = count_statuses(results)
    yes     = counts.get("yes", 0)
    no      = counts.get("no", 0)
    skipped = counts.get("skipped", 0)
    errors  = counts.get("error", 0)
    print(
        f"\nSUMMARY: {yes} repos have inline comments | "
        f"{no} missing | {skipped} skipped (non-GitHub) | {errors} errors\n"
    )


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def save_results(all_model_results, path=None):
    """Write per-model Results + Detail + Summary sheets, and a Model Comparison sheet."""
    if path is None:
        path = Path(__file__).resolve().parent / "results/inline_comments_results.xlsx"

    wb = Workbook()
    border = thin_border()
    center = alignment_center()
    wrap = alignment_wrap_left()

    deployments = list(all_model_results.keys())
    first_sheet = True

    for deployment in deployments:
        model_data = all_model_results[deployment]
        results = model_data["results"]
        token_usage = model_data["token_usage"]

        # ── Results sheet (per model) ────────────────────────────────────
        if first_sheet:
            ws1 = wb.active
            first_sheet = False
        else:
            ws1 = wb.create_sheet()
        ws1.title = safe_sheet_name("Results", deployment)

        hdrs1 = ["#", "Status", "Title", "Repo",
                 "Comment Types", "Evidence", "Files Checked", "Note",
                 "Tokens Used"]
        widths1 = [5, 10, 45, 40, 35, 55, 50, 35, 12]
        write_header_row(ws1, hdrs1, widths1, fill_hex="2F5496", border=border)

        def results_row_data(r, num):
            return [
                num,
                (r.get("status") or "").upper(),
                r.get("title") or "",
                r.get("repo") or "",
                ", ".join(r.get("comment_types") or []),
                r.get("evidence") or "",
                ", ".join(r.get("files_checked") or []),
                r.get("note") or "",
                r.get("tokens_used") or "",
            ]

        write_results_data_rows(
            ws1, results, results_row_data,
            border=border,
            row_height_fn=lambda vals: auto_row_height(vals),
        )

        # ── Per-File Detail sheet (per model) ────────────────────────────
        ws2 = wb.create_sheet(safe_sheet_name("Detail", deployment))
        hdrs2 = ["Paper #", "Paper Title", "Repo", "File Path",
                 "Description", "GitHub Link"]
        widths2 = [8, 50, 45, 55, 60, 70]
        write_header_row(ws2, hdrs2, widths2, fill_hex="2F5496", border=border)

        detail_row = 2
        merge_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merge_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for paper_num, r in enumerate(results, 1):
            owner, repo_name = parse_github_repo(r.get("repo", ""))
            file_rows = [
                fc for fc in r.get("files_with_comments", [])
                if (fc.get("path") or "").strip()
            ]
            if not file_rows:
                continue

            start_row = detail_row
            for fc in file_rows:
                fpath = fc["path"].strip()
                link = (
                    f"https://github.com/{owner}/{repo_name}/blob/HEAD/{fpath}"
                    if owner else ""
                )
                for col_idx, value in enumerate([fpath, fc.get("description") or "", link], 4):
                    cell = ws2.cell(row=detail_row, column=col_idx, value=value)
                    cell.font = Font(name="Arial", size=10)
                    cell.border = border
                    cell.alignment = wrap
                    if col_idx == 6 and value:
                        cell.hyperlink = value
                        cell.font = Font(name="Arial", size=10,
                                         color="0563C1", underline="single")
                ws2.row_dimensions[detail_row].height = auto_row_height(
                    [fpath, fc.get("description") or "", link], line_height=15 * 1.3,
                )
                detail_row += 1

            end_row = detail_row - 1
            ws2.cell(row=start_row, column=1, value=paper_num).font = Font(name="Arial", size=10)
            ws2.cell(row=start_row, column=2, value=r.get("title") or "").font = Font(name="Arial", size=10)
            ws2.cell(row=start_row, column=3, value=r.get("repo") or "").font = Font(name="Arial", size=10)
            for col_idx in (1, 2, 3):
                cell = ws2.cell(row=start_row, column=col_idx)
                cell.border = border
                cell.alignment = merge_center if col_idx == 1 else merge_left
            if end_row > start_row:
                ws2.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws2.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                ws2.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

        # ── Summary sheet (per model) ────────────────────────────────────
        ws3 = wb.create_sheet(safe_sheet_name("Summary", deployment))
        write_summary_sheet(
            ws3, results,
            positive_label="Have Inline Comments",
            negative_label="Missing Inline Comments",
            token_usage=token_usage,
            deployment=deployment,
            fill_hex="2F5496",
            border=border,
        )

    # ── Model Comparison sheet ───────────────────────────────────────────
    if len(deployments) > 1:
        ws_cmp = wb.create_sheet("Model Comparison")
        write_comparison_sheet(
            ws_cmp, all_model_results, fill_hex="2F5496", border=border,
        )

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Full results saved to {path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Entry point: run the inline-comments analysis pipeline."""
    _content_cache.clear()
    run_pipeline(
        papers=PAPERS,
        deployments=AZURE_OPENAI_DEPLOYMENTS,
        make_check_paper_fn=make_check_paper_fn,
        print_results_fn=print_results,
        save_results_fn=save_results,
        description="inline comments",
        github_token=GITHUB_TOKEN,
    )


if __name__ == "__main__":
    main()
