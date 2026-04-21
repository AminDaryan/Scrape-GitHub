"""
Installation-instructions checker for academic paper GitHub repositories.

Logic:
  1. List every file in the repo via the GitHub API.                  → collect_repo_content()
  2. Select setup-relevant files by name (README, requirements.txt,
     setup.py, Dockerfile, environment.yml, etc.); fall back to the
     root README if nothing named matches.                            → collect_repo_content()
  3. Sort selected files: root README first, then other setup files,
     then nested READMEs last — so content truncation preferentially
     keeps the most informative file intact.                          → collect_repo_content()
  4. Fetch file contents up to a total character limit.               → collect_repo_content()
  5. Pass the concatenated documentation to the LLM, asking whether
     installation instructions are present, what types they cover,
     and which file contains them.                                     → make_check_paper_fn()
  6. Run across all configured models and export to Excel.            → main(), save_results()

Results export to Excel.
"""

import os
import sys
from pathlib import Path

# Add parent directories to Python path for local imports
_this = Path(__file__).resolve()
sys.path.insert(0, str(_this.parent.parent))          # 5.1.4.code_documentation_quality/
sys.path.insert(0, str(_this.parent.parent.parent))   # requirement_checks/

from openpyxl import Workbook
from common.fetch_and_parse_github_repo import (
    load_dotenv, list_all_repo_files, fetch_file_content,
)
from common.repo_content_helpers import (
    fetch_paths_with_char_budget,
    first_readme_path,
    path_priority_with_readme_first,
)
from common.llm_helpers import llm_call_parse_retry
from common.checker_pipeline import run_pipeline
from common.excel_output import (
    count_statuses,
    write_header_row, write_results_data_rows, write_summary_sheet,
    write_comparison_sheet, safe_sheet_name,
    thin_border, auto_row_height,
)
from shared import check_paper_generic

load_dotenv()

from papers_from_database import PAPERS
from openai_client import client, AZURE_OPENAI_DEPLOYMENTS
from prompts import SYSTEM_PROMPT
from config import TARGET_FILENAMES, MAX_CONTENT_CHARS

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")

# ─── GitHub API helpers ───────────────────────────────────────────────────────


def collect_repo_content(owner, repo):
    """Fetch and concatenate setup-relevant files from a GitHub repo.

    Targets files that commonly contain installation instructions (TARGET_FILENAMES).
    If none are found, falls back to just the root README so the LLM always gets
    at least some documentation to evaluate rather than returning empty-handed.
    """
    all_files = list_all_repo_files(owner, repo)
    if not all_files:
        return "", []

    readme_fallback = first_readme_path(all_files)
    selected_paths = []
    for f in all_files:
        fpath_lower = f["path"].lower()
        basename = fpath_lower.split("/")[-1]
        if fpath_lower in TARGET_FILENAMES or basename in TARGET_FILENAMES:
            selected_paths.append(f["path"])

    # Fallback: grab just the root README if nothing matched
    if not selected_paths and readme_fallback:
        selected_paths.append(readme_fallback)

    # Always put the root-level README first so it gets the most budget,
    # then setup files, then nested READMEs last
    selected_paths.sort(
        key=lambda p: path_priority_with_readme_first(p, nested_readme_last=True)
    )

    blocks, fetched, _ = fetch_paths_with_char_budget(
        owner,
        repo,
        selected_paths,
        MAX_CONTENT_CHARS,
        fetch_content=fetch_file_content,
        pause_seconds=0.2,
        header_label="FILE",
    )

    return "\n\n".join(blocks), fetched


# ─── LLM analysis ─────────────────────────────────────────────────────────────
# Prompt templates are defined in neighboring prompts.py.


EMPTY_INSTALLATION_PAYLOAD = {
    "has_installation_instructions": None,
    "evidence": "Empty LLM response after retry — content may exceed context window",
    "instruction_types": [],
}

# Content cache — avoids re-fetching GitHub repos when running multiple models.
_content_cache = {}


def _cached_collect_repo_content(owner, repo):
    """Return cached collect_repo_content result, fetching only on first call."""
    key = f"{owner}/{repo}"
    if key not in _content_cache:
        _content_cache[key] = collect_repo_content(owner, repo)
    return _content_cache[key]


# ─── Per-paper orchestration ──────────────────────────────────────────────────

def _map_installation(result, llm_result, owner, repo):
    """Copy installation-specific fields from the LLM result into the paper result.

    Also builds a direct GitHub link to the installation file when the LLM
    identifies one, so auditors can navigate straight to the relevant file.
    """
    result["instruction_types"] = llm_result.get("instruction_types", [])
    install_file = llm_result.get("installation_file")
    if install_file and owner:
        result["installation_link"] = f"https://github.com/{owner}/{repo}/blob/HEAD/{install_file}"


def make_check_paper_fn(deployment, token_usage):
    """Factory: return a check_paper function bound to a specific LLM deployment.

    Using a factory instead of a plain function lets run_pipeline swap in different
    models without re-fetching GitHub content — the content cache is shared across
    all model instances for the same repo.
    """

    def _llm_check(repo_content, paper_title):
        """Call the LLM and parse the installation-instructions verdict for one paper."""
        def build_msg(content):
            return (
                f"Paper: {paper_title}\n\n"
                "Below are the contents of key files fetched from its GitHub repository.\n"
                "Does this repository contain installation instructions?\n\n"
                + (content if content else "[No relevant files found in the repository]")
            )

        # 1000 tokens is enough — the answer is binary yes/no plus a short evidence list.
        return llm_call_parse_retry(
            client=client,
            deployment=deployment,
            system_prompt=SYSTEM_PROMPT,
            build_user_message=build_msg,
            content=repo_content,
            token_usage=token_usage,
            empty_payload=EMPTY_INSTALLATION_PAYLOAD,
            required_list_fields=("instruction_types",),
            max_completion_tokens=1000,
            retry_truncate_chars=15_000,
        )

    def check_paper(paper):
        return check_paper_generic(
            paper,
            extra_defaults={
                "instruction_types": [],
                "installation_link": "",
            },
            collect_content_fn=_cached_collect_repo_content,
            llm_check_fn=_llm_check,
            map_llm_result_fn=_map_installation,
            boolean_key="has_installation_instructions",
        )

    return check_paper


# ─── Console reporting ────────────────────────────────────────────────────────


def print_results(results):
    """Print a compact console table of per-paper results."""
    W = 110
    print("\n" + "=" * W)
    print(f"{'#':<4} {'STATUS':<10} {'INSTRUCTION TYPES':<36} TITLE")
    print("=" * W)

    for i, r in enumerate(results, 1):
        icon = {
            "yes":     "YES",
            "no":      "NO",
            "skipped": "SKIP",
            "error":   "ERR",
        }.get(r["status"], "?")

        types_str = ", ".join(r.get("instruction_types", []))[:34]
        if not types_str and r.get("note"):
            types_str = r["note"][:34]

        title_short = r["title"][:52] + ("..." if len(r["title"]) > 52 else "")
        print(f"{i:<4} {icon:<10} {types_str:<36} {title_short}")

        if r.get("evidence"):
            print(f"       -> {r['evidence']}")
        if r.get("files_checked"):
            print(f"       files: {', '.join(r['files_checked'])}")

    print("=" * W)
    counts  = count_statuses(results)
    yes     = counts.get("yes", 0)
    no      = counts.get("no", 0)
    skipped = counts.get("skipped", 0)
    errors  = counts.get("error", 0)
    print(
        f"\nSUMMARY: {yes} have install instructions | "
        f"{no} missing | {skipped} skipped (non-GitHub) | {errors} errors\n"
    )



def save_results(all_model_results, path=None):
    """Export per-model results and summary sheets, plus a comparison sheet."""
    if path is None:
        path = Path(__file__).resolve().parent / "results/installation_instructions_results.xlsx"

    wb = Workbook()
    border = thin_border()

    deployments = list(all_model_results.keys())
    first_sheet = True

    for deployment in deployments:
        model_data = all_model_results[deployment]
        results = model_data["results"]
        token_usage = model_data["token_usage"]

        # ── Results sheet (per model) ────────────────────────────────────
        if first_sheet:
            ws = wb.active
            first_sheet = False
        else:
            ws = wb.create_sheet()
        ws.title = safe_sheet_name("Results", deployment)

        headers = ["#", "Status", "Title", "Repo",
                   "Instruction Types", "Evidence", "Installation Link",
                   "Files Checked", "Note"]
        col_widths = [5, 10, 45, 40, 35, 50, 45, 45, 30]
        write_header_row(ws, headers, col_widths, fill_hex="2F5496", border=border)

        def row_data_fn(r, num):
            return [
                num,
                (r.get("status") or "").upper(),
                r.get("title") or "",
                r.get("repo") or "",
                ", ".join(r.get("instruction_types") or []),
                r.get("evidence") or "",
                r.get("installation_link") or "",
                ", ".join(r.get("files_checked") or []),
                r.get("note") or "",
            ]

        write_results_data_rows(
            ws, results, row_data_fn,
            border=border,
            link_cols={7},
            row_height_fn=lambda vals: auto_row_height(vals),
        )

        # ── Summary sheet (per model) ────────────────────────────────────
        ws2 = wb.create_sheet(safe_sheet_name("Summary", deployment))
        write_summary_sheet(
            ws2, results,
            positive_label="Have Installation Instructions",
            negative_label="Missing Installation Instructions",
            token_usage=token_usage,
            deployment=deployment,
            fill_hex="2F5496",
            border=border,
        )

    # ── Model Comparison sheet ───────────────────────────────────────────
    if len(deployments) > 1:
        ws_cmp = wb.create_sheet("Model Comparison")
        write_comparison_sheet(
            ws_cmp, all_model_results, fill_hex="2F5496", border=border,
        )

    wb.save(path)
    print(f"Full results saved to {path}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    """Entry point: run the analysis pipeline."""
    _content_cache.clear()
    run_pipeline(
        papers=PAPERS,
        deployments=AZURE_OPENAI_DEPLOYMENTS,
        make_check_paper_fn=make_check_paper_fn,
        print_results_fn=print_results,
        save_results_fn=save_results,
        description="installation instructions",
        github_token=GITHUB_TOKEN,
    )


if __name__ == "__main__":
    main()