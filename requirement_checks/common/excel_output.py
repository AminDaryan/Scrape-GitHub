"""Shared Excel output helpers for requirement-check scripts.

Also contains the four paper statuses (yes/no/skipped/error), their Excel row
colors, and helpers for counting them and computing coverage.
"""

from pathlib import Path
from typing import Any, Dict, Mapping, Sequence

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Status definitions (merged from result_status.py)
# ---------------------------------------------------------------------------

# Each paper we check gets one of these four statuses.
# The hex values are the background colors used for that row in the Excel output.
STATUS_FILL_COLORS = {
    "yes": "C6EFCE",      # green  — the repo HAS the feature (e.g. has inline comments)
    "no": "FFDDC1",       # orange — the repo does NOT have the feature
    "skipped": "FFEB9C",  # yellow — not a GitHub repo, so we could not check
    "error": "FFC7CE",    # red    — something went wrong during checking
}


def count_statuses(
    results: Sequence[Mapping[str, Any]],
    status_key: str = "status",
) -> Dict[str, int]:
    """Go through all paper results and count how many fall into each status.

    Statuses:
        "yes"     — the repo has the checked feature (e.g. has inline comments)
        "no"      — the repo was checked but the feature is missing
        "skipped" — the repo URL was not a GitHub link, so it was not checked
        "error"   — the check failed (e.g. GitHub API error, LLM timeout)

    Input:  a list of result dicts, each having a "status" key.
    Output: {"yes": 5, "no": 2, "skipped": 1, "error": 0}
            (all four keys always present, even when the count is 0)
    """
    counts: Dict[str, int] = {k: 0 for k in STATUS_FILL_COLORS}

    for row in results:
        status = str(row.get(status_key, "")).strip().lower()
        if not status:
            continue
        counts[status] = counts.get(status, 0) + 1

    return counts


def coverage_formula(positive_count: int, total_count: int) -> str:
    """Calculate what percentage of repos have the feature, as an Excel formula.

    positive_count: number of repos with status "yes" (have the feature).
    total_count:    total number of repos checked.

    Returns a string like "=5/10*100" that gets placed in an Excel cell.
    Excel evaluates it and shows the result (50 in this case, meaning 50%).
    """
    denom = total_count if total_count else 1
    return f"={positive_count}/{denom}*100"


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------


def default_results_path(script_file, filename):
    """Return ``<script_dir>/results/<filename>`` as a Path."""
    return Path(script_file).resolve().parent / "results" / filename


def save_workbook(workbook, output_path):
    """Save a workbook, creating parent directories if needed."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def thin_border(color="CCCCCC"):
    """Create a thin border on all four sides."""
    edge = Side(style="thin", color=color)
    return Border(left=edge, right=edge, top=edge, bottom=edge)


def alignment_center(wrap_text=True):
    """Centered alignment preset."""
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)


def alignment_wrap_left():
    """Left-aligned, top-anchored, wrap-text alignment."""
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_cell(
    ws,
    row,
    col,
    value,
    *,
    fill_hex="2F5496",
    border=None,
    alignment=None,
    font_name="Arial",
    font_size=11,
    font_color="FFFFFF",
):
    """Write and style a single header cell. Returns the cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=font_name, bold=True, size=font_size, color=font_color)
    cell.fill = PatternFill("solid", start_color=fill_hex)
    cell.alignment = alignment or alignment_center(wrap_text=True)
    if border is not None:
        cell.border = border
    return cell


def write_header_row(
    ws,
    headers,
    widths,
    *,
    row=1,
    fill_hex="2F5496",
    border=None,
    height=20,
    font_name="Arial",
    font_size=11,
    font_color="FFFFFF",
):
    """Write a styled header row and set column widths."""
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        style_header_cell(
            ws,
            row,
            col_idx,
            header,
            fill_hex=fill_hex,
            border=border,
            font_name=font_name,
            font_size=font_size,
            font_color=font_color,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = height


def estimate_wrapped_lines(value, col_width=60):
    """Heuristic: how many visual lines a cell value will occupy."""
    text = "" if value is None else str(value)
    lines = text.split("\n")
    return sum(max(1, len(line) // col_width + 1) for line in lines)


def auto_row_height(values, col_width=60, min_height=20, line_height=19.5):
    """Estimate row height from a list of cell values."""
    if not values:
        return min_height
    max_lines = max(estimate_wrapped_lines(v, col_width=col_width) for v in values)
    return max(min_height, max_lines * line_height)


# ── Results-sheet & summary-sheet helpers ─────────────────────────────────

def write_results_data_rows(
    ws,
    results,
    row_data_fn,
    *,
    status_fill_map=None,
    status_col=2,
    center_cols=None,
    border=None,
    link_cols=None,
    row_height_fn=None,
):
    """Write styled data rows for a Results sheet.

    *row_data_fn(result, 1-based_num)* returns a list of cell values per row.
    Columns in *link_cols* are rendered as clickable hyperlinks.
    """
    if status_fill_map is None:
        status_fill_map = STATUS_FILL_COLORS
    if center_cols is None:
        center_cols = {1, 2, 3}
    if border is None:
        border = thin_border("CCCCCC")
    link_cols = link_cols or set()

    center = alignment_center(wrap_text=True)
    wrap = alignment_wrap_left()

    for row_idx, r in enumerate(results, 2):
        row_data = row_data_fn(r, row_idx - 1)
        fill_color = status_fill_map.get(r.get("status"), "FFFFFF")
        row_fill = PatternFill("solid", start_color=fill_color)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = center if col_idx in center_cols else wrap
            if col_idx == status_col:
                cell.fill = row_fill
            if col_idx in link_cols and value:
                cell.hyperlink = value
                cell.font = Font(
                    name="Arial", size=10, color="0563C1", underline="single",
                )

        if row_height_fn:
            ws.row_dimensions[row_idx].height = row_height_fn(row_data)


def write_summary_sheet(
    ws,
    results,
    *,
    positive_label="Have Target Feature",
    negative_label="Missing Target Feature",
    extra_rows=None,
    token_usage=None,
    deployment="",
    fill_hex="2F5496",
    border=None,
):
    """Write a standard Metric / Value summary sheet.

    *extra_rows* is an optional list of (label, value) tuples appended after
    the standard yes/no/skipped/error rows.

    If *token_usage* (a ``TokenUsageTracker``) is provided, LLM usage stats
    are appended at the bottom.
    """
    if border is None:
        border = thin_border("CCCCCC")
    center = alignment_center(wrap_text=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 15

    style_header_cell(ws, 1, 1, "Metric", fill_hex=fill_hex, border=border, alignment=center)
    style_header_cell(ws, 1, 2, "Value", fill_hex=fill_hex, border=border, alignment=center)

    counts = count_statuses(results)
    yes = counts.get("yes", 0)
    no = counts.get("no", 0)
    skipped = counts.get("skipped", 0)
    errors = counts.get("error", 0)
    total = len(results)

    rows = [
        ("Total Repos Checked", total),
        (positive_label, yes),
        (negative_label, no),
        ("Skipped (non-GitHub)", skipped),
        ("Errors", errors),
        ("Coverage (%)", coverage_formula(yes, total)),
    ]
    if extra_rows:
        rows.extend(extra_rows)

    if token_usage is not None:
        rows.append(("", ""))
        rows.append(("LLM Token Usage", ""))
        if deployment:
            rows.append(("  Model / Deployment", deployment))
        rows.append(("  Requests", token_usage.request_count))
        rows.append(("  Input Tokens", f"{token_usage.prompt_tokens:,}"))
        rows.append(("  Output Tokens", f"{token_usage.completion_tokens:,}"))
        rows.append(("  Total Tokens", f"{token_usage.total_tokens:,}"))

    for r_idx, (label, value) in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=label).font = Font(name="Arial", size=10)
        ws.cell(row=r_idx, column=2, value=value).font = Font(name="Arial", size=10)
        ws.cell(row=r_idx, column=2).alignment = center


# ── Multi-model helpers ───────────────────────────────────────────────────


def safe_sheet_name(base, model="", max_len=31):
    """Create an Excel-safe sheet name, truncating the model suffix if needed."""
    if not model:
        return base[:max_len]
    name = f"{base} ({model})"
    if len(name) <= max_len:
        return name
    available = max_len - len(base) - 3      # room inside " ()"
    return f"{base} ({model[:max(available, 3)]})"


def write_comparison_sheet(
    ws,
    all_model_results,
    *,
    fill_hex="2F5496",
    border=None,
):
    """Write a model comparison sheet: per-paper status across deployments.

    *all_model_results* is ``{deployment: {"results": [...], ...}}``.
    """
    if border is None:
        border = thin_border("CCCCCC")
    center = alignment_center(wrap_text=True)
    wrap = alignment_wrap_left()

    deployments = list(all_model_results.keys())
    headers = ["#", "Title", "Repo"] + deployments + ["Agreement"]
    widths = [5, 45, 40] + [15] * len(deployments) + [12]

    write_header_row(ws, headers, widths, fill_hex=fill_hex, border=border)

    first_results = all_model_results[deployments[0]]["results"]
    agree_count = 0

    for row_idx, paper in enumerate(first_results, 2):
        title = paper.get("title", "")
        repo = paper.get("repo", "")

        statuses = []
        for dep in deployments:
            dep_results = all_model_results[dep]["results"]
            paper_idx = row_idx - 2
            if paper_idx < len(dep_results):
                status = (dep_results[paper_idx].get("status") or "").upper()
            else:
                status = "N/A"
            statuses.append(status)

        all_agree = len(set(statuses)) == 1
        if all_agree:
            agree_count += 1

        row_vals = [row_idx - 1, title, repo] + statuses + [
            "Yes" if all_agree else "No",
        ]

        for col_idx, value in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = center if col_idx == 1 or col_idx > 3 else wrap
            if col_idx == len(row_vals):
                color = "C6EFCE" if all_agree else "FFC7CE"
                cell.fill = PatternFill("solid", start_color=color)

    # Agreement summary
    total = len(first_results)
    summary_row = total + 3
    ws.cell(row=summary_row, column=2, value="Agreement Rate").font = Font(
        name="Arial", size=10, bold=True,
    )
    pct = (
        f"{agree_count}/{total} ({agree_count / total * 100:.0f}%)"
        if total else "N/A"
    )
    ws.cell(row=summary_row, column=3, value=pct).font = Font(
        name="Arial", size=10,
    )

    # Token usage per model
    token_row = summary_row + 2
    ws.cell(row=token_row, column=1, value="Token Usage per Model").font = Font(
        name="Arial", size=11, bold=True,
    )
    token_row += 1
    token_headers = ["Model", "Requests", "Input Tokens",
                     "Output Tokens", "Total Tokens"]
    for col_idx, hdr in enumerate(token_headers, 1):
        cell = ws.cell(row=token_row, column=col_idx, value=hdr)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=fill_hex)
        cell.border = border
        cell.alignment = center
    token_row += 1
    for dep in deployments:
        tu = all_model_results[dep].get("token_usage")
        row_vals = [
            dep,
            tu.request_count if tu else 0,
            f"{tu.prompt_tokens:,}" if tu else "0",
            f"{tu.completion_tokens:,}" if tu else "0",
            f"{tu.total_tokens:,}" if tu else "0",
        ]
        for col_idx, value in enumerate(row_vals, 1):
            cell = ws.cell(row=token_row, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = center if col_idx >= 2 else wrap
        token_row += 1
