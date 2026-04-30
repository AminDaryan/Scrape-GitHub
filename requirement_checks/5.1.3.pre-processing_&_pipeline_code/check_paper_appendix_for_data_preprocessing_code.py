"""
Question 5.1.3 — Pre-processing & Pipeline Code classifier (repo-only)

Checks ONLY the GitHub repository for actual preprocessing / pipeline source
code (tokenization, filtering / deduplication / cleaning, or prompting-wrapper
construction). The paper text is no longer consulted.

Logic:
  1. Parse the GitHub URL from each paper entry. Non-GitHub repos are
     classified as NOT_APPLICABLE.                                     → check_paper()
  2. List every file in the repo via the GitHub API.                   → collect_repo_content()
  3. Sort files into three prioritised buckets:
       preprocessing-named source files (preprocess.py, tokenize.py …)
       files inside preprocessing-relevant folders (data/, pipeline/ …)
       other source files as a fallback.                               → collect_repo_content()
  4. Fetch the buckets within a shared character budget.               → collect_repo_content()
  5. Pass the combined text to the LLM and parse the verdict.          → make_check_paper_fn()
  6. Run across all configured models and export to Excel.             → main(), save_results()

Labels:
  EXISTS         — repo contains real preprocessing / pipeline code
  NOT_APPLICABLE — everything else (no code, non-GitHub URL, empty repo, or error)
"""

import os
import sys
from pathlib import Path
import re

from dotenv import load_dotenv

load_dotenv()

_this = Path(__file__).resolve()
sys.path.insert(0, str(_this.parent.parent.parent))
sys.path.insert(0, str(_this.parent.parent))

from common.fetch_and_parse_github_repo import (
    parse_github_repo, list_all_repo_files, fetch_file_content,
)
from common.repo_content_helpers import fetch_paths_with_char_budget
from common.llm_helpers import llm_call_parse_retry
from common.checker_pipeline import run_pipeline
from common.excel_output import (
    count_statuses,
    write_header_row, write_results_data_rows, write_summary_sheet,
    write_comparison_sheet, safe_sheet_name,
    thin_border, auto_row_height, alignment_center, alignment_wrap_left,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from openai_client import client, AZURE_OPENAI_DEPLOYMENTS
from papers_from_database import PAPERS
from prompts import SYSTEM_PROMPT, USER_PROMPT
from config import (
    PREPROC_NAME_RE, PREPROC_FOLDER_PREFIXES, SOURCE_EXTS,
    SKIP_FOLDER_PREFIXES, MAX_TOTAL_CHARS, PER_FILE_CHAR_CAP,
    MAX_NAMED_FILES, MAX_FOLDER_FILES, MAX_GENERIC_FILES,
)

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")

# ---------------------------------------------------------------------------
# Repo URL parsing
# ---------------------------------------------------------------------------

def parse_repo_url(url):
    """Extract (owner, repo) from a GitHub URL, including GitHub Pages links.

    Returns None for non-GitHub URLs (HuggingFace, anonymous repos, etc.).
    """
    if not url:
        return None
    url = url.strip()

    try:
        result = parse_github_repo(url)
    except Exception:
        result = None

    if result:
        if isinstance(result, tuple) and len(result) >= 2:
            return (result[0], result[1])
        if isinstance(result, dict):
            owner = result.get("owner")
            repo  = result.get("repo")
            if owner and repo:
                return (owner, repo)

    # GitHub Pages URL: https://<owner>.github.io/<repo>
    m = re.match(r"https?://([^./]+)\.github\.io/([^/?#]+)", url)
    if m:
        return (m.group(1), m.group(2))

    return None

# ---------------------------------------------------------------------------
# Repo content collection
# ---------------------------------------------------------------------------

def _is_skippable(path_lower):
    return any(path_lower.startswith(p) for p in SKIP_FOLDER_PREFIXES)


def _is_source(path_lower):
    if "." not in path_lower:
        return False
    ext = "." + path_lower.rsplit(".", 1)[-1]
    return ext in SOURCE_EXTS


def collect_repo_content(owner, repo):
    """Fetch preprocessing-relevant files from a GitHub repo.

    Files are sorted into three priority buckets and fetched within a shared
    character budget so that named preprocessing files are guaranteed to make
    it into the LLM's context before generic source code can crowd them out.
    Returns a (combined_text, fetched_paths) tuple.
    """
    print(f"  [GitHub] Listing {owner}/{repo}")
    try:
        all_files = list_all_repo_files(owner, repo)
    except Exception as e:
        print(f"  [GitHub] ERROR during listing: {type(e).__name__}: {e}")
        return "", []

    if not all_files:
        print(f"  [GitHub] Repo empty or unreachable")
        return "", []

    print(f"  [GitHub] {len(all_files)} files in repo")

    named_paths   = []   # source files whose basename matches PREPROC_NAME_RE
    folder_paths  = []   # source files inside a preprocessing-relevant folder
    generic_paths = []   # other source files — fallback context

    for f in all_files:
        fpath     = f["path"] if isinstance(f, dict) else f
        fpath_low = fpath.lower()
        basename  = fpath_low.split("/")[-1]

        if _is_skippable(fpath_low):
            continue
        if not _is_source(fpath_low):
            continue

        if PREPROC_NAME_RE.search(basename):
            named_paths.append(fpath)
        elif any(fpath_low.startswith(p) for p in PREPROC_FOLDER_PREFIXES):
            folder_paths.append(fpath)
        else:
            generic_paths.append(fpath)

    # Sort by depth then name for determinism. Shallower paths come first
    # because top-level scripts are usually the canonical entry points.
    for lst in (named_paths, folder_paths, generic_paths):
        lst.sort(key=lambda p: (p.count("/"), p.lower()))

    named_paths   = named_paths[:MAX_NAMED_FILES]
    folder_paths  = folder_paths[:MAX_FOLDER_FILES]
    generic_paths = generic_paths[:MAX_GENERIC_FILES]

    print(f"  [GitHub] Selected: {len(named_paths)} named, "
          f"{len(folder_paths)} in folders, {len(generic_paths)} generic")

    combined    = []
    fetched     = []
    total_chars = 0

    for label, paths in (("PREPROCESS-NAMED",  named_paths),
                         ("PREPROCESS-FOLDER", folder_paths),
                         ("GENERIC SOURCE",    generic_paths)):
        if not paths:
            continue
        blocks, got, total_chars = fetch_paths_with_char_budget(
            owner, repo, paths, MAX_TOTAL_CHARS,
            fetch_content=fetch_file_content,
            start_total_chars=total_chars,
            pause_seconds=0.15,
            header_label=label,
            per_file_char_cap=PER_FILE_CHAR_CAP,
        )
        combined.extend(blocks)
        fetched.extend(got)

    return "\n\n".join(combined), fetched


# ---------------------------------------------------------------------------
# LLM call wrapper
# ---------------------------------------------------------------------------

EMPTY_PREPROCESSING_PAYLOAD = {
    "classification":      None,
    "evidence":            "Empty LLM response after retry — content may exceed context window",
    "key_quotes":          [],
    "matched_categories":  [],
    "preprocessing_files": [],
    "diagnostic_notes":    "Empty response",
    "final_reason":        "",
}

# Content cache — avoids re-fetching GitHub repos when running multiple models.
_content_cache = {}


def _cached_collect_repo_content(owner, repo):
    """Return cached collect_repo_content result, fetching only on first call."""
    key = f"{owner}/{repo}"
    if key not in _content_cache:
        _content_cache[key] = collect_repo_content(owner, repo)
    return _content_cache[key]


# ---------------------------------------------------------------------------
# Per-paper orchestration
# ---------------------------------------------------------------------------

def make_check_paper_fn(deployment, token_usage):
    """Factory: return a check_paper function bound to a specific LLM deployment.

    Using a factory lets run_pipeline swap in different models without
    re-fetching GitHub content — the content cache is shared across all model
    instances for the same repo.
    """

    def _llm_check(repo_content, paper_title, repo_url):
        def build_msg(content):
            return USER_PROMPT.format(
                title=paper_title,
                repo_url=repo_url,
                repo_content=content if content else "[No source code files found in the repository]",
            )

        token_budget = 4000 if len(repo_content) > 50_000 else 2000

        return llm_call_parse_retry(
            client=client,
            deployment=deployment,
            system_prompt=SYSTEM_PROMPT,
            build_user_message=build_msg,
            content=repo_content,
            token_usage=token_usage,
            empty_payload=EMPTY_PREPROCESSING_PAYLOAD,
            required_list_fields=("key_quotes", "matched_categories", "preprocessing_files"),
            max_completion_tokens=token_budget,
            retry_truncate_chars=40_000,
            retry_max_tokens=4000,
            preview_chars=500,
        )

    def check_paper(paper):
        title = paper.get("title", "")
        url   = paper.get("repo", "") or ""

        result = {
            "title":               title,
            "semanticscholarid":   paper.get("semanticscholarid", ""),
            "repo":                url,
            "status":              None,
            "classification":      "NOT_APPLICABLE",
            "evidence":            "",
            "matched_categories":  [],
            "preprocessing_files": [],
            "files_checked":       [],
            "note":                "",
        }

        parsed = parse_repo_url(url)
        if not parsed:
            result["status"] = "skipped"
            result["note"] = (
                f"Not a GitHub repo — manual review needed. URL: {url}"
                if url else "No repo URL provided"
            )
            return result

        owner, repo_name = parsed

        try:
            repo_content, files_checked = _cached_collect_repo_content(owner, repo_name)
            result["files_checked"] = files_checked

            if not files_checked:
                result["status"] = "no"
                result["note"] = "No source files retrievable from the repo"
                return result

            llm_result = _llm_check(repo_content, title, url)
            clf = llm_result.get("classification")

            result["classification"]      = clf or "NOT_APPLICABLE"
            result["evidence"]            = (
                "; ".join(llm_result.get("key_quotes", [])) or
                llm_result.get("diagnostic_notes", "")
            )
            result["matched_categories"]  = llm_result.get("matched_categories", [])
            result["preprocessing_files"] = llm_result.get("preprocessing_files", [])
            result["note"]                = llm_result.get("final_reason", "")

            if clf is None:
                result["status"] = "error"
                result["note"]   = "Empty LLM response — content may be too large for this model"
            elif clf == "EXISTS":
                result["status"] = "yes"
            else:
                result["status"] = "no"

        except RuntimeError as e:
            result["status"] = "error"
            result["note"]   = str(e)
        except Exception as e:
            result["status"] = "error"
            result["note"]   = f"Unexpected error: {e}"

        return result

    return check_paper


# ---------------------------------------------------------------------------
# Console reporting
# ---------------------------------------------------------------------------

def print_results(results):
    """Print a compact console table of per-paper results."""
    W = 110
    print("\n" + "=" * W)
    print(f"{'#':<4} {'STATUS':<10} {'CATEGORIES':<36} TITLE")
    print("=" * W)

    for i, r in enumerate(results, 1):
        icon = "EXISTS" if r["status"] == "yes" else "NOT_APPL"

        cats_str = ", ".join(r.get("matched_categories", []))[:34]
        if not cats_str and r.get("note"):
            cats_str = r["note"][:34]

        title_short = r["title"][:46] + ("..." if len(r["title"]) > 46 else "")
        print(f"{i:<4} {icon:<10} {cats_str:<36} {title_short}")

        if r.get("evidence"):
            print(f"       evidence: {r['evidence'][:120]}")
        if r.get("files_checked"):
            extras = len(r["files_checked"]) - 5
            print(f"       files: {', '.join(r['files_checked'][:5])}"
                  + (f" (+{extras} more)" if extras > 0 else ""))

    print("=" * W)
    exists_count = sum(1 for r in results if r["status"] == "yes")
    na_count     = len(results) - exists_count
    print(
        f"\nSUMMARY: {exists_count} EXISTS | "
        f"{na_count} NOT_APPLICABLE | {len(results)} total\n"
    )


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_results(all_model_results, path=None):
    """Write per-model Results + Detail + Summary sheets, and a Model Comparison sheet."""
    if path is None:
        path = Path(__file__).resolve().parent / "results" / "preprocessing_code_results.xlsx"

    wb     = Workbook()
    border = thin_border()
    center = alignment_center()
    wrap   = alignment_wrap_left()

    deployments = list(all_model_results.keys())
    first_sheet = True

    for deployment in deployments:
        model_data  = all_model_results[deployment]
        results     = model_data["results"]
        token_usage = model_data["token_usage"]

        # ── Results sheet (per model) ────────────────────────────────────
        if first_sheet:
            ws1 = wb.active
            first_sheet = False
        else:
            ws1 = wb.create_sheet()
        ws1.title = safe_sheet_name("Results", deployment)

        hdrs1   = ["#", "Classification", "Title", "Repo",
                   "Categories", "Preprocessing Files", "Evidence",
                   "Files Checked", "Note", "Tokens Used"]
        widths1 = [5, 10, 45, 40, 30, 40, 55, 50, 35, 12]
        write_header_row(ws1, hdrs1, widths1, fill_hex="2F4F6F", border=border)

        def _to_classification(r):
            return "EXISTS" if r.get("status") == "yes" else "NOT_APPLICABLE"

        def results_row_data(r, num):
            return [
                num,
                _to_classification(r),
                r.get("title") or "",
                r.get("repo") or "",
                ", ".join(r.get("matched_categories") or []),
                ", ".join(r.get("preprocessing_files") or []),
                r.get("evidence") or "",
                ", ".join(r.get("files_checked") or []),
                r.get("note") or "",
                r.get("tokens_used") or "",
            ]

        write_results_data_rows(
            ws1, results, results_row_data,
            border=border,
            row_height_fn=lambda vals: auto_row_height(vals),
        )

        # ── Per-File Detail sheet (per model) ────────────────────────────
        ws2 = wb.create_sheet(safe_sheet_name("Detail", deployment))
        hdrs2   = ["Paper #", "Paper Title", "Repo",
                   "Preprocessing File", "GitHub Link"]
        widths2 = [8, 50, 45, 60, 75]
        write_header_row(ws2, hdrs2, widths2, fill_hex="2F4F6F", border=border)

        detail_row   = 2
        merge_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merge_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        for paper_num, r in enumerate(results, 1):
            owner, repo_name = parse_github_repo(r.get("repo", ""))
            file_rows = [f for f in r.get("preprocessing_files", []) if f and f.strip()]
            if not file_rows:
                continue

            start_row = detail_row
            for fpath in file_rows:
                fpath = fpath.strip()
                link  = (
                    f"https://github.com/{owner}/{repo_name}/blob/HEAD/{fpath}"
                    if owner else ""
                )
                for col_idx, value in enumerate([fpath, link], 4):
                    cell = ws2.cell(row=detail_row, column=col_idx, value=value)
                    cell.font   = Font(name="Arial", size=10)
                    cell.border = border
                    cell.alignment = wrap
                    if col_idx == 5 and value:
                        cell.hyperlink = value
                        cell.font = Font(name="Arial", size=10,
                                         color="0563C1", underline="single")
                ws2.row_dimensions[detail_row].height = 18
                detail_row += 1

            end_row = detail_row - 1
            ws2.cell(row=start_row, column=1, value=paper_num).font = Font(name="Arial", size=10)
            ws2.cell(row=start_row, column=2, value=r.get("title") or "").font = Font(name="Arial", size=10)
            ws2.cell(row=start_row, column=3, value=r.get("repo") or "").font = Font(name="Arial", size=10)
            for col_idx in (1, 2, 3):
                cell = ws2.cell(row=start_row, column=col_idx)
                cell.border = border
                cell.alignment = merge_center if col_idx == 1 else merge_left
            if end_row > start_row:
                ws2.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws2.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                ws2.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

        # ── Summary sheet (per model) ────────────────────────────────────
        ws3 = wb.create_sheet(safe_sheet_name("Summary", deployment))
        write_summary_sheet(
            ws3, results,
            positive_label="EXISTS",
            negative_label="NOT_APPLICABLE",
            token_usage=token_usage,
            deployment=deployment,
            fill_hex="2F4F6F",
            border=border,
        )

    # ── Model Comparison sheet ───────────────────────────────────────────
    if len(deployments) > 1:
        ws_cmp = wb.create_sheet("Model Comparison")
        write_comparison_sheet(
            ws_cmp, all_model_results, fill_hex="2F4F6F", border=border,
        )

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Full results saved to {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    """Entry point: run the preprocessing-code analysis pipeline."""
    _content_cache.clear()
    run_pipeline(
        papers=PAPERS,
        deployments=AZURE_OPENAI_DEPLOYMENTS,
        make_check_paper_fn=make_check_paper_fn,
        print_results_fn=print_results,
        save_results_fn=save_results,
        description="preprocessing / pipeline code",
        github_token=GITHUB_TOKEN,
    )


if __name__ == "__main__":
    main()
