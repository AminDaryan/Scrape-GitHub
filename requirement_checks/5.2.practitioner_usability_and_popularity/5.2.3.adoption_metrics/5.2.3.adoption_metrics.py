"""
Question 5.2.3 — Adoption Metrics.

Question: "What quantitative indicators of adoption exist?"

Expected values: free text (actual numbers) or "NOT STATED".
Fields tracked per paper:
  - GitHub stars
  - Forks
  - PyPI / package downloads (monthly, if the repo publishes a Python package)

How it works:
  1. For each paper, fetch the basic GitHub repo metadata (stars + forks)
     via the same Git Trees / repo API used elsewhere in the project.
  2. Try to detect a PyPI package name by reading the repo's
     ``pyproject.toml``, ``setup.cfg``, or ``setup.py`` (in that order).
  3. If a package name is found, query pypistats.org for the last month's
     download count.
  4. Write everything to a formatted Excel workbook.

Notes vs Question 5.2.2:
  Q 5.2.2 also surfaces stars/forks but inside a maintenance-indicator view
  (alongside recent commits, contributors, archived state, …).  This script
  is narrower: only adoption metrics, plus PyPI downloads which Q 5.2.2
  doesn't collect.

Usage:
  python 5.2.3.adoption_metrics.py
"""

import json
import re
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from urllib.error import HTTPError, URLError

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import load_dotenv

load_dotenv()

_this = Path(__file__).resolve()
# Folder layout for path resolution:
#   _this.parent             = 5.2.3.adoption_metrics/
#   _this.parent.parent      = 5.2.practitioner_usability_and_popularity/
#   _this.parent.parent.parent = requirement_checks/   ← needed for common/, data/
sys.path.insert(0, str(_this.parent.parent.parent))

from common.github_helpers import (
    github_get, parse_github_repo, fetch_file_content,
)
from common.excel_output import thin_border
from data.papers_source import load_papers

PAPERS = load_papers()

NOT_STATED = "NOT STATED"


# =============================================================================
# LOGGING
# =============================================================================

def log(msg):
    """Print a timestamped log message."""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


# =============================================================================
# GITHUB BASIC METRICS
# =============================================================================

def fetch_github_metrics(owner, repo):
    """Return ``{"stars": int, "forks": int}`` or None if the API call fails."""
    data = github_get(f"repos/{owner}/{repo}")
    if not data:
        return None
    return {
        "stars": data.get("stargazers_count", 0),
        "forks": data.get("forks_count", 0),
    }


# =============================================================================
# PYPI PACKAGE-NAME DETECTION
# =============================================================================

# PEP 621 / Poetry / setup.cfg name patterns.  Each is anchored loosely
# enough to tolerate quote-style variation (single, double, triple) and
# trailing whitespace, but strict enough to avoid matching unrelated
# ``name = ...`` lines elsewhere in the file.
_PYPROJECT_NAME_RE = re.compile(
    r"""(?ms)^\s*name\s*=\s*['"]([A-Za-z0-9_.\-]+)['"]"""
)
_SETUP_CFG_NAME_RE = re.compile(
    r"(?im)^\s*name\s*=\s*([A-Za-z0-9_.\-]+)"
)
_SETUP_PY_NAME_RE = re.compile(
    r"""setup\s*\([^)]*?\bname\s*=\s*['"]([A-Za-z0-9_.\-]+)['"]""",
    re.IGNORECASE | re.DOTALL,
)


def detect_pypi_package_name(owner, repo):
    """Try to find the package name this repo would publish on PyPI.

    Looks at three files in order, returning the first match:
      1. ``pyproject.toml``  (PEP 621 ``[project]`` or ``[tool.poetry]``)
      2. ``setup.cfg``       (``[metadata]`` section)
      3. ``setup.py``        (regex for ``setup(name="...")``)

    Returns the detected package name string, or None when none of the
    files exist or no name can be parsed.

    Note: setup.py parsing is intentionally regex-based — a real
    ``ast``-based parse would be more robust but also much more code,
    and most modern Python repos use pyproject.toml anyway.
    """
    for path, pattern in (
        ("pyproject.toml", _PYPROJECT_NAME_RE),
        ("setup.cfg",      _SETUP_CFG_NAME_RE),
        ("setup.py",       _SETUP_PY_NAME_RE),
    ):
        content = fetch_file_content(owner, repo, path)
        if not content:
            continue
        match = pattern.search(content)
        if match:
            return match.group(1)
    return None


# =============================================================================
# PYPI DOWNLOAD COUNT
# =============================================================================

def fetch_pypi_monthly_downloads(package_name):
    """Return last-month download count from pypistats.org, or None on failure.

    pypistats.org exposes a free, unauthenticated API:
      GET https://pypistats.org/api/packages/{name}/recent

    Response shape:
      {"data": {"last_day": int, "last_week": int, "last_month": int}, ...}

    Returns None when the package is not on PyPI (404) or any other
    network/parse error occurs — callers treat None as "not stated".
    """
    url = f"https://pypistats.org/api/packages/{urllib.parse.quote(package_name, safe='')}/recent"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "scrape-github-adoption/1.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            payload = json.loads(resp.read().decode())
    except (HTTPError, URLError, json.JSONDecodeError, TimeoutError):
        return None

    data = payload.get("data") or {}
    monthly = data.get("last_month")
    if isinstance(monthly, int):
        return monthly
    return None


# =============================================================================
# PER-PAPER ASSESSMENT
# =============================================================================

def assess_paper(paper):
    """Collect adoption metrics for one paper entry.

    Result dict:
      title, repo, stars, forks, pypi_package, pypi_monthly_downloads, note

    Numeric fields are integers when known and the string ``NOT_STATED`` when
    we can't determine them (e.g. non-GitHub URL, repo unreachable, no PyPI
    package detected).  Note explains skips/errors.
    """
    title = paper.get("title", "")
    url   = paper.get("repo", "") or ""

    result = {
        "title": title,
        "repo": url,
        "stars": NOT_STATED,
        "forks": NOT_STATED,
        "pypi_package": NOT_STATED,
        "pypi_monthly_downloads": NOT_STATED,
        "note": "",
    }

    if not url:
        result["note"] = "No repo URL provided"
        return result

    owner, repo = parse_github_repo(url)
    if not owner:
        result["note"] = "Not a GitHub repo URL"
        return result

    # ── GitHub stars / forks ────────────────────────────────────────────
    log(f"  GitHub: {owner}/{repo}")
    metrics = fetch_github_metrics(owner, repo)
    if metrics is None:
        result["note"] = "GitHub API call failed (private repo, deleted, or rate-limited?)"
        return result

    result["stars"] = metrics["stars"]
    result["forks"] = metrics["forks"]

    # ── PyPI package + monthly downloads ────────────────────────────────
    log(f"  Detecting PyPI package...")
    package_name = detect_pypi_package_name(owner, repo)
    if not package_name:
        result["pypi_package"] = NOT_STATED
        result["pypi_monthly_downloads"] = NOT_STATED
        return result

    result["pypi_package"] = package_name
    log(f"  Querying pypistats for '{package_name}'...")
    monthly = fetch_pypi_monthly_downloads(package_name)
    if monthly is None:
        result["pypi_monthly_downloads"] = NOT_STATED
        if not result["note"]:
            result["note"] = (
                f"Package '{package_name}' detected from repo but not found on PyPI / pypistats."
            )
    else:
        result["pypi_monthly_downloads"] = monthly

    return result


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

ALT_ROW_FILL    = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL      = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL     = PatternFill("solid", fgColor="BDD7EE")
HEADER_FONT     = Font(name="Arial", bold=True, size=11, color="1F4E79")
DATA_FONT       = Font(name="Arial", size=10)
HEADER_ALIGN    = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Single-row headers sit by themselves on row 1 spanning two rows.
# Sub-headers under "PyPI" sit on row 2.
SINGLE_HEADERS = [
    (1, "#"),
    (2, "Paper Title"),
    (3, "Repo URL"),
    (4, "GitHub Stars"),
    (5, "Forks"),
    (8, "Notes"),
]
PYPI_SUB_HEADERS = [
    (6, "Package"),
    (7, "Monthly Downloads"),
]
EXCEL_COL_WIDTHS = [5, 50, 55, 14, 12, 26, 22, 50]
NUMERIC_COLS     = {4, 5, 7}   # Stars, Forks, Monthly downloads
DATA_START_ROW   = 3           # Two rows of headers


def _style_header_cell(cell, border):
    """Apply the shared header look (light-blue fill, dark-blue text, centered)."""
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border    = border


def save_to_excel(results, filename):
    """Write adoption metrics to a formatted Excel workbook (one sheet).

    Layout uses a two-row header so the PyPI columns share a parent label:

        | # | Paper Title | Repo URL | GitHub Stars | Forks |     PyPI      | Notes |
        |   |             |          |              |       | Pkg | Monthly |       |

    Single-cell headers (cols A–E, H) are merged vertically across rows 1–2
    so they line up visually with the PyPI sub-headers on row 2.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adoption Metrics"
    border = thin_border("B8CCE4")

    # Column widths.
    for idx, width in enumerate(EXCEL_COL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    # ── Header row 1 ────────────────────────────────────────────────────
    # Single-cell headers go in row 1 and get merged with row 2.
    for col_idx, label in SINGLE_HEADERS:
        _style_header_cell(ws.cell(row=1, column=col_idx, value=label), border)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        # Style the merged "ghost" cell on row 2 too so its borders draw.
        _style_header_cell(ws.cell(row=2, column=col_idx), border)

    # PyPI parent header spans columns 6–7 on row 1.
    _style_header_cell(ws.cell(row=1, column=6, value="PyPI"), border)
    _style_header_cell(ws.cell(row=1, column=7), border)
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)

    # ── Header row 2 ────────────────────────────────────────────────────
    for col_idx, label in PYPI_SUB_HEADERS:
        _style_header_cell(ws.cell(row=2, column=col_idx, value=label), border)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # ── Data rows ───────────────────────────────────────────────────────
    for row_idx, r in enumerate(results, start=DATA_START_ROW):
        row_fill = ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL
        values = [
            row_idx - DATA_START_ROW + 1,
            r.get("title", ""),
            r.get("repo", ""),
            r.get("stars", NOT_STATED),
            r.get("forks", NOT_STATED),
            r.get("pypi_package", NOT_STATED),
            r.get("pypi_monthly_downloads", NOT_STATED),
            r.get("note", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in {1} | NUMERIC_COLS else "left",
                vertical="center",
                wrap_text=True,
            )
            if col_idx in NUMERIC_COLS and isinstance(value, (int, float)):
                cell.number_format = "#,##0"
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = f"A{DATA_START_ROW}"

    Path(filename).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filename)
    log(f"Saved to {filename}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Run adoption-metric collection over every paper and save to Excel."""
    log("\n" + "=" * 80)
    log("Question 5.2.3 — Adoption Metrics")
    log("=" * 80)
    log(f"Processing {len(PAPERS)} papers...\n")

    results = []
    started = time.time()

    for idx, paper in enumerate(PAPERS, 1):
        title = paper.get("title", "")
        log(f"[{idx}/{len(PAPERS)}] {title[:60]}")
        result = assess_paper(paper)
        results.append(result)

        stars = result["stars"]
        forks = result["forks"]
        pypi  = result["pypi_monthly_downloads"]
        log(f"  => stars={stars}, forks={forks}, pypi/month={pypi}")
        log("")
        time.sleep(0.3)

    elapsed = time.time() - started
    log("=" * 80)
    log(f"Done in {elapsed:.1f}s")
    log("=" * 80)

    output_path = _this.parent / "adoption_metrics.xlsx"
    save_to_excel(results, str(output_path))


if __name__ == "__main__":
    main()
