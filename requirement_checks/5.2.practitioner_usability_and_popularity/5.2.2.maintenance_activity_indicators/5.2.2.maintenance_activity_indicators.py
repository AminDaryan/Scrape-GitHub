"""
GitHub repository scraper — metadata + maintenance indicators.

Collects two layers of information for each paper's GitHub repository:

A. Basic repository metadata (single API call):
     - Name / URL
     - Stars / Forks / Open Issues counts
     - Primary language
     - License
     - Last push / last update dates
     - README presence (root-level readme.* file)
     - Setup files presence (setup.py, requirements.txt, Dockerfile, …)

B. Maintenance indicators (answers "What indicators of ongoing maintenance exist?"):
     1. RECENT_COMMITS         — Commits within 6 months of the evaluation date.
     2. ACTIVE_ISSUE_RESPONSES — Issues/PRs that received maintainer responses.
     3. MULTIPLE_CONTRIBUTORS  — More than the original author(s) contributed.
     4. VERSIONED_RELEASES     — Tagged releases using semantic versioning.
     5. STALE                  — No activity (commits, issues, releases) for > 1 year.
     6. ARCHIVED               — Repository explicitly archived or marked deprecated.
     7. NOT_ASSESSABLE         — Repository is inaccessible or API calls failed.

Both layers are written to one Excel workbook with two sheets:
  - "Results"  — one row per paper with metadata and indicator columns.
  - "Summary"  — counts of each overall maintenance status.

Usage:
  python 5.2.2.maintenance_activity_indicators.py
"""

import os
import re
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from dotenv import load_dotenv

load_dotenv()

_this = Path(__file__).resolve()
# Folder layout for path resolution:
#   _this.parent             = 5.2.2.maintenance_activity_indicators/
#   _this.parent.parent      = 5.2.practitioner_usability_and_popularity/
#   _this.parent.parent.parent = requirement_checks/   ← needed for common/, data/, openai_client
sys.path.insert(0, str(_this.parent.parent.parent))

from data.papers_source import load_papers

PAPERS = load_papers()

# Setup-related filenames that indicate the repo has installation instructions.
# This is a lightweight replacement for the original NLP-based detector — if
# any of these are present at the repo root, we say "Yes, the repo has setup
# information" without paying the cost of an LLM call.
SETUP_FILE_BASENAMES = {
    "setup.py", "setup.cfg", "pyproject.toml",
    "requirements.txt", "requirements-dev.txt",
    "environment.yml", "environment.yaml", "conda.yml",
    "dockerfile", "docker-compose.yml", "docker-compose.yaml",
    "makefile", "install.sh", "install.md",
    "package.json", "yarn.lock", "package-lock.json",
    "pipfile", "pipfile.lock", "poetry.lock",
    "cargo.toml", "go.mod", "gemfile",
}

README_PREFIXES = ("readme",)

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")
HEADERS = {"Authorization": f"Bearer {GITHUB_TOKEN}"} if GITHUB_TOKEN else None

NOW = datetime.now(timezone.utc)
SIX_MONTHS_AGO = NOW - timedelta(days=182)
ONE_YEAR_AGO = NOW - timedelta(days=365)

SEMVER_RE = re.compile(r"v?\d+\.\d+(\.\d+)?")


# =============================================================================
# LOGGING
# =============================================================================

def log(msg):
    """Print a timestamped log message to the console."""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


# =============================================================================
# GITHUB API HELPERS
# =============================================================================

def parse_github_url(url):
    """Extract (owner, repo) from a GitHub URL.

    Strips trailing slashes and .git suffixes before splitting.
    """
    url = url.rstrip("/")
    if url.endswith(".git"):
        url = url[:-4]
    parts = url.split("/")
    return parts[-2], parts[-1]


def _gh_get(endpoint, headers=None, params=None):
    """Make a GET request to the GitHub API.

    Returns the parsed JSON on success, or None on failure.
    """
    url = f"https://api.github.com{endpoint}"
    resp = requests.get(url, headers=headers, params=params or {})
    if resp.status_code == 200:
        return resp.json()
    return None


def _parse_date(date_str):
    """Parse an ISO 8601 date string from the GitHub API into a datetime."""
    if not date_str:
        return None
    return datetime.fromisoformat(date_str.replace("Z", "+00:00"))


# =============================================================================
# MAINTENANCE INDICATOR CHECKS
# =============================================================================

def check_recent_commits(owner, repo, headers=None):
    """Check if the repo has commits within the last 6 months.

    Uses the commits API with a 'since' parameter to only fetch recent
    commits. Returns (bool, int) — whether recent commits exist and how
    many were found.
    """
    data = _gh_get(
        f"/repos/{owner}/{repo}/commits",
        headers=headers,
        params={"since": SIX_MONTHS_AGO.isoformat(), "per_page": 100},
    )
    if data is None:
        return None, 0
    count = len(data)
    return count > 0, count


def check_active_issue_responses(owner, repo, headers=None):
    """Check if maintainers respond to issues and PRs.

    Fetches the 30 most recently updated issues (including PRs). An issue
    counts as "responded to" if it has at least one comment, because that
    typically means a maintainer or community member engaged with it.

    Returns (bool, str) — whether active responses exist and a summary
    like "12/30 issues have responses".
    """
    data = _gh_get(
        f"/repos/{owner}/{repo}/issues",
        headers=headers,
        params={"state": "all", "sort": "updated", "per_page": 30},
    )
    if data is None:
        return None, "API unavailable"
    if len(data) == 0:
        return False, "No issues found"

    responded = sum(1 for issue in data if issue.get("comments", 0) > 0)
    total = len(data)
    has_responses = responded > 0
    summary = f"{responded}/{total} issues have responses"
    return has_responses, summary


def check_multiple_contributors(owner, repo, headers=None):
    """Check if more than the original author(s) contributed.

    Fetches the contributor list. A repo with more than 1 contributor
    indicates community involvement beyond the original authors.

    Returns (bool, int) — whether multiple contributors exist and the count.
    """
    data = _gh_get(
        f"/repos/{owner}/{repo}/contributors",
        headers=headers,
        params={"per_page": 100},
    )
    if data is None:
        return None, 0
    count = len(data)
    return count > 1, count


def check_versioned_releases(owner, repo, headers=None):
    """Check if the repo uses tagged releases with semantic versioning.

    Fetches releases first (curated by maintainers), then falls back to
    tags. A release/tag matches if its name looks like a semver string
    (e.g. v1.2.3, 2.0, v0.1.0).

    Returns (bool, str) — whether versioned releases exist and the latest
    tag name (or "None").
    """
    releases = _gh_get(
        f"/repos/{owner}/{repo}/releases",
        headers=headers,
        params={"per_page": 10},
    )
    if releases:
        semver_releases = [r for r in releases if SEMVER_RE.match(r.get("tag_name", ""))]
        if semver_releases:
            latest = semver_releases[0]["tag_name"]
            return True, f"{latest} ({len(semver_releases)} semver releases)"

    tags = _gh_get(
        f"/repos/{owner}/{repo}/tags",
        headers=headers,
        params={"per_page": 10},
    )
    if tags:
        semver_tags = [t for t in tags if SEMVER_RE.match(t.get("name", ""))]
        if semver_tags:
            latest = semver_tags[0]["name"]
            return True, f"{latest} ({len(semver_tags)} semver tags)"

    return False, "None"


def check_stale(owner, repo, repo_data, headers=None):
    """Check if the repo has had no activity for more than 1 year.

    Uses the most recent of: last push date, last issue update, or last
    release date. If all are older than 1 year, the repo is considered stale.

    Args:
        repo_data: The main repo API response dict (contains pushed_at).

    Returns (bool, str) — whether the repo is stale and the date of last
    activity.
    """
    last_push = _parse_date(repo_data.get("pushed_at"))
    last_update = _parse_date(repo_data.get("updated_at"))

    dates = [d for d in [last_push, last_update] if d is not None]

    releases = _gh_get(
        f"/repos/{owner}/{repo}/releases",
        headers=headers,
        params={"per_page": 1},
    )
    if releases:
        release_date = _parse_date(releases[0].get("published_at"))
        if release_date:
            dates.append(release_date)

    if not dates:
        return None, "No activity dates found"

    most_recent = max(dates)
    is_stale = most_recent < ONE_YEAR_AGO
    return is_stale, most_recent.strftime("%Y-%m-%d")


def check_archived(repo_data):
    """Check if the repo is explicitly archived or marked as deprecated.

    Checks the GitHub 'archived' flag and scans the description for
    keywords like 'deprecated', 'unmaintained', or 'no longer maintained'.

    Args:
        repo_data: The main repo API response dict.

    Returns (bool, str) — whether the repo is archived/deprecated and
    the reason.
    """
    if repo_data.get("archived"):
        return True, "Repository is archived"

    description = (repo_data.get("description") or "").lower()
    deprecation_keywords = ["deprecated", "unmaintained", "no longer maintained",
                            "not maintained", "obsolete", "abandoned"]
    for keyword in deprecation_keywords:
        if keyword in description:
            return True, f"Description contains '{keyword}'"

    return False, ""


# =============================================================================
# PER-REPO ASSESSMENT
# =============================================================================

def _empty_metadata():
    """Return blank metadata fields used when a repo is unreachable."""
    return {
        "stars": "N/A", "forks": "N/A", "open_issues": "N/A",
        "language": "N/A", "license": "N/A",
        "pushed_at": "N/A", "updated_at": "N/A",
        "has_readme": "N/A", "has_setup": "N/A",
    }


def assess_repo(owner, repo, headers=None):
    """Fetch basic metadata + run all maintenance indicator checks.

    Two layers of information are gathered:
      A) Basic metadata, available straight from the repo API response:
         stars, forks, open_issues, language, license, last push/update dates.
         README and setup-file presence via a scan of the repo's root file tree
         (no NLP/LLM — see SETUP_FILE_BASENAMES).
      B) Maintenance indicators, requiring extra API calls:
         recent commits, issue responses, contributors, versioned releases,
         staleness, archived/deprecated flag.

    Each indicator is stored as ``(bool, detail_str)``.  If the repo API
    call fails, every indicator is set to ``(None, "API unavailable")``
    and ``not_assessable`` is True.
    """
    log(f"  Fetching repo metadata: {owner}/{repo}")
    repo_data = _gh_get(f"/repos/{owner}/{repo}", headers=headers)

    if repo_data is None:
        log(f"  Failed to fetch repo — marking as NOT_ASSESSABLE")
        return {
            "name": f"{owner}/{repo}",
            "html_url": f"https://github.com/{owner}/{repo}",
            **_empty_metadata(),
            "recent_commits": (None, "API unavailable"),
            "active_issue_responses": (None, "API unavailable"),
            "multiple_contributors": (None, "API unavailable"),
            "versioned_releases": (None, "API unavailable"),
            "stale": (None, "API unavailable"),
            "archived": (None, "API unavailable"),
            "not_assessable": True,
        }

    # ── Layer A: Basic metadata (free — already in repo_data) ──────────
    metadata = {
        "stars":       repo_data.get("stargazers_count", 0),
        "forks":       repo_data.get("forks_count", 0),
        "open_issues": repo_data.get("open_issues_count", 0),
        "language":    repo_data.get("language") or "N/A",
        "license":     (repo_data.get("license") or {}).get("name", "N/A") if repo_data.get("license") else "N/A",
        "pushed_at":   (repo_data.get("pushed_at") or "")[:10] or "N/A",
        "updated_at":  (repo_data.get("updated_at") or "")[:10] or "N/A",
    }

    # README + setup-instructions checks via the repo file tree.
    # One API call returns the whole tree; we then scan the root entries
    # for README files and common setup-related filenames.
    log(f"  Checking for README and setup files...")
    tree = _gh_get(f"/repos/{owner}/{repo}/git/trees/HEAD", headers=headers)
    if tree and "tree" in tree:
        root_basenames = {
            entry["path"].lower()
            for entry in tree["tree"]
            if "/" not in entry.get("path", "") and entry.get("type") == "blob"
        }
        metadata["has_readme"] = any(b.startswith(README_PREFIXES) for b in root_basenames)
        metadata["has_setup"] = any(b in SETUP_FILE_BASENAMES for b in root_basenames)
    else:
        metadata["has_readme"] = "N/A"
        metadata["has_setup"] = "N/A"

    # ── Layer B: Maintenance indicators ────────────────────────────────
    log(f"  Checking recent commits...")
    recent_commits = check_recent_commits(owner, repo, headers)

    log(f"  Checking issue responses...")
    issue_responses = check_active_issue_responses(owner, repo, headers)

    log(f"  Checking contributors...")
    contributors = check_multiple_contributors(owner, repo, headers)

    log(f"  Checking versioned releases...")
    releases = check_versioned_releases(owner, repo, headers)

    log(f"  Checking staleness...")
    stale = check_stale(owner, repo, repo_data, headers)

    log(f"  Checking archived status...")
    archived = check_archived(repo_data)

    return {
        "name": repo_data.get("full_name", f"{owner}/{repo}"),
        "html_url": repo_data.get("html_url", ""),
        **metadata,
        "recent_commits": recent_commits,
        "active_issue_responses": issue_responses,
        "multiple_contributors": contributors,
        "versioned_releases": releases,
        "stale": stale,
        "archived": archived,
        "not_assessable": False,
    }


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

LIGHT_BLUE_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_ROW_FILL    = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL      = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL      = PatternFill("solid", fgColor="C6EFCE")
RED_FILL        = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL     = PatternFill("solid", fgColor="FFEB9C")
GREY_FILL       = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT     = Font(name="Arial", bold=True, size=11, color="1F4E79")
DATA_FONT       = Font(name="Arial", size=10)
THIN_BORDER     = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)

EXCEL_HEADERS = [
    "#", "Paper Title", "Repository", "URL",
    # Basic metadata (Layer A)
    "Stars", "Forks", "Open\nIssues", "Language", "License",
    "Last Push", "Last Update", "Has\nREADME", "Setup\nInstructions",
    # Maintenance indicators (Layer B)
    "Recent Commits\n(6 months)", "Active Issue\nResponses",
    "Multiple\nContributors", "Versioned\nReleases",
    "Stale\n(>1 year)", "Archived /\nDeprecated", "Overall Status",
]
EXCEL_COL_WIDTHS = [
    5, 45, 28, 50,
    10, 10, 10, 14, 35, 14, 14, 12, 26,
    22, 26, 22, 28, 22, 22, 18,
]

# 1-indexed column groups for styling
NUMERIC_COLS         = {5, 6, 7}                    # Stars, Forks, Open Issues
INDICATOR_POS_COLS   = {14, 15, 16, 17}             # positive maintenance indicators
INDICATOR_NEG_COLS   = {18, 19}                     # stale, archived (inverted colour)
OVERALL_COL          = 20


def _indicator_fill(value):
    """Return the appropriate cell fill colour for a boolean indicator.

    True = green (good for positive indicators, bad for stale/archived).
    False = red.
    None = grey (not assessable).
    """
    if value is None:
        return GREY_FILL
    return GREEN_FILL if value else RED_FILL


def _stale_archived_fill(value):
    """Inverted colour for negative indicators (stale, archived).

    True = red (bad — repo IS stale/archived).
    False = green (good — repo is NOT stale/archived).
    """
    if value is None:
        return GREY_FILL
    return RED_FILL if value else GREEN_FILL


def _overall_status(result):
    """Derive an overall maintenance status label from all indicators.

    Priority order:
      NOT_ASSESSABLE → ARCHIVED → STALE → ACTIVE / MINIMAL / INACTIVE
    """
    if result["not_assessable"]:
        return "NOT_ASSESSABLE"
    if result["archived"][0]:
        return "ARCHIVED"
    if result["stale"][0]:
        return "STALE"

    positive_count = sum(1 for key in ("recent_commits", "active_issue_responses",
                                        "multiple_contributors", "versioned_releases")
                         if result[key][0] is True)
    if positive_count >= 3:
        return "ACTIVE"
    if positive_count >= 1:
        return "MINIMAL"
    return "INACTIVE"


OVERALL_FILL = {
    "ACTIVE":         GREEN_FILL,
    "MINIMAL":        YELLOW_FILL,
    "INACTIVE":       RED_FILL,
    "STALE":          RED_FILL,
    "ARCHIVED":       RED_FILL,
    "NOT_ASSESSABLE": GREY_FILL,
}


def save_to_excel(all_results, filename="maintenance_indicators.xlsx"):
    """Write maintenance indicator results to a formatted Excel workbook.

    Creates two sheets:
      - Results: one row per paper with colour-coded indicator columns.
      - Summary: counts of each overall status category.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # ── Header row ──────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(zip(EXCEL_HEADERS, EXCEL_COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = LIGHT_BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 36

    # ── Data rows ───────────────────────────────────────────────────────
    for row_idx, entry in enumerate(all_results, start=2):
        result = entry["result"]
        overall = _overall_status(result)

        # Unpack each indicator: (bool_value, detail_string)
        rc_val, rc_detail     = result["recent_commits"]
        ir_val, ir_detail     = result["active_issue_responses"]
        mc_val, mc_detail     = result["multiple_contributors"]
        vr_val, vr_detail     = result["versioned_releases"]
        st_val, st_detail     = result["stale"]
        ar_val, ar_detail     = result["archived"]

        def _fmt(val, detail):
            if val is None:
                return "N/A"
            label = "Yes" if val else "No"
            return f"{label} — {detail}" if detail else label

        values = [
            row_idx - 1,
            entry.get("paper_title", ""),
            result["name"],
            result["html_url"],
            # Basic metadata
            result.get("stars", "N/A"),
            result.get("forks", "N/A"),
            result.get("open_issues", "N/A"),
            result.get("language", "N/A"),
            result.get("license", "N/A"),
            result.get("pushed_at", "N/A"),
            result.get("updated_at", "N/A"),
            result.get("has_readme", "N/A"),
            result.get("has_setup", "N/A"),
            # Maintenance indicators
            _fmt(rc_val, rc_detail),
            _fmt(ir_val, ir_detail),
            _fmt(mc_val, mc_detail),
            _fmt(vr_val, vr_detail),
            _fmt(st_val, st_detail),
            _fmt(ar_val, ar_detail),
            overall,
        ]

        # Column index → fill colour for indicator columns
        indicator_fills = {
            14: _indicator_fill(rc_val),
            15: _indicator_fill(ir_val),
            16: _indicator_fill(mc_val),
            17: _indicator_fill(vr_val),
            18: _stale_archived_fill(st_val),
            19: _stale_archived_fill(ar_val),
            OVERALL_COL: OVERALL_FILL.get(overall, WHITE_FILL),
        }

        # Columns that should be centre-aligned (everything except text fields).
        center_cols = {1} | NUMERIC_COLS | {8, 10, 11, 12, 13} | INDICATOR_POS_COLS | INDICATOR_NEG_COLS | {OVERALL_COL}

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx in center_cols else "left",
                vertical="center",
                wrap_text=True,
            )
            if col_idx in NUMERIC_COLS and isinstance(value, (int, float)):
                cell.number_format = "#,##0"
            if col_idx in indicator_fills:
                cell.fill = indicator_fills[col_idx]
        ws.row_dimensions[row_idx].height = 40

    ws.freeze_panes = "A2"

    # ── Summary sheet ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12

    ws2.cell(row=1, column=1, value="Overall Status").font = HEADER_FONT
    ws2.cell(row=1, column=2, value="Count").font = HEADER_FONT
    for c in (1, 2):
        ws2.cell(row=1, column=c).fill = LIGHT_BLUE_FILL
        ws2.cell(row=1, column=c).border = THIN_BORDER
        ws2.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    status_counts = {}
    for entry in all_results:
        status = _overall_status(entry["result"])
        status_counts[status] = status_counts.get(status, 0) + 1

    for row_idx, status in enumerate(["ACTIVE", "MINIMAL", "INACTIVE",
                                       "STALE", "ARCHIVED", "NOT_ASSESSABLE"], start=2):
        ws2.cell(row=row_idx, column=1, value=status).font = DATA_FONT
        ws2.cell(row=row_idx, column=1).fill = OVERALL_FILL.get(status, WHITE_FILL)
        ws2.cell(row=row_idx, column=1).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2, value=status_counts.get(status, 0)).font = DATA_FONT
        ws2.cell(row=row_idx, column=2).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")

    total_row = 8
    ws2.cell(row=total_row, column=1, value="Total Papers").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=total_row, column=2, value=len(all_results)).font = Font(name="Arial", bold=True, size=10)

    Path(filename).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filename)
    log(f"Saved to {filename}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Scrape maintenance indicators from all paper repos and save to Excel."""
    log("\n" + "=" * 80)
    log("Starting GitHub Maintenance Indicator Scraper")
    log("=" * 80)

    log(f"Processing {len(PAPERS)} papers...\n")
    all_results = []
    start_total = time.time()

    for idx, paper in enumerate(PAPERS, 1):
        url = paper.get("repo", "")
        title = paper.get("title", "")
        log(f"[{idx}/{len(PAPERS)}] {title[:60]}")

        if not url or "github.com" not in url:
            log(f"  Skipped — not a GitHub URL: {url}")
            all_results.append({
                "paper_title": title,
                "result": {
                    "name": url or "N/A",
                    "html_url": url or "",
                    **_empty_metadata(),
                    "recent_commits": (None, "Not a GitHub repo"),
                    "active_issue_responses": (None, "Not a GitHub repo"),
                    "multiple_contributors": (None, "Not a GitHub repo"),
                    "versioned_releases": (None, "Not a GitHub repo"),
                    "stale": (None, "Not a GitHub repo"),
                    "archived": (None, "Not a GitHub repo"),
                    "not_assessable": True,
                },
            })
            log("")
            continue

        owner, repo = parse_github_url(url)
        result = assess_repo(owner, repo, HEADERS)
        all_results.append({
            "paper_title": title,
            "result": result,
        })

        overall = _overall_status(result)
        log(f"  => {overall}")
        log("")
        time.sleep(0.3)

    total_time = time.time() - start_total
    log("=" * 80)
    log(f"All papers processed in {total_time:.1f}s")
    log("=" * 80)

    # Print summary to console
    status_counts = {}
    for entry in all_results:
        status = _overall_status(entry["result"])
        status_counts[status] = status_counts.get(status, 0) + 1
    for status, count in sorted(status_counts.items()):
        log(f"  {status}: {count}")

    output_path = Path(__file__).parent / "maintenance_indicators.xlsx"
    save_to_excel(all_results, str(output_path))
    log("Done!")


if __name__ == "__main__":
    main()
