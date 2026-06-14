"""
Question 5.2.4 — Post-Publication Maintenance (automatic extraction).

Question: "What is the post-publication maintenance status?"

Expected values: free text or "NOT AVAILABLE".
Fields tracked per paper:
  - Date of last commit  (committer date on the default branch's HEAD commit)
  - Total number of commits on the default branch

How it works (no LLM):
  1. Last commit date — call ``GET /repos/{owner}/{repo}/commits?per_page=1``
     and read the first (most recent) entry's committer date.
  2. Total commit count — same endpoint, but read the response's ``Link``
     header.  GitHub's pagination cursor includes a ``rel="last"`` URL
     whose ``page=N`` parameter equals the total number of commits when
     ``per_page=1``.  The shared helper
     ``github_get_with_link_header`` returns the raw header so we can
     parse it here.
  3. If the repo is empty / unreachable / not on GitHub, we report
     ``NOT AVAILABLE`` for both fields with a short note explaining why.

Usage:
  python 5.2.4.post_publication_maintenance.py
"""

import re
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import load_dotenv

load_dotenv()

_this = Path(__file__).resolve()
# Folder layout for path resolution:
#   _this.parent             = 5.2.4.post_publication_maintenance/
#   _this.parent.parent      = 5.2.practitioner_usability_and_popularity/
#   _this.parent.parent.parent = requirement_checks/   ← needed for common/, data/
sys.path.insert(0, str(_this.parent.parent.parent))

from common.github_helpers import (
    github_get, github_get_with_link_header, parse_github_repo,
)
from common.excel_output import thin_border, write_header_row
from data.papers_source import load_papers

PAPERS = load_papers()

NOT_AVAILABLE = "NOT AVAILABLE"

# Matches `<…&page=N>; rel="last"` inside a GitHub Link header.
_LINK_LAST_PAGE_RE = re.compile(
    r'<[^>]*[?&]page=(\d+)[^>]*>;\s*rel="last"'
)


# =============================================================================
# LOGGING
# =============================================================================

def log(msg):
    """Print a timestamped log message."""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


# =============================================================================
# COMMIT METRICS
# =============================================================================

def fetch_commit_metrics(owner, repo):
    """Return ``{"last_commit_date", "total_commits"}`` or None on failure.

    Both metrics come from a single API call: the commits endpoint with
    ``per_page=1``.  The first entry gives us the date, the ``Link`` header
    gives us the count via the standard pagination trick.

    Edge cases:
      * Empty repo (no commits)            → API returns 409, we return None
      * Repo with exactly one commit       → no Link header, total=1
      * Otherwise                          → total = page number from rel="last"
    """
    data, link = github_get_with_link_header(f"repos/{owner}/{repo}/commits?per_page=1")
    if not data or not isinstance(data, list):
        return None

    # The first commit returned is the most recent on the default branch.
    head = data[0] if data else None
    if not head:
        return None

    commit_obj = head.get("commit") or {}
    committer  = commit_obj.get("committer") or {}
    iso_date   = committer.get("date") or (commit_obj.get("author") or {}).get("date")

    last_commit_date = iso_date[:10] if iso_date else None  # YYYY-MM-DD

    # Total commits via pagination Link header.
    if link:
        match = _LINK_LAST_PAGE_RE.search(link)
        total_commits = int(match.group(1)) if match else len(data)
    else:
        # No Link header → response was a single page → at most len(data) commits.
        total_commits = len(data)

    return {
        "last_commit_date": last_commit_date,
        "total_commits":    total_commits,
    }


# =============================================================================
# PER-PAPER ASSESSMENT
# =============================================================================

def assess_paper(paper):
    """Run the maintenance check for one paper.

    Result dict: ``{title, repo, last_commit_date, total_commits, note}``.
    Missing values are reported as the literal string ``"NOT AVAILABLE"``.
    """
    title = paper.get("title", "")
    url   = paper.get("repo", "") or ""

    base = {
        "title":            title,
        "repo":             url,
        "last_commit_date": NOT_AVAILABLE,
        "total_commits":    NOT_AVAILABLE,
        "note":             "",
    }

    if not url:
        base["note"] = "No repo URL provided"
        return base

    owner, repo = parse_github_repo(url)
    if not owner:
        base["note"] = "Not a GitHub repo URL"
        return base

    log(f"  Querying commits: {owner}/{repo}")
    try:
        metrics = fetch_commit_metrics(owner, repo)
    except RuntimeError as e:
        # Raised by github_helpers on rate-limit (HTTP 403)
        base["note"] = str(e)
        return base
    except Exception as e:
        base["note"] = f"Unexpected error: {e}"
        return base

    if metrics is None:
        # Could be: empty repo, deleted repo, private repo, or transient API failure.
        # Verify the repo exists at all so the note is informative.
        repo_meta = github_get(f"repos/{owner}/{repo}")
        if repo_meta is None:
            base["note"] = "Repo unreachable (deleted, private, or API error)"
        else:
            base["note"] = "Repo exists but has no commits on its default branch"
        return base

    if metrics["last_commit_date"]:
        base["last_commit_date"] = metrics["last_commit_date"]
    if metrics["total_commits"] is not None:
        base["total_commits"] = metrics["total_commits"]
    return base


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
DATA_FONT    = Font(name="Arial", size=10)

EXCEL_HEADERS = [
    "#", "Paper Title", "Repo URL",
    "Date of Last Commit", "Total Commits",
    "Notes",
]
EXCEL_COL_WIDTHS = [5, 50, 55, 22, 16, 50]
NUMERIC_COLS     = {5}      # Total Commits
CENTER_COLS      = {1, 4, 5}


def save_to_excel(results, filename):
    """Write maintenance metrics to a one-sheet Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Post-Publication Maintenance"
    border = thin_border("B8CCE4")

    # Header row + column widths via the shared helper.
    write_header_row(
        ws, EXCEL_HEADERS, EXCEL_COL_WIDTHS,
        fill_hex="BDD7EE", font_color="1F4E79", height=28, border=border,
    )

    # ── Data rows ───────────────────────────────────────────────────────
    for row_idx, r in enumerate(results, start=2):
        row_fill = ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL
        values = [
            row_idx - 1,
            r.get("title", ""),
            r.get("repo", ""),
            r.get("last_commit_date", NOT_AVAILABLE),
            r.get("total_commits", NOT_AVAILABLE),
            r.get("note", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in CENTER_COLS else "left",
                vertical="center",
                wrap_text=True,
            )
            if col_idx in NUMERIC_COLS and isinstance(value, (int, float)):
                cell.number_format = "#,##0"
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "A2"

    Path(filename).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filename)
    log(f"Saved to {filename}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Run post-publication maintenance check on every paper, save Excel."""
    log("\n" + "=" * 80)
    log("Question 5.2.4 — Post-Publication Maintenance")
    log("=" * 80)
    log(f"Processing {len(PAPERS)} papers...\n")

    results = []
    started = time.time()

    for idx, paper in enumerate(PAPERS, 1):
        title = paper.get("title", "")
        log(f"[{idx}/{len(PAPERS)}] {title[:60]}")
        result = assess_paper(paper)
        results.append(result)
        log(f"  => last={result['last_commit_date']}, total={result['total_commits']}")
        log("")
        time.sleep(0.3)

    elapsed = time.time() - started
    log("=" * 80)
    log(f"Done in {elapsed:.1f}s")
    log("=" * 80)

    output_path = _this.parent / "post_publication_maintenance.xlsx"
    save_to_excel(results, str(output_path))


if __name__ == "__main__":
    main()
