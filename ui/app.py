"""Streamlit UI for the paper-reproducibility checkers.

Launch from the repo root (with the venv active):

    streamlit run ui/app.py

Three tabs — 5.1 (code availability), 5.2 (usability/popularity), and the
Beall's predatory-venue check.  Upload a papers JSON (a list, or a single paper
object), pick a check, run it, preview the result, and download the Excel.

Each check runs the existing checker script as a subprocess with the upload
injected via an env var (see ui/runners.py), so the UI and the CLI always
produce identical output.
"""

import json
import os
import re
import sys

import openpyxl
import streamlit as st
from dotenv import load_dotenv

import runners

load_dotenv(runners.REPO_ROOT / ".env")

# In-process validators, for the live "Input data quality" panel (these modules
# are import-safe — no config/prompts name collisions).
sys.path.insert(0, str(runners.RC))           # requirement_checks/
sys.path.insert(0, str(runners._BEALLS))      # bealls_list_check/ (for the venue checks)
try:
    from common import input_quality           # the one shared, field-aware validator
    _VALIDATORS_OK = True
except Exception:
    _VALIDATORS_OK = False

st.set_page_config(page_title="Paper reproducibility checks", page_icon="📄", layout="wide")

# Best-effort: show Streamlit toasts in the top-right (falls back to default).
st.markdown("<style>div[data-testid='stToast']{position:fixed;top:3.5rem;right:1rem;}</style>",
            unsafe_allow_html=True)


def notify(msg, icon=None):
    """Transient popup; never fatal if the Streamlit build lacks st.toast."""
    try:
        st.toast(msg, icon=icon)
    except Exception:
        pass

# ── Example input shown to the user per section ─────────────────────────────
_GH_EXAMPLE = json.dumps(
    [{"title": "Example paper", "repo": "https://github.com/pallets/flask",
      "semanticscholarid": ""}], indent=2)
_S2_EXAMPLE = json.dumps(
    [{"title": "Example paper", "venue": "Journal of X", "year": 2023,
      "externalIds": {"DOI": "10.1234/x"},
      "publicationVenue": {"name": "Journal of X", "type": "journal",
                           "url": "https://www.scirp.org/journal/x",
                           "issn": "1234-5678", "alternate_urls": []},
      "openAccessPdf": {"url": ""}}], indent=2)


def env_sidebar():
    st.sidebar.header("Environment")
    token = bool(os.getenv("GITHUB_TOKEN"))
    st.sidebar.write(("✅" if token else "⚠️") + " `GITHUB_TOKEN` "
                     + ("set" if token else "missing (60 req/hr limit)"))
    st.sidebar.write(f"**LLM provider:** `{os.getenv('LLM_PROVIDER', 'azure')}`")
    st.sidebar.write(f"**Deployment:** `{os.getenv('AZURE_OPENAI_DEPLOYMENT', '(unset)')}`")
    st.sidebar.caption("Keys are read from the repo's `.env` by each checker. "
                       "LLM checks cost tokens; large lists can be slow.")


def get_papers(section_key: str, corpus: bool):
    """Render the upload/paste widget and return a parsed list of papers, or None."""
    example = _S2_EXAMPLE if corpus else _GH_EXAMPLE
    fmt = ("Semantic Scholar **paper records** (with `publicationVenue`)"
           if corpus else "papers with **GitHub links** (`title` + `repo`)")
    st.caption(f"Input format: {fmt}. Upload a JSON **list**, or a **single** paper object.")
    with st.expander("See expected format"):
        st.code(example, language="json")
    mode = st.radio("Provide papers via", ["Upload file", "Paste JSON"],
                    horizontal=True, key=f"mode_{section_key}")
    raw = None
    if mode == "Upload file":
        up = st.file_uploader("papers.json", type=["json"], key=f"file_{section_key}")
        if up is not None:
            raw = up.getvalue().decode("utf-8")
    else:
        raw = st.text_area("Paste JSON here", height=160, key=f"paste_{section_key}",
                           placeholder=example)

    if not raw or not raw.strip():
        return None
    try:
        items = runners.normalize_upload(json.loads(raw))
    except json.JSONDecodeError as e:
        st.error(f"Invalid JSON: {e}")
        return None
    except ValueError as e:
        st.error(str(e))
        return None
    return items


def venue_list_input(label: str, key: str):
    """Optional JSON list of venues for whitelist/blacklist. Returns list or None."""
    raw = st.text_area(label, height=90, key=key,
                       placeholder='[{"name": "IEEE", "domain": "ieee.org"}, "Nature"]')
    if not raw or not raw.strip():
        return None
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        st.error(f"{label}: invalid JSON ({e})")
        return None
    if isinstance(parsed, dict):
        parsed = [parsed]
    if not isinstance(parsed, list):
        st.error(f"{label}: must be a JSON list of venues.")
        return None
    return parsed


def render_result(res: runners.RunResult, checker: runners.Checker):
    if res.ok:
        st.success(f'✅ Done — "{checker.label}" finished. Your Excel is ready below.')
        st.download_button(f"⬇️ Download {res.output_path.name}", data=res.output_bytes,
                           file_name=res.output_path.name, key=f"dl_{checker.id}",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        _preview(res.output_bytes)
    else:
        st.error(f"The checker exited with code {res.returncode} and produced no output file. "
                 "See the run log below.")
    with st.expander("Run log (incl. the input data-quality report)", expanded=not res.ok):
        if res.stdout:
            st.text(res.stdout[-8000:])
        if res.stderr:
            st.caption("stderr")
            st.text(res.stderr[-4000:])


def _preview(xlsx_bytes: bytes, max_rows: int = 40):
    """Show sheet names + the first sheet's first rows as a table."""
    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:                      # never let a preview crash the page
        st.caption(f"(Preview unavailable: {e})")
        return
    st.caption("Sheets: " + ", ".join(wb.sheetnames))
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    data = [dict(zip(headers, [("" if v is None else v) for v in r])) for r in rows[1:max_rows + 1]]
    st.dataframe(data, use_container_width=True)
    if len(rows) > max_rows + 1:
        st.caption(f"(showing first {max_rows} of {len(rows) - 1} rows — download for the full sheet)")


def show_input_quality(items, corpus, key):
    """Persistent panel showing the input-validation (data-quality) check + result.

    This is where the data-quality / Semantic-Scholar check is visible up front:
    it runs the offline validators in-process on the loaded papers and offers the
    report as a downloadable Excel. (Deeper checks — Crossref, repo-liveness —
    run inside the check itself and appear in the run log + the result workbook.)
    """
    if not _VALIDATORS_OK:
        return
    try:
        flagged = input_quality.validate_papers(items)   # one shared validator (offline)
    except Exception as exc:
        st.caption(f"(input validation unavailable: {exc})")
        return
    what = "Semantic Scholar venue metadata" if corpus else "GitHub-link papers"
    icon = "⚠️" if flagged else "✅"
    with st.expander(f"{icon} Input data quality ({what}) — {len(flagged)} of {len(items)} "
                     f"paper(s) flagged", expanded=bool(flagged)):
        st.download_button("⬇️ Download report (.xlsx)",
                           data=runners.build_input_quality_xlsx(items, flagged),
                           file_name="input_data_quality.xlsx", key=f"dqdl_{key}",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if not flagged:
            st.caption("No issues found by the offline checks. (Not a guarantee of correctness; "
                       "deeper checks run during the check itself and show in the run log.)")
        for i, p, flags in flagged[:5]:
            st.markdown(f"**#{i}** {(p.get('title') or '?')[:90]} — " + "; ".join(flags))
        if len(flagged) > 5:
            st.caption(f"…and {len(flagged) - 5} more — download the report above for the full list.")


SECTION_INTRO = {
    "5.1": "Is the paper's **code available and documented**?",
    "5.2": "Is the code **maintained and actually used**?",
    "bealls": "Is the paper's **journal/publisher on Beall's List** of potentially predatory venues?",
}


def _checker_badge(checker, section):
    if checker.needs_llm:
        return "🔑 Uses an LLM — needs API keys in `.env`, and costs tokens."
    if section == "bealls":
        return "⚡ No LLM — matched offline against the vendored Beall's List snapshot."
    return "⚡ No LLM — uses the GitHub API (a `GITHUB_TOKEN` in `.env` is recommended)."


def section_tab(section: str):
    st.caption(SECTION_INTRO[section])
    checkers = runners.checkers_for(section)

    # 1 — choose a check
    st.markdown("**Step 1 · Choose a check**")
    labels = [c.label for c in checkers]
    label = st.selectbox("Check", labels, key=f"sel_{section}", label_visibility="collapsed")
    checker = checkers[labels.index(label)]
    st.markdown(f"➡️ {checker.description}")
    st.caption(_checker_badge(checker, section))
    if checker.note:
        st.info("ℹ️ " + checker.note)

    # 2 — provide the papers
    st.markdown("**Step 2 · Provide the papers**")
    items = get_papers(f"{section}_{checker.id}", corpus=checker.corpus_based)
    if items is not None:
        st.success(f"✅ Loaded {len(items)} paper(s) — ready to run.")
        show_input_quality(items, checker.corpus_based, key=checker.id)

    # 3 — options
    aux_lists = None
    extra_args = None
    env_overrides = None
    with st.expander("⚙️ Options (optional)"):
        if section in ("5.1", "5.2"):
            if st.checkbox(
                    "Check that each GitHub link still works", key=f"live_{checker.id}",
                    help="Before running, sends a quick request to every repository URL and flags "
                         "links that are dead (404) or moved. Slower — it contacts GitHub once per "
                         "paper. (Input is always validated for missing/duplicate/non-GitHub links "
                         "regardless; see the panel above.)"):
                env_overrides = {"CHECK_REPO_LIVENESS": "1"}
            else:
                st.caption("Nothing to configure — input is always validated (see the panel above).")
        if section == "bealls":
            st.markdown("**Scope — whitelist** (optional)")
            wl = venue_list_input("Only check papers from these venues — leave empty to check all",
                                  f"wl_{checker.id}")
            st.markdown("**Extend the list — blacklist** (optional)")
            bl = venue_list_input("Also treat papers from these venues as predatory",
                                  f"bl_{checker.id}")
            aux_lists = {"--whitelist": wl, "--blacklist": bl}
            st.caption('Format for both boxes: a JSON list like '
                       '`[{"name": "IEEE", "domain": "ieee.org"}]` (a plain `"name"` string works too).')
            if st.checkbox(
                    "Double-check each venue against Crossref (more reliable, slower)",
                    key=f"cr_{checker.id}",
                    help="Crossref is the official registry of published articles (looked up by "
                         "each paper's DOI). When on, the journal name and ISSN that Semantic "
                         "Scholar reports are compared against Crossref's official record, and any "
                         "that disagree are flagged in the Data-quality column — catching papers "
                         "where Semantic Scholar has the wrong venue. Off is faster; on contacts "
                         "Crossref once per paper."):
                extra_args = ["--crossref", "all"]

    # run
    st.markdown("**Step 3 · Run**")
    run = st.button("▶ Run check", type="primary", key=f"run_{checker.id}",
                    disabled=items is None,
                    help=None if items is not None else "Add papers in Step 2 first.")
    if not items:
        st.caption("Add papers above to enable the button.")
    if run and items is not None:
        progress = st.progress(0.0, text="Starting…")
        log_line = st.empty()
        step = re.compile(r"\[\s*(\d+)\s*/\s*(\d+)\s*\]")   # checkers print "[i/N]"
        pulse = {"v": 0.0}

        def on_line(line):
            m = step.search(line)
            if m and int(m.group(2)):                # checks that report "[i/N]"
                progress.progress(min(int(m.group(1)) / int(m.group(2)), 1.0),
                                  text=f"{m.group(1)}/{m.group(2)} papers…")
            else:                                    # Beall's has no per-paper counter —
                pulse["v"] = (pulse["v"] + 0.05) % 0.9   # nudge so the bar looks alive
                progress.progress(pulse["v"], text="Working…")
            if line.strip():
                log_line.caption("⏳ " + line[-140:])    # latest activity from the checker

        try:
            # st.spinner shows an animated indicator the whole time, so the user
            # always sees that work is happening even during silent stretches
            # (e.g. while Beall's loads its snapshot/corpus).
            with st.spinner(f'Running "{checker.label}"… '
                            "large lists and LLM checks can take a while."):
                res = runners.run_checker(checker, items, aux_lists=aux_lists,
                                          extra_args=extra_args,
                                          env_overrides=env_overrides, on_line=on_line)
        except Exception as e:                       # surface any launch failure
            progress.empty(); log_line.empty()
            st.session_state[f"res_{checker.id}"] = None
            st.error("Could not start the check:")
            st.exception(e)
            return
        progress.progress(1.0, text="Done")
        log_line.empty()
        st.session_state[f"res_{checker.id}"] = res
        notify("Check complete ✅" if res.ok else "Check failed 🛑",
               icon="✅" if res.ok else "🛑")

    res = st.session_state.get(f"res_{checker.id}")
    if res is not None:
        render_result(res, checker)


st.title("📄 Paper reproducibility & venue checks")
st.caption("Pick a tab, choose a check, add your papers as JSON, and run — each check "
           "produces a downloadable Excel.")
env_sidebar()
with st.sidebar.expander("ℹ️ What the tabs mean"):
    st.markdown(
        "- **5.1 — Code availability:** is the paper's code public and documented?\n"
        "- **5.2 — Usability & popularity:** is that code maintained and actually used?\n"
        "- **Beall's:** is the paper's journal/publisher on Beall's List of *potentially* "
        "predatory venues?\n\n"
        "In each tab: **(1)** choose a check, **(2)** add papers (a JSON list, or a single "
        "paper object), **(3)** Run. The result downloads as an Excel file.")

tab51, tab52, tab_b = st.tabs(["5.1 · Code availability", "5.2 · Usability & popularity",
                               "Beall's · predatory venues"])
with tab51:
    section_tab("5.1")
with tab52:
    section_tab("5.2")
with tab_b:
    section_tab("bealls")
