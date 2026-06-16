"""Adapter layer between the Streamlit UI and the existing checker scripts.

Each checker is a standalone script that reads its papers from
``data.papers_source.load_papers()`` (which honours the ``PAPERS_JSON`` env var)
and writes an Excel file to a fixed location.  Rather than import the checkers
in-process (they reuse module names like ``config``/``prompts`` across folders,
which would collide), we run each as a **subprocess** with the upload injected
via an environment variable, then read back the Excel it produced.

This module is import-safe (no Streamlit) so it can be unit-tested on its own.
"""

import io
import json
import os
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font

REPO_ROOT = Path(__file__).resolve().parent.parent
RC = REPO_ROOT / "requirement_checks"
_Q1 = RC / "5.1.code_availability"
_Q2 = RC / "5.2.practitioner_usability_and_popularity"
_DOC = _Q1 / "5.1.4.code_documentation_quality"
_BEALLS = RC / "bealls_list_check"


@dataclass(frozen=True)
class Checker:
    id: str
    section: str            # "5.1" | "5.2" | "bealls"
    label: str
    script: Path            # script to run as a subprocess
    output: Path            # Excel it writes (we also scan its folder for fresher files)
    needs_llm: bool = False
    corpus_based: bool = False   # Beall's: input is Semantic Scholar records, fed via BEALLS_CORPUS_DIR
    note: str = ""
    description: str = ""        # plain-language "what this check answers", shown in the UI


CHECKERS: List[Checker] = [
    # ── 5.1 Code availability & documentation ───────────────────────────────
    Checker("5.1.3.preprocessing", "5.1", "5.1.3 — Preprocessing / pipeline code (LLM)",
            _Q1 / "5.1.3.pre-processing_&_pipeline_code" / "check_paper_appendix_for_data_preprocessing_code.py",
            _Q1 / "5.1.3.pre-processing_&_pipeline_code" / "results" / "preprocessing_code_results.xlsx",
            needs_llm=True,
            description="Does the repository contain real data-preprocessing / pipeline code "
                        "(not just an inference demo)?"),
    Checker("5.1.4.inline", "5.1", "5.1.4 — Inline comments (LLM)",
            _DOC / "check_github_repo_inline_comments" / "check_github_repo_inline_comments.py",
            _DOC / "check_github_repo_inline_comments" / "results" / "inline_comments_results.xlsx",
            needs_llm=True,
            description="Does the code contain meaningful inline comments?"),
    Checker("5.1.4.installation", "5.1", "5.1.4 — Installation instructions (LLM)",
            _DOC / "check_github_repo_installation_instructions" / "check_github_repo_installation_instructions.py",
            _DOC / "check_github_repo_installation_instructions" / "results" / "installation_instructions_results.xlsx",
            needs_llm=True,
            description="Does the repo explain how to install / set it up?"),
    Checker("5.1.4.usage", "5.1", "5.1.4 — Usage / example commands (LLM)",
            _DOC / "check_github_repo_usage_examples" / "check_github_repo_example_commands.py",
            _DOC / "check_github_repo_usage_examples" / "results" / "usage_examples_results.xlsx",
            needs_llm=True,
            description="Does the repo show example commands / how to run it?"),
    Checker("5.1.4.apidoc", "5.1", "5.1.4 — API documentation (LLM)",
            _DOC / "check_github_repo_api_documentation" / "check_github_repo_api_documentation.py",
            _DOC / "check_github_repo_api_documentation" / "results" / "api_documentation_results.xlsx",
            needs_llm=True,
            description="Does the repo document its functions / classes (API docs)?"),
    Checker("5.1.4.apidoc_norule", "5.1", "5.1.4 — API documentation (rule-based, no LLM)",
            _DOC / "check_github_repo_api_documentation" / "api_documentation_check_no_llm_used.py",
            _DOC / "check_github_repo_api_documentation" / "results" / "api_documentation_no_llm.xlsx",
            description="Same question as above, but answered by scanning for docstrings / API "
                        "docs with rules (no LLM, no tokens)."),
    Checker("5.1.5.license", "5.1", "5.1.5 — Code license (no LLM)",
            _Q1 / "5.1.5.code_license" / "5.1.5.code_license.py",
            _Q1 / "5.1.5.code_license" / "code_license.xlsx",
            description="Does the repo have an open-source license, and which one?"),
    # ── 5.2 Practitioner usability & popularity ─────────────────────────────
    Checker("5.2.2.maintenance", "5.2", "5.2.2 — Maintenance activity indicators (no LLM)",
            _Q2 / "5.2.2.maintenance_activity_indicators" / "5.2.2.maintenance_activity_indicators.py",
            _Q2 / "5.2.2.maintenance_activity_indicators" / "maintenance_indicators.xlsx",
            description="Is the repo actively maintained? (recent commits, contributors, "
                        "releases, archived/stale flags)"),
    Checker("5.2.3.adoption", "5.2", "5.2.3 — Adoption metrics: stars/forks/PyPI (no LLM)",
            _Q2 / "5.2.3.adoption_metrics" / "5.2.3.adoption_metrics.py",
            _Q2 / "5.2.3.adoption_metrics" / "adoption_metrics.xlsx",
            description="How adopted is it? GitHub stars/forks and PyPI monthly downloads."),
    Checker("5.2.4.postpub", "5.2", "5.2.4 — Post-publication maintenance (no LLM)",
            _Q2 / "5.2.4.post_publication_maintenance" / "5.2.4.post_publication_maintenance.py",
            _Q2 / "5.2.4.post_publication_maintenance" / "post_publication_maintenance.xlsx",
            description="Has it been maintained since the paper? (date of last commit, total commits)"),
    # ── Beall's predatory-venue check (corpus = Semantic Scholar records) ────
    Checker("bealls.deterministic", "bealls", "Beall's check — fast (rules only, no LLM)",
            _BEALLS / "bealls_list_check.py",
            _BEALLS / "results" / "bealls_list_results.xlsx",
            corpus_based=True,
            description="Is the paper's journal/publisher on Beall's List of potentially "
                        "predatory venues? Fast, deterministic, no tokens.",
            note="Input is Semantic Scholar paper records (with publicationVenue), not GitHub links."),
    Checker("bealls.llm", "bealls", "Beall's check — thorough (rules + LLM second pass)",
            _BEALLS / "bealls_llm_check.py",
            _BEALLS / "results" / "bealls_llm_results.xlsx",
            needs_llm=True, corpus_based=True,
            description="Same as the fast check, then an LLM re-examines each venue to catch "
                        "ones the rules miss (e.g. alias domains). Uses tokens.",
            note="Input is Semantic Scholar paper records (with publicationVenue), not GitHub links."),
]

CHECKERS_BY_ID = {c.id: c for c in CHECKERS}


def checkers_for(section: str) -> List[Checker]:
    return [c for c in CHECKERS if c.section == section]


@dataclass
class RunResult:
    ok: bool
    returncode: int
    stdout: str
    stderr: str
    output_path: Optional[Path]
    output_bytes: Optional[bytes]


def _newest_xlsx_since(folder: Path, since_ts: float) -> Optional[Path]:
    """Newest .xlsx in *folder* modified at/after *since_ts* (handles the
    timestamped fallback Beall's writes when its default file is locked)."""
    if not folder.is_dir():
        return None
    candidates = [p for p in folder.glob("*.xlsx")
                  if not p.name.startswith("~$") and p.stat().st_mtime >= since_ts - 1]
    return max(candidates, key=lambda p: p.stat().st_mtime, default=None)


def run_checker(checker: Checker, items: list, *, aux_lists: Optional[dict] = None,
                extra_args: Optional[list] = None, env_overrides: Optional[dict] = None,
                on_line=None, timeout: Optional[int] = None) -> RunResult:
    """Run one checker over *items* and return its produced Excel.

    *items* is a list of paper objects (GitHub-link papers for 5.1/5.2; Semantic
    Scholar records for Beall's).  We inject them via a temp file + env var and
    run the checker as a subprocess, inheriting the repo's ``.env`` (the checker
    calls ``load_dotenv()`` itself, so GitHub/LLM keys are picked up).

    *aux_lists* maps a CLI flag to a list written to a temp JSON and passed to
    the checker — used for Beall's ``--whitelist`` / ``--blacklist``; empty
    lists are skipped.  *on_line(line)*, if given, is called for each stdout line
    as it arrives (so the UI can show live progress); otherwise output is just
    captured.
    """
    started = time.time()
    with tempfile.TemporaryDirectory(prefix="bealls_ui_") as tmp:
        tmp = Path(tmp)
        env = os.environ.copy()
        env.update(env_overrides or {})
        if checker.corpus_based:
            # The corpus dir is globbed for *.json, so it must contain ONLY the
            # uploaded corpus — keep it in its own subdir away from the aux files.
            corpus_dir = tmp / "corpus"
            corpus_dir.mkdir()
            (corpus_dir / "uploaded_corpus.json").write_text(
                json.dumps(items, ensure_ascii=False), encoding="utf-8")
            env["BEALLS_CORPUS_DIR"] = str(corpus_dir)
        else:
            papers_file = tmp / "uploaded_papers.json"
            papers_file.write_text(json.dumps(items, ensure_ascii=False), encoding="utf-8")
            env["PAPERS_JSON"] = str(papers_file)

        # ``-u`` (unbuffered) so the child's print()s reach us line-by-line even
        # though stdout is a pipe, not a terminal — otherwise block buffering
        # hides all progress until the process exits (very visible on Beall's,
        # which prints few lines).
        cmd = [sys.executable, "-u", str(checker.script)]
        for flag, entries in (aux_lists or {}).items():
            if not entries:
                continue
            aux_file = tmp / (flag.lstrip("-") + ".json")   # in tmp, never the corpus dir
            aux_file.write_text(json.dumps(entries, ensure_ascii=False), encoding="utf-8")
            cmd += [flag, str(aux_file)]
        cmd += list(extra_args or [])

        if on_line is None:
            proc = subprocess.run(cmd, env=env, cwd=str(REPO_ROOT),
                                  capture_output=True, text=True, timeout=timeout)
            stdout, stderr, returncode = proc.stdout or "", proc.stderr or "", proc.returncode
        else:
            # Stream stdout (stderr merged in, so order is preserved and the pipe
            # can't deadlock) so the caller can show live progress.
            proc = subprocess.Popen(cmd, env=env, cwd=str(REPO_ROOT),
                                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, bufsize=1)
            captured = []
            for line in proc.stdout:
                line = line.rstrip("\n")
                captured.append(line)
                try:
                    on_line(line)
                except Exception:
                    pass                       # a UI hiccup must not kill the run
            proc.wait(timeout=timeout)
            stdout, stderr, returncode = "\n".join(captured), "", proc.returncode

    out_path = checker.output if checker.output.exists() else None
    fresh = _newest_xlsx_since(checker.output.parent, started)
    if fresh is not None:
        out_path = fresh

    output_bytes = out_path.read_bytes() if out_path else None
    return RunResult(
        ok=(returncode == 0 and output_bytes is not None),
        returncode=returncode,
        stdout=stdout,
        stderr=stderr,
        output_path=out_path,
        output_bytes=output_bytes,
    )


def build_input_quality_xlsx(items: list, flagged: list) -> bytes:
    """Build a downloadable Excel of the input data-quality report (all papers).

    *flagged* is the list of ``(index, paper, warnings)`` from
    ``input_quality.validate_papers``; every paper gets a row, with "OK" when it
    has no warnings.
    """
    flags_by_idx = {i: warns for i, _paper, warns in flagged}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input data quality"
    headers = ["#", "Title", "Repo", "Venue", "DOI", "Data-quality warnings"]
    widths = [5, 50, 45, 35, 24, 70]
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = width
    for i, paper in enumerate(items, 1):
        pv = paper.get("publicationVenue") or {}
        warns = flags_by_idx.get(i, [])
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=paper.get("title") or "")
        ws.cell(row=i + 1, column=3, value=paper.get("repo") or "")
        ws.cell(row=i + 1, column=4, value=paper.get("venue") or pv.get("name") or "")
        ws.cell(row=i + 1, column=5, value=(paper.get("externalIds") or {}).get("DOI") or "")
        wcell = ws.cell(row=i + 1, column=6, value="OK" if not warns else "\n".join(warns))
        wcell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def normalize_upload(parsed) -> list:
    """Coerce uploaded/pasted JSON into a list of paper objects.

    Accepts a list of objects or a single object (wrapped). Raises ValueError
    with a friendly message otherwise.
    """
    if isinstance(parsed, dict):
        return [parsed]
    if isinstance(parsed, list):
        if not all(isinstance(x, dict) for x in parsed):
            raise ValueError("JSON list must contain paper objects (dicts).")
        return parsed
    raise ValueError("Upload must be a JSON list of papers or a single paper object.")
