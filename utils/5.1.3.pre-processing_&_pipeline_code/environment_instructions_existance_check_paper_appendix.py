"""
Question 5.1.3 - Pre-processing & Pipeline Code classifier

Root cause of all empty responses:
  GPT-5 is a reasoning model. It spends tokens internally thinking before
  it writes output. max_completion_tokens=768 covered the reasoning budget
  but left nothing for the actual response — so it returned empty silently.
  Fix: raise max_completion_tokens to 16000, which covers both reasoning
  and output for any appendix size.
"""

import os, re, sys, time, urllib.parse, requests
from pathlib import Path
from collections import Counter
from dotenv import load_dotenv

load_dotenv()

sys.path.append(str(Path(__file__).resolve().parent.parent))
from openai_client import client, AZURE_OPENAI_DEPLOYMENT
from papers import PAPERS

PAUSE   = 1.5
ALLOWED = {"SEPARATE_APPENDIX", "MISSING"}

# ---------------------------------------------------------------------------
# Prompts
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """You are classifying ML papers for a literature review.
You will be given the full APPENDIX TEXT of a paper.

SEPARATE_APPENDIX - ONLY if the appendix contains:
- Literal tables titled "Prompt Templates", "Input Examples", "Few-shot Templates", "ASCII Art Prompts", "Jailbreak Prompt Table", etc.
- Pseudocode/Algorithm blocks that show HOW INPUT DATA IS CONSTRUCTED (e.g. Algorithm for building prompts, filtering dataset, assembling probe sets)
- Structured lists or tables of prompt templates used as experimental inputs

MISSING - everything else, including:
- Prose descriptions
- Hyperparameter tables
- Algorithm pseudocode for the METHOD (patching, steering, circuit discovery, SAE training, etc.)
- Feature-example tables, output examples, result tables
- Single-sentence mentions of tokenization/dataset
- References or bibliography

CRITICAL DISTINCTION (never forget):
SEPARATE_APPENDIX = tells you HOW THE INPUT DATA WAS BUILT
MISSING           = tells you HOW THE METHOD/ANALYSIS WORKS

Few-shot examples:

EXAMPLE 1 (SEPARATE_APPENDIX):
Appendix contains "Table 3: Prompt Templates" with exact SYSTEM/USER blocks for GPT-4 classification experiments.

EXAMPLE 2 (SEPARATE_APPENDIX):
Appendix has "ASCII art prompt template tables" or "Table of jailbreak prompts".

EXAMPLE 3 (SEPARATE_APPENDIX):
Appendix contains "Algorithm 1: Dataset Construction" that builds prompts or probe sets.

EXAMPLE 4 (MISSING):
Appendix only says "We use 40 sequences of 300 tokens..." or has a hyperparameter table.

EXAMPLE 5 (MISSING):
Appendix has "Algorithm 1: Activation Patching" or "Table of feature examples".

EXAMPLE 6 (MISSING):
Appendix is just prose or "Dataset: OpenWebText" with no table/pseudocode for construction.

When in doubt: MISSING.
Reply in EXACTLY this format, nothing else:

CLASSIFICATION: <SEPARATE_APPENDIX or MISSING>
EVIDENCE: <exact quote from appendix, max 1 sentence — must contain "Table" or "Algorithm" if SEPARATE_APPENDIX>
REASONING: <one sentence>"""

USER_PROMPT = """Paper title: {title}

=== APPENDIX TEXT ===
{appendix_text}

Classify this paper now."""

FALLBACK_PROMPT = """Paper title: {title}

The appendix could not be retrieved from arXiv.
Use your training knowledge of this paper's appendix.
If you have no reliable knowledge of it, classify as MISSING.

Classify this paper now."""

# ---------------------------------------------------------------------------
# GPT call
# max_completion_tokens=768 was the bug: GPT-5 uses reasoning tokens
# internally before writing output. 768 was exhausted by reasoning alone,
# leaving nothing for the response. 16000 covers both reasoning + output.
# max_tokens throws HTTP 400 on GPT-5 - do NOT add it back.
# ---------------------------------------------------------------------------

def call_gpt(messages):
    total_chars = sum(len(m["content"]) for m in messages)
    print(f"  [GPT] Calling {AZURE_OPENAI_DEPLOYMENT} — "
          f"{total_chars:,} total chars, max_completion_tokens=16000")
    try:
        resp = client.chat.completions.create(
            model=AZURE_OPENAI_DEPLOYMENT,
            messages=messages,
            max_completion_tokens=16000,
        )
        raw = (resp.choices[0].message.content or "").strip()
        if not raw:
            print(f"  [GPT] ERROR: Empty response — "
                  f"check Azure quota or if the deployment supports this token limit")
        else:
            print(f"  [GPT] OK — {len(raw)} chars: {raw[:200]!r}")
        return raw
    except Exception as e:
        print(f"  [GPT] ERROR: {type(e).__name__}: {e}")
        return ""

# ---------------------------------------------------------------------------
# arXiv fetcher — returns the full appendix, no arbitrary cap
# ---------------------------------------------------------------------------

PAPER_SIGNALS = re.compile(
    r"(?:abstract|introduction|related work|references|appendix|theorem|figure|table)\b",
    re.IGNORECASE)

def is_junk(text):
    """True if text is website navigation rather than paper content."""
    junk_phrases = [
        "arXivLabs", "Papers with Code", "Help | Advanced Search",
        "Subscribe to arXiv", "Privacy Policy", "Cornell University",
        "What is ScienceCast", "What is Replicate", "Hugging Face Spaces",
    ]
    junk_count = sum(1 for p in junk_phrases if p in text[:3000])
    paper_hits = len(PAPER_SIGNALS.findall(text[:5000]))
    # A real ML paper after tag stripping is always 20k+ chars.
    # Anything under 10k is a redirect, stub, or nav page — not the actual paper.
    if len(text) < 10_000:
        print(f"  [arXiv] JUNK: only {len(text):,} chars — real ML papers are 20k+, this is a stub or nav page")
        return True
    if junk_count >= 2 and paper_hits < 5:
        print(f"  [arXiv] JUNK: {junk_count} nav phrases, {paper_hits} paper signals")
        return True
    return False

def fetch_appendix(title):
    """Fetch the full appendix from arXiv HTML. Returns text or None."""

    # Step 1: resolve arXiv ID
    print(f"  [arXiv] Searching: {title[:70]}")
    try:
        query = urllib.parse.quote(f'ti:"{title}"')
        resp  = requests.get(
            f"https://export.arxiv.org/api/query?search_query={query}&max_results=1",
            timeout=15, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        m = re.search(r"<id>http://arxiv\.org/abs/([^<\s]+)</id>", resp.text)
        if not m:
            print(f"  [arXiv] ERROR: No paper ID in API response. "
                  f"Snippet: {resp.text[:200]}")
            return None
        arxiv_id = re.sub(r"v\d+$", "", m.group(1).strip())
        print(f"  [arXiv] Resolved ID: {arxiv_id}")
    except Exception as e:
        print(f"  [arXiv] ERROR during ID lookup: {type(e).__name__}: {e}")
        return None

    # Step 2: fetch HTML
    for url in [f"https://ar5iv.labs.arxiv.org/html/{arxiv_id}",
                f"https://arxiv.org/html/{arxiv_id}"]:
        print(f"  [arXiv] Fetching: {url}")
        try:
            resp = requests.get(url, timeout=30,
                                headers={"User-Agent": "Mozilla/5.0"})
            print(f"  [arXiv] HTTP {resp.status_code}, "
                  f"raw HTML: {len(resp.text):,} chars")
            if resp.status_code != 200:
                continue

            # Strip HTML tags
            text = re.sub(r"<script[^>]*>.*?</script>", " ", resp.text,
                          flags=re.DOTALL)
            text = re.sub(r"<style[^>]*>.*?</style>", " ", text,
                          flags=re.DOTALL)
            text = re.sub(r"<[^>]+>", " ", text)
            text = re.sub(r"\s{3,}", "\n\n", text).strip()
            print(f"  [arXiv] After tag stripping: {len(text):,} chars")

            if is_junk(text):
                print(f"  [arXiv] Skipping {url} — junk content")
                continue

            # Find appendix heading
            APPENDIX_RE = re.compile(
              r"(?:^|\n\n)(?:Appendix|Supplementary|Appendices|"
              r"A\.?\s*(?:Prompt|Input|Dataset|Template|Construction|"
              r"Jailbreak|Few-shot|ASCII))",
              re.IGNORECASE)
            m2 = APPENDIX_RE.search(text)
            if m2:
                appendix = text[m2.start():]
                print(f"  [arXiv] Appendix heading at char {m2.start():,} — "
                      f"full appendix: {len(appendix):,} chars")
            else:
                appendix = text[len(text) * 3 // 4:]
                print(f"  [arXiv] WARNING: No appendix heading — "
                      f"using last 25%: {len(appendix):,} chars")

            print(f"  [arXiv] Preview: {appendix[:300]!r}")
            return appendix

        except Exception as e:
            print(f"  [arXiv] ERROR at {url}: {type(e).__name__}: {e}")

    print(f"  [arXiv] ERROR: All fetch attempts failed")
    return None

# ---------------------------------------------------------------------------
# Parse GPT response
# ---------------------------------------------------------------------------

def parse(raw):
    result = {"classification": "MISSING", "evidence": "N/A", "reasoning": "N/A"}
    if not raw:
        return result

    for line in raw.splitlines():
        ls = line.strip()
        if ls.upper().startswith("CLASSIFICATION:"):
            val = ls.split(":", 1)[1].strip().upper()
            if val in ALLOWED:
                result["classification"] = val
        elif ls.upper().startswith("EVIDENCE:"):
            result["evidence"] = ls.split(":", 1)[1].strip()
        elif ls.upper().startswith("REASONING:"):
            result["reasoning"] = ls.split(":", 1)[1].strip()

    # POST-PROCESSING HEURISTIC (kills false positives)
    evidence = result["evidence"].lower()
    if result["classification"] == "SEPARATE_APPENDIX":
        if not any(kw in evidence for kw in ["table", "algorithm", "prompt template", "ascii", "few-shot", "input construction"]):
            print("  [HEURISTIC] Forcing MISSING — evidence not structured prompt table")
            result["classification"] = "MISSING"

    return result

# ---------------------------------------------------------------------------
# Per-paper pipeline
# ---------------------------------------------------------------------------

def process(entry):
    title    = entry["title"]
    appendix = fetch_appendix(title)

    if appendix:
        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": USER_PROMPT.format(
                title=title, appendix_text=appendix)},
        ]
        source = "ARXIV"
    else:
        print(f"  [PIPELINE] No appendix retrieved — "
              f"falling back to GPT training knowledge")
        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": FALLBACK_PROMPT.format(title=title)},
        ]
        source = "KNOWLEDGE"

    raw    = call_gpt(messages)
    result = parse(raw)
    result["source"]            = source
    result["title"]             = title
    result["semanticscholarid"] = entry.get("semanticscholarid", "")
    return result

# ---------------------------------------------------------------------------
# Accuracy
# ---------------------------------------------------------------------------

def accuracy_report(results):
    gt_map = {p["title"]: p["ground_truth"]
              for p in PAPERS if "ground_truth" in p}
    correct, total, wrong, per_label = 0, 0, [], {}
    for r in results:
        gt = gt_map.get(r["title"])
        if not gt:
            continue
        total += 1
        per_label.setdefault(gt, {"correct": 0, "total": 0})
        per_label[gt]["total"] += 1
        if r["classification"] == gt:
            correct += 1
            per_label[gt]["correct"] += 1
        else:
            wrong.append({"title": r["title"],
                          "pred":  r["classification"],
                          "gt":    gt})
    return {"accuracy":  correct / total if total else 0,
            "correct":   correct,
            "total":     total,
            "per_label": per_label,
            "wrong":     wrong}

# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_excel(results, path, acc):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    COLOR = {"SEPARATE_APPENDIX": "FFEB9C", "MISSING": "FCE4D6"}
    LABEL = {"SEPARATE_APPENDIX": "SEPARATE APPENDIX", "MISSING": "MISSING"}
    gt_map    = {p["title"]: p.get("ground_truth", "") for p in PAPERS}
    gt_reason = {p["title"]: p.get("gt_reason",    "") for p in PAPERS}

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["#", "Title", "Prediction", "Ground Truth", "Correct?",
               "Source", "Evidence", "Reasoning", "GT Reason"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = PatternFill("solid", start_color="2F4F6F")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[1].height = 25

    for i, r in enumerate(results, 1):
        row = i + 1
        clf = r.get("classification", "MISSING")
        gt  = gt_map.get(r["title"], "")
        ok  = (clf == gt) if gt else None
        bg  = "F9F9F9" if i % 2 == 0 else "FFFFFF"

        def cell(col, val, bold=False, fg="000000", bg_ov=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Arial", size=10, bold=bold, color=fg)
            c.alignment = Alignment(
                horizontal="center" if col in (1, 3, 4, 5, 6) else "left",
                vertical="top", wrap_text=True)
            c.fill = PatternFill("solid", start_color=bg_ov or bg)

        cell(1, i)
        cell(2, r["title"])
        for col_i, key in [(3, clf), (4, gt)]:
            c = ws.cell(row=row, column=col_i,
                        value=LABEL.get(key, key))
            c.font      = Font(bold=True, name="Arial", size=10)
            c.alignment = Alignment(horizontal="center", vertical="top",
                                    wrap_text=True)
            c.fill      = PatternFill("solid",
                                      start_color=COLOR.get(key, "F2F2F2"))
        if ok is True:
            cell(5, "✓", bold=True, fg="276221", bg_ov="C6EFCE")
        elif ok is False:
            cell(5, "✗", bold=True, fg="9C0006", bg_ov="FFC7CE")
        else:
            cell(5, "—")
        src_color = {"ARXIV":     "BDD7EE",
                     "KNOWLEDGE": "FFE4B5"}.get(r.get("source", ""), bg)
        c6 = ws.cell(row=row, column=6, value=r.get("source", ""))
        c6.font      = Font(name="Arial", size=10, bold=True)
        c6.alignment = Alignment(horizontal="center", vertical="top")
        c6.fill      = PatternFill("solid", start_color=src_color)
        cell(7, r.get("evidence",  ""))
        cell(8, r.get("reasoning", ""))
        cell(9, gt_reason.get(r["title"], ""))
        ws.row_dimensions[row].height = 60

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Overall Accuracy"
    ws2["B1"] = f"{acc['accuracy']:.1%}  ({acc['correct']}/{acc['total']})"
    ws2["A1"].font = ws2["B1"].font = Font(bold=True, name="Arial", size=12)
    for col, h in enumerate(["Label", "Correct", "Total", "Accuracy"], 1):
        ws2.cell(row=3, column=col,
                 value=h).font = Font(bold=True, name="Arial")
    r2 = 4
    for lbl, s in sorted(acc["per_label"].items()):
        a = s["correct"] / s["total"] if s["total"] else 0
        ws2.cell(row=r2, column=1, value=lbl)
        ws2.cell(row=r2, column=2, value=s["correct"])
        ws2.cell(row=r2, column=3, value=s["total"])
        ws2.cell(row=r2, column=4, value=f"{a:.1%}")
        r2 += 1
    r2 += 1
    ws2.cell(row=r2, column=1,
             value="Wrong predictions").font = Font(bold=True)
    r2 += 1
    for w in acc["wrong"]:
        ws2.cell(row=r2, column=1, value=w["title"])
        ws2.cell(row=r2, column=2, value=w["pred"])
        ws2.cell(row=r2, column=3, value=w["gt"])
        r2 += 1
    for col, w in [("A", 10), ("B", 22), ("C", 22), ("D", 10)]:
        ws2.column_dimensions[col].width = w
    for col, w in [("A", 5), ("B", 48), ("C", 22), ("D", 22),
                   ("E", 10), ("F", 12), ("G", 50), ("H", 50), ("I", 50)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"\n  Saved: {path}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    results = []
    for entry in PAPERS:
        title = entry.get("title", "")
        print(f"\n{'='*60}")
        print(f"  {title[:70]}")
        print(f"{'='*60}")
        try:
            result = process(entry)
        except Exception as e:
            print(f"  [FATAL] {type(e).__name__}: {e}")
            result = {
                "title":             title,
                "semanticscholarid": entry.get("semanticscholarid", ""),
                "classification":    "MISSING",
                "evidence":          f"ERROR: {e}",
                "reasoning":         "Exception during processing",
                "source":            "ERROR",
            }
        results.append(result)
        print(f"  -> FINAL: [{result['source']}] {result['classification']}")
        time.sleep(PAUSE)

    print(f"\n{'='*60}\nSUMMARY\n{'='*60}")
    print(f"  Labels:  {dict(Counter(r['classification'] for r in results))}")
    print(f"  Sources: {dict(Counter(r['source']         for r in results))}")
    acc = accuracy_report(results)
    print(f"\n  Accuracy: {acc['accuracy']:.1%} "
          f"({acc['correct']}/{acc['total']})")
    for lbl, s in sorted(acc["per_label"].items()):
        a = s["correct"] / s["total"] if s["total"] else 0
        print(f"    {lbl:25s}: {a:.0%} ({s['correct']}/{s['total']})")
    if acc["wrong"]:
        print(f"\n  Wrong ({len(acc['wrong'])}):")
        for w in acc["wrong"]:
            print(f"    PRED={w['pred']:22s}  GT={w['gt']:22s}  "
                  f"{w['title'][:55]}")
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "results_513.xlsx")
    save_excel(results, out, acc)