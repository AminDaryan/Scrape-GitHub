
import os
import sys
import base64
import requests
from pathlib import Path
from typing import Optional
from dotenv import load_dotenv
# Load environment variables from .env file
load_dotenv()

# Add parent folder to sys.path if needed
sys.path.append(str(Path(__file__).resolve().parent.parent))

from openai_client import client, AZURE_OPENAI_DEPLOYMENT


"""
Analyzes a GitHub repository to determine if data pre-processing code
(tokenization, filtering, prompting wrappers) is included.

Returns one of:
- INCLUDED_IN_REPO
- SEPARATE_APPENDIX
- DESCRIBED_TEXT_ONLY
- MISSING
- NOT_APPLICABLE
"""
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")

# ── GitHub helpers ────────────────────────────────────────────────────────────
 
def github_headers() -> dict:
    h = {"Accept": "application/vnd.github+json"}
    if GITHUB_TOKEN:
        h["Authorization"] = f"Bearer {GITHUB_TOKEN}"
    return h
 
 
def get_repo_tree(owner: str, repo: str) -> list:
    """Return flat file tree for the default branch."""
    url = f"https://api.github.com/repos/{owner}/{repo}/git/trees/HEAD?recursive=1"
    resp = requests.get(url, headers=github_headers(), timeout=30)
    resp.raise_for_status()
    return resp.json().get("tree", [])
 
 
def get_file_content(owner: str, repo: str, path: str) -> Optional[str]:
    """Fetch decoded text content of a single file (skips binaries / large files)."""
    url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"
    resp = requests.get(url, headers=github_headers(), timeout=30)
    if resp.status_code != 200:
        return None
    data = resp.json()
    if data.get("encoding") != "base64":
        return None
    size = data.get("size", 0)
    if size > 100_000:          # skip files > 100 KB
        return None
    try:
        return base64.b64decode(data["content"]).decode("utf-8", errors="replace")
    except Exception:
        return None
 
 
def get_readme(owner: str, repo: str) -> str:
    """Return README content (any extension)."""
    url = f"https://api.github.com/repos/{owner}/{repo}/readme"
    resp = requests.get(url, headers=github_headers(), timeout=30)
    if resp.status_code != 200:
        return ""
    data = resp.json()
    try:
        return base64.b64decode(data["content"]).decode("utf-8", errors="replace")
    except Exception:
        return ""
 
 
# ── Repo scanning ─────────────────────────────────────────────────────────────
 
PREPROCESSING_KEYWORDS = [
    "tokeniz", "tokenise", "filter", "prompt", "preprocess", "pre-process",
    "pre_process", "data_processing", "data-processing", "wrapper",
    "clean", "normalize", "normalise", "truncat", "pad", "mask",
    "encode", "decode", "vocab", "bpe", "sentencepiece",
]
 
CODE_EXTENSIONS = {
    ".py", ".ipynb", ".r", ".R", ".jl", ".sh", ".js", ".ts",
    ".rb", ".java", ".scala", ".cpp", ".c", ".go",
}
 
APPENDIX_PATTERNS = ["appendix", "supplement", "supp_", "supp-", "_supp"]
 
 
def classify_files(tree: list) -> dict:
    """
    Scan the file tree and return a summary dict with:
      - preprocessing_scripts : list of paths that look like preprocessing code
      - appendix_scripts       : preprocessing scripts that live in appendix/supplement dirs
      - all_paths              : full list of blob paths
    """
    preprocessing_scripts = []
    appendix_scripts = []
 
    for item in tree:
        if item.get("type") != "blob":
            continue
        path: str = item["path"]
        _, ext = os.path.splitext(path.lower())
 
        if ext not in CODE_EXTENSIONS:
            continue
 
        path_lower = path.lower()
        in_appendix = any(p in path_lower for p in APPENDIX_PATTERNS)
        has_keyword = any(kw in path_lower for kw in PREPROCESSING_KEYWORDS)
 
        if has_keyword:
            if in_appendix:
                appendix_scripts.append(path)
            else:
                preprocessing_scripts.append(path)
 
    return {
        "preprocessing_scripts": preprocessing_scripts,
        "appendix_scripts": appendix_scripts,
        "all_paths": [i["path"] for i in tree if i.get("type") == "blob"],
    }
 
 
def build_context(owner: str, repo: str, scan: dict) -> str:
    """Build a text context to feed the LLM."""
    lines = [f"Repository: {owner}/{repo}", ""]
 
    # README
    readme = get_readme(owner, repo)
    if readme:
        lines.append("=== README (first 3000 chars) ===")
        lines.append(readme[:3000])
        lines.append("")
 
    # Relevant script snippets
    relevant = scan["preprocessing_scripts"] + scan["appendix_scripts"]
    for path in relevant[:5]:     # cap at 5 files to stay within token budget
        content = get_file_content(owner, repo, path)
        if content:
            lines.append(f"=== {path} (first 2000 chars) ===")
            lines.append(content[:2000])
            lines.append("")
 
    # Full file listing (truncated)
    lines.append("=== File tree (first 200 entries) ===")
    lines.extend(scan["all_paths"][:200])
 
    return "\n".join(lines)
 
 
# ── AzureOpenAI classification ────────────────────────────────────────────────
 
SYSTEM_PROMPT = """You are an expert code reviewer. Your task is to determine whether a GitHub
repository includes code for data pre-processing (tokenization, filtering, prompting wrappers).
 
You MUST respond in exactly this format and nothing else:
CLASSIFICATION: <one of the values below>
LOCATION: <file path(s) or description of where it was found, or "N/A" if not applicable>
 
Allowed classification values:
INCLUDED_IN_REPO
SEPARATE_APPENDIX
DESCRIBED_TEXT_ONLY
MISSING
NOT_APPLICABLE
 
Definitions:
- INCLUDED_IN_REPO    : Pre-processing scripts are present as part of the main codebase.
- SEPARATE_APPENDIX   : Pre-processing code is provided only in a clearly labelled appendix or supplement folder/file.
- DESCRIBED_TEXT_ONLY : The README or documentation describes the pre-processing steps in text, but no script is provided.
- MISSING             : No pre-processing code or description is present at all.
- NOT_APPLICABLE      : The repository does not involve any NLP / ML pipeline where pre-processing would be relevant.
 
Example response:
CLASSIFICATION: INCLUDED_IN_REPO
LOCATION: src/data/tokenizer.py, scripts/preprocess.py
"""
 
USER_PROMPT_TEMPLATE = """Below is information collected from the GitHub repository.
Based on this information, classify the repository and provide the location of the pre-processing code.
 
{context}
 
Respond in the required format:"""
 
 
def classify_with_llm(context: str) -> dict:
    response = client.chat.completions.create(
        model=AZURE_OPENAI_DEPLOYMENT,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": USER_PROMPT_TEMPLATE.format(context=context)},
        ],
        max_completion_tokens=5000,
    )
 
    if not response.choices:
        print(f"  [WARN] Empty choices. Full response: {response}")
        return {"classification": "MISSING", "location": "N/A"}
 
    choice = response.choices[0]
    if choice.message.content is None:
        print(f"  [WARN] content=None. finish_reason={choice.finish_reason}")
        return {"classification": "MISSING", "location": "N/A"}
 
    raw = choice.message.content.strip()
 
    # Parse CLASSIFICATION and LOCATION lines
    classification = "MISSING"
    location = "N/A"
    allowed = {"INCLUDED_IN_REPO", "SEPARATE_APPENDIX", "DESCRIBED_TEXT_ONLY", "MISSING", "NOT_APPLICABLE"}
 
    for line in raw.splitlines():
        line = line.strip()
        if line.upper().startswith("CLASSIFICATION:"):
            value = line.split(":", 1)[1].strip().upper()
            if value in allowed:
                classification = value
        elif line.upper().startswith("LOCATION:"):
            location = line.split(":", 1)[1].strip()
 
    # Fallback: scan full response if parsing failed
    if classification == "MISSING":
        for value in allowed:
            if value in raw.upper():
                classification = value
                break
 
    return {"classification": classification, "location": location}
 
 
# ── Main entry point ──────────────────────────────────────────────────────────
 
def analyze_repo(repo_url: str) -> str:
    """
    Main function.
 
    Parameters
    ----------
    repo_url : str
        Full GitHub URL, e.g. "https://github.com/owner/repo"
 
    Returns
    -------
    str
        One of: INCLUDED_IN_REPO | SEPARATE_APPENDIX | DESCRIBED_TEXT_ONLY | MISSING | NOT_APPLICABLE
    """
    # Parse owner/repo from URL
    parts = repo_url.rstrip("/").split("/")
    if len(parts) < 2:
        raise ValueError(f"Cannot parse owner/repo from URL: {repo_url}")
    owner, repo = parts[-2], parts[-1]
 
    print(f"Fetching file tree for {owner}/{repo} …")
    tree = get_repo_tree(owner, repo)
    print(f"  {len(tree)} entries found.")
 
    scan = classify_files(tree)
    print(f"  Pre-processing scripts (main):    {len(scan['preprocessing_scripts'])}")
    print(f"  Pre-processing scripts (appendix): {len(scan['appendix_scripts'])}")
 
    context = build_context(owner, repo, scan)
    print("Calling AzureOpenAI for classification …")
    result = classify_with_llm(context)
    return result  # dict with "classification" and "location"
 
 
# ── Repo list ─────────────────────────────────────────────────────────────────

REPOS = [
    # ── INCLUDED_IN_REPO (clear cases) ───────────────────────────────────────
    "https://github.com/huggingface/transformers",          # tokenizers throughout src/
    "https://github.com/openai/openai-cookbook",            # preprocessing notebooks
    "https://github.com/JohnSnowLabs/spark-nlp",            # full NLP pipeline library
    "https://github.com/google/sentencepiece",              # BPE/unigram tokenizer
    "https://github.com/nltk/nltk",                         # classic NLP toolkit
    "https://github.com/explosion/spaCy",                   # industrial NLP, tokenizers
    "https://github.com/facebookresearch/fairseq",          # seq2seq with preprocessing scripts
    "https://github.com/stanfordnlp/stanza",                # Stanford NLP pipeline
    "https://github.com/huggingface/datasets",              # data processing + filtering
    "https://github.com/EleutherAI/gpt-neox",               # preprocessing in tools/
 
    # ── INCLUDED_IN_REPO (tricky — preprocessing hidden in utils/helpers) ────
    "https://github.com/karpathy/nanoGPT",                  # data.py does tokenization
    "https://github.com/pytorch/fairseq",                   # preprocessing buried in fairseq/data/
    "https://github.com/facebookresearch/llama",            # tokenizer.py in root
    "https://github.com/openai/gpt-2",                     # encoder.py + encode.py scripts
    "https://github.com/google-research/bert",              # tokenization.py in root
    "https://github.com/microsoft/DeepSpeed",               # data utils in deepspeed/runtime/
    "https://github.com/allenai/allennlp",                  # readers + tokenizers in allennlp/data/
    "https://github.com/flairNLP/flair",                    # tokenization in flair/tokenization.py
    "https://github.com/NVIDIA/Megatron-LM",                # tools/preprocess_data.py
    "https://github.com/huggingface/trl",                   # reward model data processing
 
    # ── SEPARATE_APPENDIX (preprocessing isolated in supplement/appendix) ───
    "https://github.com/facebookresearch/ParlAI",           # data folder separated from core
    "https://github.com/google-research/text-to-text-transfer-transformer",  # data/ appendix style
    "https://github.com/sebastianruder/NLP-progress",       # data scripts in separate folder
    "https://github.com/Unbabel/OpenKiwi",                  # preprocessing in separate kiwi/data/
    "https://github.com/rsennrich/subword-nmt",             # BPE scripts as standalone tools
 
    # ── DESCRIBED_TEXT_ONLY (README describes steps but no script provided) ──
    "https://github.com/chiphuyen/stanford-tensorflow-tutorials",  # instructions only
    "https://github.com/graykode/nlp-tutorial",             # README describes tokenization, minimal code
    "https://github.com/keon/awesome-nlp",                  # curated list, no code
    "https://github.com/oxford-cs-deepnlp-2017/lectures",  # lecture repo, text descriptions only
    "https://github.com/mihail911/nlp-library",             # reading list, no preprocessing code
 
    # ── DESCRIBED_TEXT_ONLY (tricky — has code but preprocessing only in docs) ─
    "https://github.com/minimaxir/gpt-2-simple",            # README explains encoding but delegates to OpenAI's tokenizer
    "https://github.com/jessevig/bertviz",                  # visualization tool, preprocessing described not implemented
    "https://github.com/huggingface/accelerate",            # training framework; preprocessing described as user responsibility
    "https://github.com/Lightning-AI/pytorch-lightning",    # framework only, preprocessing in docs/examples
    "https://github.com/openai/CLIP",                       # tokenizer imported externally, only described in README
 
    # ── MISSING (no preprocessing code or description at all) ────────────────
    "https://github.com/aymericdamien/TensorFlow-Examples", # generic TF examples, no NLP preprocessing
    "https://github.com/trekhleb/javascript-algorithms",    # algorithms only, no ML preprocessing
    "https://github.com/donnemartin/system-design-primer", # system design, no NLP at all
    "https://github.com/public-apis/public-apis",           # API list, no code
    "https://github.com/sindresorhus/awesome",              # meta-list, no code
 
    # ── NOT_APPLICABLE (no NLP/ML pipeline at all) ───────────────────────────
    "https://github.com/torvalds/linux",                    # OS kernel
    "https://github.com/django/django",                     # web framework
    "https://github.com/vuejs/vue",                         # frontend JS framework
    "https://github.com/expressjs/express",                 # Node.js web framework
    "https://github.com/redis/redis",                       # in-memory database
 
    # ── TRICKY edge cases (ambiguous / hard to detect) ───────────────────────
    "https://github.com/openai/whisper",                    # audio model — has tokenizer but looks like audio repo
    "https://github.com/facebookresearch/metaseq",          # preprocessing exists but deeply nested
    "https://github.com/EleutherAI/lm-evaluation-harness", # eval only — but has prompt templates that look like preprocessing
    "https://github.com/BerriAI/litellm",                  # LLM wrapper — prompt wrapping could be mistaken for preprocessing
    "https://github.com/ggerganov/llama.cpp",              # C++ inference — tokenizer exists but no Python preprocessing
]
 
 
# ── Excel export ──────────────────────────────────────────────────────────────
 

def save_to_excel(results: dict, path: str = "results.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
 
    STATUS_COLORS = {
        "INCLUDED_IN_REPO":    "C6EFCE",  # green
        "SEPARATE_APPENDIX":   "FFEB9C",  # yellow
        "DESCRIBED_TEXT_ONLY": "BDD7EE",  # blue
        "MISSING":             "FCE4D6",  # orange
        "NOT_APPLICABLE":      "E2EFDA",  # light green
    }
 
    DISPLAY_LABELS = {
        "INCLUDED_IN_REPO":    "INCLUDED IN REPO (Part of the main codebase)",
        "SEPARATE_APPENDIX":   "SEPARATE APPENDIX (Provided in appendix/supplement)",
        "DESCRIBED_TEXT_ONLY": "DESCRIBED TEXT ONLY (Text description but no script)",
        "MISSING":             "MISSING (Not provided)",
        "NOT_APPLICABLE":      "NOT APPLICABLE",
    }
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
 
    # Header
    headers = ["#", "Repository", "Classification", "Location"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color="2F4F6F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
 
    ws.row_dimensions[1].height = 25
 
    # Data rows
    for i, (url, result) in enumerate(results.items(), 1):
        row = i + 1
        if isinstance(result, dict):
            classification = result["classification"]
            location = result["location"]
        else:
            classification = result  # ERROR string
            location = "N/A"
 
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="top")
        
        # Repo as hyperlink
        repo_cell = ws.cell(row=row, column=2, value=url)
        repo_cell.hyperlink = url
        repo_cell.font = Font(color="0563C1", underline="single", name="Arial", size=10)
        repo_cell.alignment = Alignment(vertical="top", wrap_text=True)
 
        # Classification with color
        display = DISPLAY_LABELS.get(classification, classification)
        cls_cell = ws.cell(row=row, column=3, value=display)
        cls_cell.font = Font(bold=True, name="Arial", size=10)
        cls_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        color = STATUS_COLORS.get(classification, "F2F2F2")
        cls_cell.fill = PatternFill("solid", start_color=color)
 
        # Location
        loc_cell = ws.cell(row=row, column=4, value=location)
        loc_cell.font = Font(name="Arial", size=10)
        loc_cell.alignment = Alignment(vertical="top", wrap_text=True)
 
        # Alternating row background for non-classified columns
        bg = "F9F9F9" if i % 2 == 0 else "FFFFFF"
        for col in [1, 2, 4]:
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color=bg)
 
        ws.row_dimensions[row].height = 40
 
    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 60
 
    # Freeze header row
    ws.freeze_panes = "A2"
 
    wb.save(path)
    print(f"\nResults saved to: {path}")
 
 
# ── Main ──────────────────────────────────────────────────────────────────────
 
if __name__ == "__main__":
    results = {}
    for repo_url in REPOS:
        print(f"\n{'='*60}")
        print(f"Analyzing: {repo_url}")
        print('='*60)
        try:
            answer = analyze_repo(repo_url)
            results[repo_url] = answer
        except Exception as e:
            results[repo_url] = {"classification": f"ERROR: {e}", "location": "N/A"}
 
    print(f"\n{'='*60}")
    print("SUMMARY")
    print('='*60)
    for url, result in results.items():
        if isinstance(result, dict):
            print(f"{url}")
            print(f"  → {result['classification']}")
            print(f"  → Location: {result['location']}\n")
        else:
            print(f"{url}\n  → {result}\n")
 
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results.xlsx")
    save_to_excel(results, output_path)